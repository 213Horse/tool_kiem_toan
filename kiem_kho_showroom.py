#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ứng dụng Kiểm Kho - Quét mã vạch để kiểm tra tồn kho thực tế
Chạy trên Windows và macOS
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
# Lazy import pandas - chỉ import khi cần thiết để tăng tốc độ khởi động
# import pandas as pd  # Đã chuyển sang lazy import
import os
from pathlib import Path
import sys
import json
import shutil
import time
import traceback
import base64
import signal
import atexit

class KiemKhoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Kiểm Kho Showroom - Quét Mã Vạch")
        # Tăng chiều cao để hiển thị đủ tất cả các phần tử
        self.root.geometry("1200x800")
        # Đặt kích thước tối thiểu để đảm bảo ô nhập ISBN luôn hiển thị trên màn hình nhỏ
        self.root.minsize(800, 600)  # Tối thiểu 800x600 để đảm bảo có đủ không gian
        # Màu nền nhẹ nhàng hơn
        self.root.configure(bg='#F5F5F5')
        
        # Import pandas ngay sau khi UI được tạo (để có thể dùng trong toàn bộ class)
        # Import sớm hơn một chút để tránh lỗi "name 'pd' is not defined"
        try:
            import pandas as pd
            self.pd = pd  # Lưu vào instance để dùng trong các method khác
        except ImportError:
            self.pd = None
            # Sẽ báo lỗi khi load_data được gọi
        
        # Biến lưu trữ dữ liệu
        self.df = None
        self.current_box_data = None
        self.current_box_number = None
        self.scanned_items = {}  # Lưu các item đã quét: {isbn: {tua, ton_thuc_te, so_thung, ton_trong_thung, ghi_chu}}
        self.edit_entry = None  # Entry widget để chỉnh sửa trực tiếp
        self.so_thung_original_value = ''  # Lưu giá trị số thùng ban đầu để chặn sửa khi đã có dữ liệu quét
        self.editing_item = None  # Item đang được chỉnh sửa
        self.error_highlights = {}  # Lưu các highlight widgets: {item_id: [entry1, entry2]}
        self.is_processing_edit = False  # Flag để tránh xử lý edit 2 lần
        self.template_file_path = None  # Đường dẫn CHỌN ĐƯỜNG DẪN FILE TỔNG HỢP MẶC ĐỊNH
        self.auto_save_folder = None  # Thư mục tự động lưu file Excel 2 (Kiemkecuoinam)
        self.config_folder = None  # Thư mục lưu file config (do người dùng chọn)
        self.config_file = self.get_config_file_path()  # Đường dẫn file config
        self.tong_hop_data = []  # Lưu tổng hợp các data đã kiểm kê
        self.notebook = None  # Notebook widget để chứa các tab
        self.tong_hop_tree = None  # Treeview trong tab Tổng hợp
        self.so_tua_da_quet_var = None  # Biến để hiển thị số tựa đã quét
        self.tong_hop_edit_entry = None  # Entry widget để chỉnh sửa trong tab Tổng hợp
        self.tong_hop_editing_item = None  # Item đang được chỉnh sửa trong tab Tổng hợp
        self.tong_hop_editing_column = None  # Cột đang được chỉnh sửa trong tab Tổng hợp
        self._tong_hop_finish_scheduled = None  # ID của scheduled finish_tong_hop_edit call
        self.is_processing_tong_hop_edit = False  # Flag để tránh xử lý 2 lần
        
        # Lưu đường dẫn backup file ngay khi khởi tạo để tránh lỗi __file__ không được định nghĩa
        try:
            self._backup_file_path = self._get_backup_file_path_init()
        except Exception:
            # Fallback: sử dụng thư mục hiện tại
            self._backup_file_path = Path.cwd() / "kiem_kho_showroom_backup.json"
        
        # Load cấu hình từ file (nếu có)
        saved_config = self.load_config()
        need_setup = False
        
        if saved_config:
            template_path = saved_config.get('template_file_path')
            auto_folder = saved_config.get('auto_save_folder')
            config_folder = saved_config.get('config_folder')
            
            # Kiểm tra cả 3 đường dẫn có tồn tại và hợp lệ không
            if template_path and auto_folder and config_folder:
                # Windows-safe: Normalize paths và kiểm tra tồn tại
                try:
                    template_path_normalized = str(Path(template_path).resolve())
                    auto_folder_normalized = str(Path(auto_folder).resolve())
                    config_folder_normalized = str(Path(config_folder).resolve())
                    
                    # Kiểm tra tồn tại với normalized paths
                    if (os.path.isfile(template_path_normalized) and 
                        os.path.isdir(auto_folder_normalized) and 
                        os.path.isdir(config_folder_normalized)):
                        # Cả 3 đường dẫn đều hợp lệ - không cần setup lại
                        self.template_file_path = template_path_normalized
                        self.auto_save_folder = auto_folder_normalized
                        self.config_folder = config_folder_normalized
                        # Cập nhật config_file để dùng đúng thư mục
                        self.config_file = Path(self.config_folder) / "kiem_kho_showroom_config.json"
                    else:
                        # Một hoặc nhiều đường dẫn không tồn tại - cần setup lại
                        need_setup = True
                except Exception as e:
                    # Nếu có lỗi khi normalize (ví dụ: đường dẫn không tồn tại), cần setup lại
                    print(f"Lỗi khi kiểm tra đường dẫn: {str(e)}")
                    need_setup = True
            else:
                # Thiếu một hoặc nhiều đường dẫn - cần setup lại
                need_setup = True
        else:
            # Không có cấu hình - cần setup lần đầu
            need_setup = True
        
        # Nếu cần cấu hình, hiển thị dialog (chỉ lần đầu hoặc khi config không hợp lệ)
        if need_setup:
            # Đảm bảo root window được hiển thị trước khi tạo dialog
            self.root.deiconify()
            self.root.update()
            self.setup_paths()
            
            # Sau khi setup xong, kiểm tra lại xem có cấu hình hợp lệ không (cả 3 đường dẫn)
            if not self.template_file_path or not self.auto_save_folder or not self.config_folder:
                # Nếu vẫn không có cấu hình hợp lệ, đóng app
                self.root.quit()
                return
        
        # Tạo giao diện TRƯỚC để hiển thị nhanh hơn (tối ưu tốc độ khởi động)
        self.create_ui()
        
        # Tự động điều chỉnh kích thước cửa sổ để hiển thị đủ tất cả các phần tử
        self.root.update_idletasks()
        # Lấy chiều cao yêu cầu của cửa sổ
        req_height = self.root.winfo_reqheight()
        req_width = self.root.winfo_reqwidth()
        # Đảm bảo chiều cao đủ để hiển thị tất cả (thêm một chút padding)
        current_height = self.root.winfo_height()
        if req_height > current_height:
            # Điều chỉnh geometry để hiển thị đủ
            self.root.geometry(f"{req_width}x{req_height + 50}")
        
        # Bind Enter key để hỗ trợ quét mã vạch
        self.root.bind('<Return>', self.on_enter_pressed)
        
        # Bind sự kiện đóng cửa sổ để kiểm tra dữ liệu chưa lưu
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Kiểm tra và khôi phục dữ liệu backup nếu có
        self.check_and_restore_backup()
        
        # Load dữ liệu SAU KHI UI đã hiển thị (defer loading để tăng tốc độ khởi động)
        # Sử dụng after() để load dữ liệu sau khi UI đã render xong (100ms delay)
        self.root.after(100, self.load_data_deferred)
        
        # Bắt đầu auto-save định kỳ (mỗi 30 giây)
        self.start_auto_save()
        
        # Đăng ký xử lý signal để lưu backup khi shutdown (cúp điện, tắt máy)
        self.setup_signal_handlers()
    
    def get_config_location_file(self):
        """Lấy đường dẫn file pointer trỏ đến vị trí config thực sự"""
        if getattr(sys, 'frozen', False):
            # Chạy từ executable - lưu file pointer cùng thư mục với exe
            exe_dir = Path(sys.executable).parent
            return exe_dir / ".set_up"
        else:
            # Chạy từ source code
            return Path(__file__).parent / ".set_up"
    
    def _set_file_hidden(self, file_path):
        """Đặt thuộc tính hidden/system cho file trên Windows để chặn người dùng mở"""
        if sys.platform == 'win32':
            try:
                import ctypes
                # Đặt thuộc tính hidden và system
                # FILE_ATTRIBUTE_HIDDEN = 2
                # FILE_ATTRIBUTE_SYSTEM = 4
                # FILE_ATTRIBUTE_HIDDEN | FILE_ATTRIBUTE_SYSTEM = 6
                ctypes.windll.kernel32.SetFileAttributesW(str(file_path), 6)
            except Exception as e:
                # Nếu không thể đặt thuộc tính, không quan trọng lắm
                pass
    
    def _encode_path(self, path_str):
        """Mã hóa đường dẫn để lưu vào file pointer"""
        try:
            # Sử dụng base64 với một key đơn giản để mã hóa
            key = "KiemKho2025SecretKey"
            # XOR với key (đơn giản nhưng hiệu quả)
            encoded_bytes = []
            key_bytes = key.encode('utf-8')
            path_bytes = path_str.encode('utf-8')
            for i, byte in enumerate(path_bytes):
                encoded_bytes.append(byte ^ key_bytes[i % len(key_bytes)])
            # Encode base64 để có thể lưu dạng text
            encoded = base64.b64encode(bytes(encoded_bytes)).decode('utf-8')
            return encoded
        except Exception as e:
            print(f"Lỗi khi mã hóa đường dẫn: {str(e)}")
            return None
    
    def _decode_path(self, encoded_str):
        """Giải mã đường dẫn từ file pointer"""
        try:
            # Decode base64
            encoded_bytes = base64.b64decode(encoded_str.encode('utf-8'))
            # XOR với key để giải mã
            key = "KiemKho2025SecretKey"
            decoded_bytes = []
            key_bytes = key.encode('utf-8')
            for i, byte in enumerate(encoded_bytes):
                decoded_bytes.append(byte ^ key_bytes[i % len(key_bytes)])
            # Decode về string
            decoded = bytes(decoded_bytes).decode('utf-8')
            return decoded
        except Exception as e:
            print(f"Lỗi khi giải mã đường dẫn: {str(e)}")
            return None
    
    def get_config_file_path(self):
        """Lấy đường dẫn file config - đọc từ file pointer hoặc tìm ở nhiều vị trí"""
        # Nếu đã có config_folder do người dùng chọn, dùng nó (ưu tiên cao nhất)
        if self.config_folder:
            return Path(self.config_folder) / "kiem_kho_showroom_config.json"
        
        # Ưu tiên 1: Đọc từ file pointer (trỏ đến vị trí config thực sự)
        location_file = self.get_config_location_file()
        if location_file.exists():
            try:
                with open(location_file, 'r', encoding='utf-8') as f:
                    encoded_path = f.read().strip()
                    if encoded_path:
                        # Giải mã đường dẫn
                        config_path = self._decode_path(encoded_path)
                        if config_path:
                            config_file_path = Path(config_path)
                            if config_file_path.exists():
                                # Đọc config_folder từ file config này
                                try:
                                    with open(config_file_path, 'r', encoding='utf-8') as cf:
                                        config = json.load(cf)
                                        if 'config_folder' in config and config['config_folder']:
                                            config_folder_path = Path(config['config_folder'])
                                            if config_folder_path.exists():
                                                self.config_folder = str(config_folder_path.resolve())
                                                return Path(self.config_folder) / "kiem_kho_showroom_config.json"
                                        # Nếu không có config_folder, dùng file này luôn
                                        return config_file_path
                                except:
                                    return config_file_path
            except Exception as e:
                print(f"Lỗi khi đọc file pointer: {str(e)}")
        
        # Ưu tiên 2: Tìm file config ở nhiều vị trí để lấy config_folder
        search_locations = []
        
        if getattr(sys, 'frozen', False):
            # Chạy từ executable
            exe_dir = Path(sys.executable).parent
            search_locations.extend([
                exe_dir / "kiem_kho_showroom_config.json",
                exe_dir.parent / "kiem_kho_showroom_config.json",
            ])
            
            user_home = Path.home()
            search_locations.extend([
                user_home / "Desktop" / "kiem_kho_showroom_config.json",
                user_home / "Documents" / "kiem_kho_showroom_config.json",
                user_home / "kiem_kho_showroom_config.json",
            ])
        else:
            # Chạy từ source code
            search_locations.append(Path(__file__).parent / "kiem_kho_showroom_config.json")
        
        # Tìm file config ở các vị trí và đọc config_folder từ đó
        for config_file_path in search_locations:
            if config_file_path.exists():
                try:
                    with open(config_file_path, 'r', encoding='utf-8') as f:
                        old_config = json.load(f)
                        # Nếu file này có config_folder, dùng nó
                        if 'config_folder' in old_config and old_config['config_folder']:
                            config_folder_path = Path(old_config['config_folder'])
                            if config_folder_path.exists() and config_folder_path.is_dir():
                                self.config_folder = str(config_folder_path.resolve())
                                # Trả về đường dẫn file config trong config_folder (vị trí thực sự)
                                return Path(self.config_folder) / "kiem_kho_showroom_config.json"
                        # Nếu không có config_folder nhưng có đầy đủ thông tin, dùng file này
                        elif 'template_file_path' in old_config and 'auto_save_folder' in old_config:
                            # File này có thể là file config hợp lệ, nhưng chưa có config_folder
                            # Trả về file này để load
                            return config_file_path
                except Exception as e:
                    # Bỏ qua lỗi và tiếp tục tìm
                    continue
        
        # Nếu không tìm thấy, dùng vị trí mặc định
        if getattr(sys, 'frozen', False):
            return Path(sys.executable).parent / "kiem_kho_showroom_config.json"
        else:
            return Path(__file__).parent / "kiem_kho_showroom_config.json"
    
    def load_config(self):
        """Load cấu hình từ file - Windows-safe, tìm ở nhiều vị trí"""
        try:
            # Danh sách các vị trí tìm file config (theo thứ tự ưu tiên)
            search_locations = []
            
            # Ưu tiên 1: Đọc từ file pointer (trỏ đến vị trí config thực sự)
            location_file = self.get_config_location_file()
            if location_file.exists():
                try:
                    with open(location_file, 'r', encoding='utf-8') as f:
                        encoded_path = f.read().strip()
                        if encoded_path:
                            # Giải mã đường dẫn
                            config_path = self._decode_path(encoded_path)
                            if config_path:
                                config_file_path = Path(config_path)
                                if config_file_path.exists():
                                    search_locations.insert(0, config_file_path)  # Ưu tiên cao nhất
                except Exception as e:
                    print(f"Lỗi khi đọc file pointer: {str(e)}")
            
            # Ưu tiên 2: File config trong config_folder (nếu đã biết)
            if self.config_folder:
                search_locations.append(Path(self.config_folder) / "kiem_kho_showroom_config.json")
            
            # Ưu tiên 3: File config hiện tại
            if self.config_file:
                search_locations.append(self.config_file)
            
            # Ưu tiên 3: Tìm ở các vị trí khác
            if getattr(sys, 'frozen', False):
                exe_dir = Path(sys.executable).parent
                search_locations.extend([
                    exe_dir / "kiem_kho_showroom_config.json",
                    exe_dir.parent / "kiem_kho_showroom_config.json",
                ])
                
                user_home = Path.home()
                search_locations.extend([
                    user_home / "Desktop" / "kiem_kho_showroom_config.json",
                    user_home / "Documents" / "kiem_kho_showroom_config.json",
                    user_home / "kiem_kho_showroom_config.json",
                ])
            else:
                search_locations.append(Path(__file__).parent / "kiem_kho_showroom_config.json")
            
            # Loại bỏ trùng lặp và chỉ giữ các file tồn tại
            # Windows-safe: Normalize paths để so sánh (case-insensitive trên Windows)
            unique_locations = []
            seen = set()
            for loc in search_locations:
                try:
                    if loc.exists():
                        # Windows-safe: Normalize path và lowercase để so sánh
                        if sys.platform == 'win32':
                            loc_str = str(loc.resolve()).lower().replace('\\', '/')
                        else:
                            loc_str = str(loc.resolve())
                        
                        if loc_str not in seen:
                            seen.add(loc_str)
                            unique_locations.append(loc)
                except Exception:
                    # Bỏ qua nếu không thể resolve
                    continue
            
            # Đọc từ các vị trí theo thứ tự ưu tiên
            for config_file_path in unique_locations:
                max_retries = 3
                retry_count = 0
                while retry_count < max_retries:
                    try:
                        with open(config_file_path, 'r', encoding='utf-8') as f:
                            config = json.load(f)
                        
                        # Format mới: có cả template_file_path, auto_save_folder và config_folder
                        if 'template_file_path' in config and 'auto_save_folder' in config:
                            # Normalize paths cho Windows (quan trọng: phải normalize đúng cách)
                            template_path = None
                            auto_folder = None
                            config_folder = None
                            
                            try:
                                if config['template_file_path']:
                                    template_path = str(Path(config['template_file_path']).resolve())
                            except:
                                pass
                            
                            try:
                                if config['auto_save_folder']:
                                    auto_folder = str(Path(config['auto_save_folder']).resolve())
                            except:
                                pass
                            
                            try:
                                if config.get('config_folder'):
                                    config_folder = str(Path(config['config_folder']).resolve())
                            except:
                                pass
                            
                            # QUAN TRỌNG: Nếu file này có config_folder, đọc file config từ đó (vị trí thực sự)
                            if config_folder:
                                try:
                                    actual_config_file = Path(config_folder) / "kiem_kho_showroom_config.json"
                                    # Windows-safe: So sánh đường dẫn case-insensitive
                                    if actual_config_file.exists():
                                        current_path_str = str(config_file_path.resolve())
                                        actual_path_str = str(actual_config_file.resolve())
                                        
                                        # So sánh case-insensitive trên Windows
                                        if sys.platform == 'win32':
                                            if current_path_str.lower() != actual_path_str.lower():
                                                # Đọc lại từ file config thực sự
                                                with open(actual_config_file, 'r', encoding='utf-8') as f2:
                                                    config = json.load(f2)
                                                # Normalize lại paths
                                                if config.get('template_file_path'):
                                                    template_path = str(Path(config['template_file_path']).resolve())
                                                if config.get('auto_save_folder'):
                                                    auto_folder = str(Path(config['auto_save_folder']).resolve())
                                                if config.get('config_folder'):
                                                    config_folder = str(Path(config['config_folder']).resolve())
                                        else:
                                            if current_path_str != actual_path_str:
                                                # Đọc lại từ file config thực sự
                                                with open(actual_config_file, 'r', encoding='utf-8') as f2:
                                                    config = json.load(f2)
                                                # Normalize lại paths
                                                if config.get('template_file_path'):
                                                    template_path = str(Path(config['template_file_path']).resolve())
                                                if config.get('auto_save_folder'):
                                                    auto_folder = str(Path(config['auto_save_folder']).resolve())
                                                if config.get('config_folder'):
                                                    config_folder = str(Path(config['config_folder']).resolve())
                                except Exception as e:
                                    print(f"Lỗi khi đọc file config thực sự: {str(e)}")
                                    # Nếu không đọc được, dùng config từ file đầu tiên
                            
                            # Cập nhật config_folder và config_file
                            if config_folder:
                                self.config_folder = config_folder
                                self.config_file = Path(self.config_folder) / "kiem_kho_showroom_config.json"
                            
                            return {
                                'template_file_path': template_path,
                                'auto_save_folder': auto_folder,
                                'config_folder': config_folder
                            }
                        
                        # Format cũ - hỗ trợ backward compatibility
                        elif 'auto_save_folder' in config:
                            auto_folder = str(Path(config['auto_save_folder']).resolve()) if config['auto_save_folder'] else None
                            return {
                                'template_file_path': None,
                                'auto_save_folder': auto_folder,
                                'config_folder': None
                            }
                        
                        break  # Thành công, thoát khỏi loop
                    except PermissionError:
                        retry_count += 1
                        if retry_count < max_retries:
                            time.sleep(0.2)
                        else:
                            # Chuyển sang file tiếp theo
                            break
                    except Exception as e:
                        # Chuyển sang file tiếp theo
                        break
        except Exception as e:
            print(f"Lỗi khi đọc config: {str(e)}")
            traceback.print_exc()
        return None
    
    def save_config(self, template_path, folder_path, config_folder=None):
        """Lưu cấu hình vào file - Windows-safe"""
        try:
            # Normalize paths cho Windows
            template_path_normalized = str(Path(template_path).resolve())
            folder_path_normalized = str(Path(folder_path).resolve())
            
            # Nếu có config_folder, normalize và cập nhật
            if config_folder:
                config_folder_normalized = str(Path(config_folder).resolve())
                self.config_folder = config_folder_normalized
                # Cập nhật lại config_file path
                self.config_file = Path(self.config_folder) / "kiem_kho_showroom_config.json"
            
            config = {
                'template_file_path': template_path_normalized,
                'auto_save_folder': folder_path_normalized,
                'config_folder': str(Path(self.config_folder).resolve()) if self.config_folder else None
            }
            
            # Windows-specific: Retry nếu file bị lock
            max_retries = 3
            retry_count = 0
            save_success = False
            
            while retry_count < max_retries and not save_success:
                try:
                    with open(self.config_file, 'w', encoding='utf-8') as f:
                        json.dump(config, f, ensure_ascii=False, indent=2)
                    save_success = True
                except PermissionError:
                    retry_count += 1
                    if retry_count < max_retries:
                        time.sleep(0.2)
                    else:
                        print(f"Lỗi khi lưu config: File bị lock sau {max_retries} lần thử")
                except Exception as e:
                    print(f"Lỗi khi lưu config: {str(e)}")
                    traceback.print_exc()
                    break
        except Exception as e:
            print(f"Lỗi khi lưu config: {str(e)}")
            traceback.print_exc()
    
    def setup_paths(self):
        """Hiển thị dialog để cấu hình 2 đường dẫn"""
        try:
            # Đảm bảo root window được hiển thị và update trước
            self.root.deiconify()
            self.root.update()
            self.root.update_idletasks()
        except Exception as e:
            print(f"Lỗi khi hiển thị root window: {str(e)}")
        
        try:
            dialog = tk.Toplevel(self.root)
            dialog.title("Cấu hình đường dẫn")
            dialog.geometry("700x480")
            dialog.configure(bg='#F5F5F5')
            dialog.transient(self.root)
            dialog.grab_set()  # Modal dialog
            
            # Đặt cửa sổ ở giữa màn hình
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (700 // 2)
            y = (dialog.winfo_screenheight() // 2) - (480 // 2)
            dialog.geometry(f"700x480+{x}+{y}")
            
            # Đảm bảo dialog được focus và hiển thị trên cùng
            dialog.lift()
            dialog.focus_force()
            dialog.attributes('-topmost', True)  # Luôn hiển thị trên cùng
            dialog.update()
        except Exception as e:
            print(f"Lỗi khi tạo dialog: {str(e)}")
            import traceback
            traceback.print_exc()
            # Nếu không thể tạo dialog, hiển thị messagebox thay thế
            try:
                messagebox.showerror("Lỗi", f"Không thể hiển thị dialog cấu hình:\n{str(e)}")
            except:
                pass
            return
        
        # Label hướng dẫn
        label_text = "Cấu hình 3 đường dẫn:\n1. CHỌN ĐƯỜNG DẪN FILE TỔNG HỢP MẶC ĐỊNH (để copy khi SAVE)\n2. CHỌN ĐƯỜNG DẪN FILE THEO DÕI CHI TIẾT (Kiemkecuoinam)\n3. Thư mục lưu file cấu hình (kiem_kho_showroom_config.json)"
        tk.Label(dialog, text=label_text, bg='#F5F5F5', fg='#000000', 
                font=('Arial', 11), justify=tk.LEFT, wraplength=650).pack(pady=15, padx=20)
        
        # Load giá trị đã lưu nếu có
        saved_config = self.load_config()
        
        # Đường dẫn 1: CHỌN ĐƯỜNG DẪN FILE TỔNG HỢP MẶC ĐỊNH
        tk.Label(dialog, text="1. CHỌN ĐƯỜNG DẪN FILE TỔNG HỢP MẶC ĐỊNH:", bg='#F5F5F5', fg='#000000', 
                font=('Arial', 10, 'bold'), anchor='w').pack(pady=(10, 5), padx=20, fill=tk.X)
        
        input_frame1 = tk.Frame(dialog, bg='#F5F5F5')
        input_frame1.pack(pady=5, padx=20, fill=tk.X)
        
        path1_var = tk.StringVar()
        if saved_config and saved_config.get('template_file_path'):
            path1_var.set(saved_config['template_file_path'])
        
        path1_entry = tk.Entry(input_frame1, textvariable=path1_var, width=50, 
                              font=('Arial', 10), relief=tk.SOLID, bd=1,
                              bg='#FFFFFF', fg='#000000', insertbackground='#000000')
        path1_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        def browse_file1():
            # Tạm thời hạ dialog cấu hình xuống để dialog chọn file có thể hiển thị
            dialog.attributes('-topmost', False)
            dialog.grab_release()
            dialog.update()
            
            file_path = filedialog.askopenfilename(
                title="Chọn CHỌN ĐƯỜNG DẪN FILE TỔNG HỢP MẶC ĐỊNH",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            # Khôi phục lại dialog cấu hình lên trên cùng
            dialog.grab_set()
            dialog.attributes('-topmost', True)
            dialog.lift()
            dialog.focus_force()
            dialog.update()
            
            if file_path:
                path1_var.set(file_path)
        
        browse1_btn = tk.Button(input_frame1, text="Chọn file", command=browse_file1,
                               bg='#C8E6C9', fg='#000000', font=('Arial', 9, 'bold'), 
                               relief=tk.RAISED, bd=2, padx=12, pady=4,
                               activebackground='#A5D6A7', activeforeground='#000000',
                               cursor='hand2')
        browse1_btn.pack(side=tk.RIGHT)
        
        # Đường dẫn 2: CHỌN ĐƯỜNG DẪN FILE THEO DÕI CHI TIẾT
        tk.Label(dialog, text="2. CHỌN ĐƯỜNG DẪN FILE THEO DÕI CHI TIẾT (Kiemkecuoinam):", bg='#F5F5F5', fg='#000000', 
                font=('Arial', 10, 'bold'), anchor='w').pack(pady=(15, 5), padx=20, fill=tk.X)
        
        input_frame2 = tk.Frame(dialog, bg='#F5F5F5')
        input_frame2.pack(pady=5, padx=20, fill=tk.X)
        
        path2_var = tk.StringVar()
        if saved_config and saved_config.get('auto_save_folder'):
            path2_var.set(saved_config['auto_save_folder'])
        
        path2_entry = tk.Entry(input_frame2, textvariable=path2_var, width=50, 
                              font=('Arial', 10), relief=tk.SOLID, bd=1,
                              bg='#FFFFFF', fg='#000000', insertbackground='#000000')
        path2_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        def browse_folder2():
            # Tạm thời hạ dialog cấu hình xuống để dialog chọn thư mục có thể hiển thị
            dialog.attributes('-topmost', False)
            dialog.grab_release()
            dialog.update()
            
            folder = filedialog.askdirectory(title="Chọn CHỌN ĐƯỜNG DẪN FILE THEO DÕI CHI TIẾT")
            
            # Khôi phục lại dialog cấu hình lên trên cùng
            dialog.grab_set()
            dialog.attributes('-topmost', True)
            dialog.lift()
            dialog.focus_force()
            dialog.update()
            
            if folder:
                path2_var.set(folder)
        
        browse2_btn = tk.Button(input_frame2, text="Chọn thư mục", command=browse_folder2,
                               bg='#C8E6C9', fg='#000000', font=('Arial', 9, 'bold'), 
                               relief=tk.RAISED, bd=2, padx=12, pady=4,
                               activebackground='#A5D6A7', activeforeground='#000000',
                               cursor='hand2')
        browse2_btn.pack(side=tk.RIGHT)
        
        # Đường dẫn 3: Thư mục lưu file config
        tk.Label(dialog, text="3. Thư mục lưu file cấu hình (kiem_kho_showroom_config.json):", bg='#F5F5F5', fg='#000000', 
                font=('Arial', 10, 'bold'), anchor='w').pack(pady=(15, 5), padx=20, fill=tk.X)
        
        input_frame3 = tk.Frame(dialog, bg='#F5F5F5')
        input_frame3.pack(pady=5, padx=20, fill=tk.X)
        
        path3_var = tk.StringVar()
        if saved_config and saved_config.get('config_folder'):
            path3_var.set(saved_config['config_folder'])
        elif self.config_folder:
            path3_var.set(self.config_folder)
        else:
            # Mặc định: thư mục chứa DuLieuDauVaoShowroom.xlsx
            default_config_dir = self.get_config_file_path().parent
            path3_var.set(str(default_config_dir))
        
        path3_entry = tk.Entry(input_frame3, textvariable=path3_var, width=50, 
                              font=('Arial', 10), relief=tk.SOLID, bd=1,
                              bg='#FFFFFF', fg='#000000', insertbackground='#000000')
        path3_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        def browse_folder3():
            # Tạm thời hạ dialog cấu hình xuống để dialog chọn thư mục có thể hiển thị
            dialog.attributes('-topmost', False)
            dialog.grab_release()
            dialog.update()
            
            folder = filedialog.askdirectory(title="Chọn thư mục lưu file cấu hình")
            
            # Khôi phục lại dialog cấu hình lên trên cùng
            dialog.grab_set()
            dialog.attributes('-topmost', True)
            dialog.lift()
            dialog.focus_force()
            dialog.update()
            
            if folder:
                path3_var.set(folder)
        
        browse3_btn = tk.Button(input_frame3, text="Chọn thư mục", command=browse_folder3,
                               bg='#C8E6C9', fg='#000000', font=('Arial', 9, 'bold'), 
                               relief=tk.RAISED, bd=2, padx=12, pady=4,
                               activebackground='#A5D6A7', activeforeground='#000000',
                               cursor='hand2')
        browse3_btn.pack(side=tk.RIGHT)
        
        # Button OK và Cancel
        button_frame = tk.Frame(dialog, bg='#F5F5F5')
        button_frame.pack(pady=20)
        
        def on_ok():
            template_path = path1_var.get().strip()
            folder_path = path2_var.get().strip()
            config_folder_path = path3_var.get().strip()
            
            if not template_path or not folder_path or not config_folder_path:
                messagebox.showwarning("Cảnh báo", "Vui lòng nhập đầy đủ 3 đường dẫn!")
                return
            
            # Kiểm tra file Excel có tồn tại không
            if not os.path.isfile(template_path):
                messagebox.showerror("Lỗi", f"File Excel không tồn tại!\n{template_path}")
                return
            
            # Kiểm tra CHỌN ĐƯỜNG DẪN FILE THEO DÕI CHI TIẾT có tồn tại không
            if not os.path.isdir(folder_path):
                messagebox.showerror("Lỗi", f"CHỌN ĐƯỜNG DẪN FILE THEO DÕI CHI TIẾT không tồn tại!\n{folder_path}")
                return
            
            # Kiểm tra thư mục lưu file config có tồn tại không
            if not os.path.isdir(config_folder_path):
                messagebox.showerror("Lỗi", f"Thư mục lưu file cấu hình không tồn tại!\n{config_folder_path}")
                return
            
            # Kiểm tra quyền ghi vào CHỌN ĐƯỜNG DẪN FILE THEO DÕI CHI TIẾT - Windows-safe
            try:
                test_file = Path(folder_path) / '.kiem_kho_test'
                # Windows-specific: Normalize path
                test_file = test_file.resolve()
                with open(test_file, 'w', encoding='utf-8') as f:
                    f.write('test')
                # Windows-specific: Retry nếu file bị lock
                max_retries = 3
                retry_count = 0
                delete_success = False
                while retry_count < max_retries and not delete_success:
                    try:
                        test_file.unlink()
                        delete_success = True
                    except PermissionError:
                        retry_count += 1
                        if retry_count < max_retries:
                            time.sleep(0.2)
                        else:
                            # File có thể bị lock, nhưng không quan trọng vì chỉ là file test
                            pass
            except Exception as e:
                error_msg = str(e)
                traceback.print_exc()
                messagebox.showerror("Lỗi", f"Không có quyền ghi vào CHỌN ĐƯỜNG DẪN FILE THEO DÕI CHI TIẾT!\n{error_msg}")
                return
            
            # Kiểm tra quyền ghi vào thư mục lưu file config - Windows-safe
            try:
                test_file_config = Path(config_folder_path) / '.kiem_kho_test_config'
                # Windows-specific: Normalize path
                test_file_config = test_file_config.resolve()
                with open(test_file_config, 'w', encoding='utf-8') as f:
                    f.write('test')
                # Windows-specific: Retry nếu file bị lock
                max_retries = 3
                retry_count = 0
                delete_success = False
                while retry_count < max_retries and not delete_success:
                    try:
                        test_file_config.unlink()
                        delete_success = True
                    except PermissionError:
                        retry_count += 1
                        if retry_count < max_retries:
                            time.sleep(0.2)
                        else:
                            # File có thể bị lock, nhưng không quan trọng vì chỉ là file test
                            pass
            except Exception as e:
                error_msg = str(e)
                traceback.print_exc()
                messagebox.showerror("Lỗi", f"Không có quyền ghi vào thư mục lưu file cấu hình!\n{error_msg}")
                return
            
            # Lưu cấu hình (lưu vào thư mục config mới)
            self.save_config(template_path, folder_path, config_folder_path)
            
            # Cập nhật biến (quan trọng: phải cập nhật sau khi save để lần sau load được)
            self.template_file_path = template_path
            self.auto_save_folder = folder_path
            self.config_folder = config_folder_path
            # Cập nhật config_file để trỏ đến đúng vị trí
            self.config_file = Path(self.config_folder) / "kiem_kho_showroom_config.json"
            
            # QUAN TRỌNG: Lưu file pointer trỏ đến vị trí config thực sự (đã mã hóa)
            # File này sẽ được lưu ở vị trí cố định (cùng thư mục với exe) để dễ tìm
            try:
                location_file = self.get_config_location_file()
                # Mã hóa đường dẫn trước khi lưu
                encoded_path = self._encode_path(str(self.config_file.resolve()))
                if encoded_path:
                    with open(location_file, 'w', encoding='utf-8') as f:
                        f.write(encoded_path)
                    # Đặt thuộc tính hidden/system trên Windows để chặn người dùng mở
                    self._set_file_hidden(location_file)
                    print(f"[OK] Đã lưu file pointer (đã mã hóa và ẩn) vào: {location_file}")
                    print(f"[OK] File pointer trỏ đến: {self.config_file}")
                else:
                    print(f"[WARNING] Không thể mã hóa đường dẫn")
            except Exception as e:
                print(f"[WARNING] Không thể lưu file pointer: {str(e)}")
                # Không quan trọng lắm, vẫn có thể tìm bằng cách khác
            
            # Đảm bảo config đã được lưu thành công
            if self.config_file.exists():
                print(f"[OK] Đã lưu cấu hình vào: {self.config_file}")
            
            dialog.destroy()
        
        def on_cancel():
            # Nếu không nhập đầy đủ và bấm "Bỏ qua", tắt phần mềm
            template_path = path1_var.get().strip()
            folder_path = path2_var.get().strip()
            config_folder_path = path3_var.get().strip()
            if not template_path or not folder_path or not config_folder_path:
                messagebox.showinfo("Thông báo", "Phần mềm sẽ đóng vì chưa cấu hình đầy đủ đường dẫn.")
                self.root.quit()
                sys.exit(0)
            else:
                dialog.destroy()
                self.root.quit()
                sys.exit(0)
        
        ok_btn = tk.Button(button_frame, text="OK", command=on_ok,
                          bg='#BBDEFB', fg='#000000', font=('Arial', 11, 'bold'),
                          relief=tk.RAISED, bd=2, padx=25, pady=8,
                          activebackground='#90CAF9', activeforeground='#000000',
                          cursor='hand2')
        ok_btn.pack(side=tk.LEFT, padx=8)
        
        cancel_btn = tk.Button(button_frame, text="Bỏ qua", command=on_cancel,
                               bg='#FFCDD2', fg='#000000', font=('Arial', 11, 'bold'),
                               relief=tk.RAISED, bd=2, padx=25, pady=8,
                               activebackground='#EF9A9A', activeforeground='#000000',
                               cursor='hand2')
        cancel_btn.pack(side=tk.LEFT, padx=8)
        
        # Bind Enter key
        path1_entry.bind('<Return>', lambda e: path2_entry.focus())
        path2_entry.bind('<Return>', lambda e: path3_entry.focus())
        path3_entry.bind('<Return>', lambda e: on_ok())
        
        # Focus vào ô đầu tiên
        path1_entry.focus()
        path1_entry.select_range(0, tk.END)
        
        # Đợi dialog đóng
        dialog.wait_window()
        
    def get_original_dist_path(self):
        """Lấy đường dẫn thư mục dist gốc từ file config (được đóng gói trong exe)"""
        if not getattr(sys, 'frozen', False):
            return None
        
        try:
            # Ưu tiên 1: Tìm file config trong thư mục chứa exe (nếu copy cùng với exe)
            exe_dir = Path(sys.executable).parent
            config_file_py = exe_dir / "dist_path_config.py"
            
            # Thử đọc từ file Python trong thư mục exe trước
            if config_file_py.exists():
                try:
                    import importlib.util
                    spec = importlib.util.spec_from_file_location("dist_path_config", config_file_py)
                    if spec and spec.loader:
                        config_module = importlib.util.module_from_spec(spec)
                        spec.loader.exec_module(config_module)
                        if hasattr(config_module, 'DIST_PATH') and config_module.DIST_PATH:
                            dist_path = Path(config_module.DIST_PATH)
                            if dist_path.exists():
                                return dist_path
                except Exception as e:
                    print(f"Lỗi khi đọc dist_path_config.py từ exe dir: {str(e)}")
            
            # Ưu tiên 2: Đọc từ file Python đã đóng gói trong exe (bundle)
            try:
                if getattr(sys, '_MEIPASS', None):
                    config_path = Path(sys._MEIPASS) / "dist_path_config.py"
                    if config_path.exists():
                        import importlib.util
                        spec = importlib.util.spec_from_file_location("dist_path_config", config_path)
                        if spec and spec.loader:
                            config_module = importlib.util.module_from_spec(spec)
                            spec.loader.exec_module(config_module)
                            if hasattr(config_module, 'DIST_PATH') and config_module.DIST_PATH:
                                dist_path = Path(config_module.DIST_PATH)
                                if dist_path.exists():
                                    return dist_path
            except Exception as e:
                print(f"Lỗi khi đọc dist_path_config.py từ bundle: {str(e)}")
            
            # Ưu tiên 3: Tìm file txt trong thư mục chứa exe
            config_file = exe_dir / "dist_path.txt"
            
            # Ưu tiên 3: Tìm trong thư mục cha
            if not config_file.exists():
                parent_dir = exe_dir.parent
                config_file = parent_dir / "dist_path.txt"
            
            # Ưu tiên 4: Tìm trong các thư mục phổ biến
            if not config_file.exists():
                user_home = Path.home()
                possible_locations = [
                    user_home / "Desktop" / "dist" / "dist_path.txt",
                    user_home / "Documents" / "dist" / "dist_path.txt",
                    user_home / "Downloads" / "dist" / "dist_path.txt",
                ]
                for loc in possible_locations:
                    if loc.exists():
                        config_file = loc
                        break
            
            # Đọc đường dẫn từ file txt (backup)
            if config_file.exists():
                with open(config_file, 'r', encoding='utf-8') as f:
                    dist_path = f.read().strip()
                    if dist_path and Path(dist_path).exists():
                        return Path(dist_path)
        except Exception as e:
            print(f"Lỗi khi đọc dist_path config: {str(e)}")
        
        return None
    
    def load_data(self):
        """Load dữ liệu từ file Excel"""
        # Sử dụng pandas đã được import trong __init__
        if self.pd is None:
            try:
                import pandas as pd
                self.pd = pd
            except ImportError:
                messagebox.showerror("Lỗi", "Không thể import pandas! Vui lòng cài đặt: pip install pandas")
                sys.exit(1)
        
        pd = self.pd  # Alias để dùng trong hàm này
        
        try:
            # Kiểm tra nếu đang chạy từ executable (PyInstaller)
            if getattr(sys, 'frozen', False):
                # Chạy từ executable
                # ƯU TIÊN: Tìm file từ thư mục dist gốc (nơi build ban đầu)
                original_dist_path = self.get_original_dist_path()
                
                if original_dist_path:
                    # Tìm file từ thư mục dist gốc
                    excel_path = original_dist_path / "DuLieuDauVaoShowroom.xlsx"
                    search_dir = original_dist_path
                else:
                    # Fallback: Tìm trong thư mục chứa executable
                    exe_dir = Path(sys.executable).parent
                    excel_path = exe_dir / "DuLieuDauVaoShowroom.xlsx"
                    search_dir = exe_dir
                
                # Nếu không có trong thư mục dist gốc hoặc thư mục exe, thử tìm trong bundle
                if not excel_path.exists():
                    base_path = Path(sys._MEIPASS)
                    excel_path = base_path / "DuLieuDauVaoShowroom.xlsx"
                
                # Danh sách file thay thế nếu file chính không đọc được
                xls_alternatives = [
                    search_dir / "KIEM KE Năm -2025 - BP ONLINE.xls",
                    search_dir / "KIEM KE Năm -2025 - BP ONLINE copy.xls",
                    search_dir / "DuLieuDauVao.xls",
                ]
                # Fallback: thử trong bundle nếu không tìm thấy
                if getattr(sys, '_MEIPASS', None):
                    base_path = Path(sys._MEIPASS)
                    xls_alternatives.extend([
                        base_path / "KIEM KE Năm -2025 - BP ONLINE.xls",
                        base_path / "KIEM KE Năm -2025 - BP ONLINE copy.xls",
                        base_path / "DuLieuDauVao.xls",
                    ])
            else:
                # Chạy từ source code
                excel_path = Path(__file__).parent / "DuLieuDauVaoShowroom.xlsx"
                # Danh sách file thay thế nếu file chính không đọc được
                xls_alternatives = [
                    Path(__file__).parent / "KIEM KE Năm -2025 - BP ONLINE.xls",
                    Path(__file__).parent / "KIEM KE Năm -2025 - BP ONLINE copy.xls",
                    Path(__file__).parent / "DuLieuDauVao.xls",
                ]
            
            # Kiểm tra file .xlsx có hợp lệ không
            if excel_path.exists() and excel_path.suffix.lower() == '.xlsx':
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(excel_path, read_only=True)
                    if len(wb.sheetnames) == 0:
                        # File .xlsx không có worksheet, tìm file .xls thay thế
                        wb.close()
                        for alt_path in xls_alternatives:
                            if alt_path.exists():
                                excel_path = alt_path
                                break
                        else:
                            raise ValueError("File Excel không có worksheet nào và không tìm thấy file thay thế!")
                    else:
                        wb.close()
                except Exception as e:
                    # Nếu file .xlsx lỗi, thử tìm file .xls thay thế
                    for alt_path in xls_alternatives:
                        if alt_path.exists():
                            excel_path = alt_path
                            break
            
            # Nếu vẫn không tìm thấy file hợp lệ
            if not excel_path.exists():
                # Thử tìm file .xls
                for alt_path in xls_alternatives:
                    if alt_path.exists():
                        excel_path = alt_path
                        break
                
                # Nếu vẫn không có, cho phép người dùng chọn file
                if not excel_path.exists():
                    excel_path = filedialog.askopenfilename(
                        title="Chọn file dữ liệu Excel",
                        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
                    )
                    if not excel_path:
                        messagebox.showerror("Lỗi", "Không tìm thấy file dữ liệu!")
                        sys.exit(1)
                    excel_path = Path(excel_path)
            
            # Đọc file Excel (hỗ trợ cả .xls và .xlsx)
            # Ưu tiên thử đọc với header=0 (dòng 1) trước
            header_row = None
            
            if excel_path.suffix.lower() == '.xls':
                # Thử đọc với header=0 trước (dòng 1 trong Excel)
                try:
                    temp_df = pd.read_excel(excel_path, engine='xlrd', header=0, nrows=2)
                    temp_df.columns = temp_df.columns.str.strip()
                    # Kiểm tra xem có đủ cột cần thiết không
                    cols_lower = [str(c).lower() if pd.notna(c) else '' for c in temp_df.columns]
                    cols_str = ' '.join(cols_lower)
                    keywords_found = sum([
                        'isbn' in cols_str,
                        'số thùng' in cols_str or 'so thung' in cols_str or 'thùng' in cols_str,
                        'tựa' in cols_str or 'tua' in cols_str or 'titles' in cols_str,
                        'tồn' in cols_str or 'ton' in cols_str or 'qty' in cols_str
                    ])
                    if keywords_found >= 3:  # Nếu tìm thấy ít nhất 3 từ khóa, dùng header=0
                        header_row = 0
                except:
                    pass
                
                # Nếu header=0 không hợp lệ, tìm header tự động
                if header_row is None:
                    try:
                        temp_df = pd.read_excel(excel_path, engine='xlrd', header=None, nrows=30)
                        for idx, row in temp_df.iterrows():
                            row_values = [str(cell).lower() if pd.notna(cell) else '' for cell in row.values]
                            row_str = ' '.join(row_values)
                            if 'isbn' in row_str:
                                keywords_found = sum([
                                    'isbn' in row_str,
                                    'số thùng' in row_str or 'so thung' in row_str or 'thùng' in row_str,
                                    'tựa' in row_str or 'tua' in row_str or 'titles' in row_str,
                                    'tồn' in row_str or 'ton' in row_str or 'qty' in row_str
                                ])
                                if keywords_found >= 2:
                                    header_row = idx
                                    break
                    except:
                        pass
                
                # Đọc file với header row đã tìm được
                try:
                    if header_row is not None:
                        self.df = pd.read_excel(excel_path, engine='xlrd', header=header_row)
                    else:
                        # Mặc định dùng header=0
                        self.df = pd.read_excel(excel_path, engine='xlrd', header=0)
                except Exception as e2:
                    raise ValueError(f"Không thể đọc file .xls: {str(e2)}")
            else:
                # File .xlsx - thử đọc với header=0 trước
                try:
                    self.df = pd.read_excel(excel_path, engine='openpyxl', header=0)
                except Exception as e2:
                    raise ValueError(f"Không thể đọc file .xlsx: {str(e2)}")
            
            # Loại bỏ các dòng rỗng hoàn toàn
            self.df = self.df.dropna(how='all')
            
            # Kiểm tra DataFrame có rỗng không (trước khi filter ISBN)
            if self.df.empty:
                raise ValueError("File Excel không có dữ liệu!")
            
            # Loại bỏ các dòng có ISBN rỗng hoặc không hợp lệ (chỉ khi đã có cột ISBN)
            if 'isbn' in self.df.columns or 'ISBN' in self.df.columns:
                isbn_col = 'isbn' if 'isbn' in self.df.columns else 'ISBN'
                original_count = len(self.df)
                
                # Loại bỏ các dòng có ISBN rỗng
                self.df = self.df[self.df[isbn_col].notna()]
                
                # Loại bỏ các dòng có ISBN là số thứ tự đơn giản (1, 2, 3...) hoặc số nhỏ hơn 4 chữ số
                if len(self.df) > 0:
                    # Chuyển ISBN sang string để kiểm tra
                    isbn_str = self.df[isbn_col].astype(str)
                    # Loại bỏ các ISBN chỉ là số đơn giản (1.0, 2.0, ...) hoặc có ít hơn 4 ký tự số
                    mask = ~isbn_str.str.match(r'^\d+\.0?$')  # Không phải chỉ số đơn giản
                    # Giữ lại các ISBN có độ dài hợp lý (ít nhất 4 ký tự sau khi loại bỏ .0)
                    mask = mask & (isbn_str.str.replace('.0', '').str.len() >= 4)
                    self.df = self.df[mask]
                
                # Kiểm tra lại sau khi filter
                if self.df.empty:
                    raise ValueError(f"File Excel không có dữ liệu hợp lệ! Đã loại bỏ {original_count} dòng không hợp lệ.")
            
            # Kiểm tra lại DataFrame có rỗng không (sau khi filter)
            if self.df.empty:
                raise ValueError("File Excel không có dữ liệu sau khi lọc!")
            
            # Xử lý DataFrame
            self._process_dataframe()
            
        except Exception as e:
            error_msg = f"Không thể đọc file Excel: {str(e)}\n\n"
            error_msg += "Vui lòng kiểm tra:\n"
            error_msg += "1. File Excel có đúng định dạng không (.xlsx hoặc .xls)\n"
            error_msg += "2. File có chứa dữ liệu không\n"
            error_msg += "3. File không bị hỏng\n\n"
            error_msg += "Bạn có muốn chọn file khác không?"
            
            result = messagebox.askyesno("Lỗi", error_msg)
            if result:
                # Cho phép chọn file khác
                excel_path = filedialog.askopenfilename(
                    title="Chọn file dữ liệu Excel",
                    filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
                )
                if excel_path:
                    try:
                        # Đảm bảo pandas đã được import
                        if self.pd is None:
                            import pandas as pd
                            self.pd = pd
                        pd = self.pd
                        
                        # Thử đọc lại với file mới
                        if Path(excel_path).suffix.lower() == '.xls':
                            self.df = pd.read_excel(excel_path, engine='xlrd')
                        else:
                            self.df = pd.read_excel(excel_path, engine='openpyxl')
                        # Nếu thành công, tiếp tục xử lý
                        self._process_dataframe()
                        return
                    except Exception as e2:
                        messagebox.showerror("Lỗi", f"Vẫn không thể đọc file: {str(e2)}")
                        sys.exit(1)
                else:
                    messagebox.showerror("Lỗi", "Không có file nào được chọn!")
                    sys.exit(1)
            else:
                sys.exit(1)
    
    def load_data_deferred(self):
        """Load dữ liệu sau khi UI đã hiển thị (deferred loading để tăng tốc độ khởi động)"""
        try:
            # Cập nhật cursor để hiển thị đang load
            self.root.config(cursor='wait')
            self.root.update_idletasks()
            
            # Load dữ liệu
            self.load_data()
        finally:
            # Khôi phục cursor
            self.root.config(cursor='')
    
    def _process_dataframe(self):
        """Xử lý DataFrame sau khi đọc thành công"""
        # Đảm bảo pandas đã được import
        if self.pd is None:
            try:
                import pandas as pd
                self.pd = pd
            except ImportError:
                messagebox.showerror("Lỗi", "Không thể import pandas! Vui lòng cài đặt: pip install pandas")
                return
        
        pd = self.pd  # Alias để dùng trong hàm này
        
        # Chuẩn hóa tên cột (loại bỏ khoảng trắng thừa, chuyển sang lowercase)
        self.df.columns = self.df.columns.str.strip()
        
        # Tìm các cột cần thiết (case-insensitive)
        col_mapping = {}
        for col in self.df.columns:
            if pd.isna(col):
                continue
            col_str = str(col).strip()
            col_lower = col_str.lower()
            
            # Bỏ logic tìm cột số thùng - Showroom không cần số thùng trong Excel
            
            # ISBN
            if 'isbn' in col_lower:
                if 'isbn' not in col_mapping:
                    col_mapping['isbn'] = col
            
            # Tựa/Tên sách / Titles
            if ('tựa' in col_lower or 'tua' in col_lower or 'tên' in col_lower or 
                'titles' in col_lower or 'title' in col_lower):
                if 'tua' not in col_mapping:
                    col_mapping['tua'] = col
            
            # Tồn từng tựa / Qty tựa trong thùng / Qty
            # Ưu tiên các cột có tên đầy đủ trước
            if 'qty tựa trong thùng' in col_lower or 'qty tua trong thung' in col_lower:
                if 'ton_tung_tua' not in col_mapping:
                    col_mapping['ton_tung_tua'] = col
            elif (('tồn' in col_lower and 'tựa' in col_lower) or 
                  ('ton' in col_lower and 'tua' in col_lower)):
                if 'ton_tung_tua' not in col_mapping:
                    col_mapping['ton_tung_tua'] = col
            elif col_lower == 'qty' and 'ton_tung_tua' not in col_mapping:
                # Chỉ dùng Qty nếu chưa tìm thấy cột nào khác
                col_mapping['ton_tung_tua'] = col
        
        # Kiểm tra xem có đủ cột không (chỉ cần 3 cột: isbn, tựa, tồn tựa)
        if len(col_mapping) < 3:
            messagebox.showwarning("Cảnh báo", 
                f"Không tìm thấy đủ các cột cần thiết. Cần: isbn, tựa, tồn tựa\n"
                f"Tìm thấy: {list(col_mapping.keys())}\n"
                f"Các cột trong file: {list(self.df.columns)}")
        
        # Đổi tên cột để dễ sử dụng
        if col_mapping:
            self.df = self.df.rename(columns=col_mapping)
        
        # Làm sạch dữ liệu
        if 'isbn' in self.df.columns:
            self.df['isbn'] = self.df['isbn'].astype(str).str.strip()
    
    def create_ui(self):
        """Tạo giao diện người dùng"""
        # Màu sắc
        bg_color = '#F5F5F5'  # Nền chính nhẹ nhàng
        input_bg = '#FFFFFF'  # Nền input trắng
        input_bg_yellow = '#FFF9C4'  # Nền vàng nhẹ nhàng hơn
        label_fg = '#333333'  # Màu chữ đen nhẹ
        label_required_fg = '#C62828'  # Đỏ đậm cho label bắt buộc
        button_bg = '#E3F2FD'  # Nền button xanh nhẹ
        
        # Tạo Notebook để chứa các tab
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # === TAB 1: KIỂM KÊ ===
        tab_kiemke = tk.Frame(self.notebook, bg=bg_color)
        self.notebook.add(tab_kiemke, text="Kiểm kê")
        
        # Frame chính cho tab Kiểm kê - dùng grid để kiểm soát layout tốt hơn
        main_frame = tk.Frame(tab_kiemke, bg=bg_color)
        main_frame.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)
        tab_kiemke.grid_rowconfigure(0, weight=1)
        tab_kiemke.grid_columnconfigure(0, weight=1)
        
        # === PHẦN NHẬP THÔNG TIN THÙNG ===
        info_frame = tk.Frame(main_frame, bg=bg_color)
        info_frame.grid(row=0, column=0, sticky='ew', pady=(0, 10))
        main_frame.grid_columnconfigure(0, weight=1)
        
        # Màu chữ trong input - đen đậm để dễ nhìn
        input_fg = '#000000'  # Đen đậm
        
        # Nhập/Xuất
        tk.Label(info_frame, text="Nhập/Xuất:", bg=bg_color, fg=label_fg, font=('Arial', 11)).grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.nhap_xuat_var = tk.StringVar(value="KIEM KE Năm 2025 - BP Online")
        entry1 = tk.Entry(info_frame, textvariable=self.nhap_xuat_var, width=40, bg=input_bg_yellow, 
                          fg=input_fg, font=('Arial', 10), relief=tk.SOLID, bd=1, insertbackground='#000000')
        entry1.grid(row=0, column=1, columnspan=2, padx=5, pady=5, sticky='ew')
        
        # Thùng / vị trí mới
        tk.Label(info_frame, text="Thùng / vị trí mới:", bg=bg_color, fg=label_fg, font=('Arial', 11)).grid(row=0, column=3, sticky='w', padx=5, pady=5)
        self.vi_tri_moi_var = tk.StringVar()
        self.vi_tri_moi_entry = tk.Entry(info_frame, textvariable=self.vi_tri_moi_var, width=20, bg=input_bg_yellow,
                          fg=input_fg, font=('Arial', 10), relief=tk.SOLID, bd=1, insertbackground='#000000')
        self.vi_tri_moi_entry.grid(row=0, column=4, padx=5, pady=5, sticky='ew')
        # Thêm validation khi người dùng nhập xong
        self.vi_tri_moi_entry.bind('<FocusOut>', lambda e: self.validate_vi_tri_moi())
        self.vi_tri_moi_entry.bind('<Return>', lambda e: self.validate_vi_tri_moi())
        
        # Số thùng (*)
        tk.Label(info_frame, text="Số thùng (*):", bg=bg_color, fg=label_required_fg, font=('Arial', 11, 'bold')).grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.so_thung_var = tk.StringVar()
        self.so_thung_entry = tk.Entry(info_frame, textvariable=self.so_thung_var, width=40, bg=input_bg_yellow,
                                       fg=input_fg, font=('Arial', 11), relief=tk.SOLID, bd=1, insertbackground='#000000')
        self.so_thung_entry.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky='ew')
        # Bỏ nút Load và logic load_box_data - Showroom không cần load theo số thùng
        # BỎ event handler để tránh reset "Đã quét" khi nhập số thùng mới
        # Chỉ cập nhật current_box_number khi quét ISBN, không cập nhật "Đã quét" khi nhập số thùng
        # self.so_thung_entry.bind('<FocusOut>', lambda e: self.on_so_thung_changed())
        # self.so_thung_entry.bind('<Return>', lambda e: self.on_so_thung_changed())
        
        # Lưu giá trị số thùng ban đầu để kiểm tra khi người dùng cố gắng sửa
        self.so_thung_original_value = ''
        
        # Thêm event handler để chặn việc sửa số thùng khi đã có dữ liệu quét
        self.so_thung_entry.bind('<FocusIn>', self.on_so_thung_focus_in)
        
        # Ngày (*)
        tk.Label(info_frame, text="Ngày (*):", bg=bg_color, fg=label_required_fg, font=('Arial', 11, 'bold')).grid(row=2, column=0, sticky='w', padx=5, pady=5)
        from datetime import datetime
        self.ngay_var = tk.StringVar(value=datetime.now().strftime("%d/%m/%y"))
        entry3 = tk.Entry(info_frame, textvariable=self.ngay_var, width=40, bg=input_bg_yellow,
                         fg=input_fg, font=('Arial', 10), relief=tk.SOLID, bd=1, insertbackground='#000000')
        entry3.grid(row=2, column=1, columnspan=2, padx=5, pady=5, sticky='ew')
        
        # Tổ (*) - Bắt buộc
        tk.Label(info_frame, text="Tổ (*):", bg=bg_color, fg=label_required_fg, font=('Arial', 11, 'bold')).grid(row=3, column=0, sticky='w', padx=5, pady=5)
        self.to_var = tk.StringVar()
        self.to_entry = tk.Entry(info_frame, textvariable=self.to_var, width=40, bg=input_bg_yellow,
                          fg=input_fg, font=('Arial', 10), relief=tk.SOLID, bd=1, insertbackground='#000000')
        self.to_entry.grid(row=3, column=1, columnspan=2, padx=5, pady=5, sticky='ew')
        
        # Note thùng
        tk.Label(info_frame, text="Note thùng:", bg=bg_color, fg=label_required_fg, font=('Arial', 11, 'bold')).grid(row=4, column=0, sticky='w', padx=5, pady=5)
        self.note_thung_var = tk.StringVar()
        entry5 = tk.Entry(info_frame, textvariable=self.note_thung_var, width=40, bg=input_bg_yellow,
                         fg=input_fg, font=('Arial', 10), relief=tk.SOLID, bd=1, insertbackground='#000000')
        entry5.grid(row=4, column=1, columnspan=2, padx=5, pady=5, sticky='ew')
        
        # Nút SAVE
        save_btn = tk.Button(info_frame, text="SAVE", command=self.save_data, 
                            bg='#4CAF50', fg='white', font=('Arial', 12, 'bold'), 
                            width=15, height=2, relief=tk.RAISED, bd=2, cursor='hand2')
        save_btn.grid(row=1, column=3, rowspan=2, padx=20, pady=5, sticky='n')
        
        # Nút RESET
        reset_btn = tk.Button(info_frame, text="RESET", command=self.reset_scanned_data, 
                            bg='#FF9800', fg='white', font=('Arial', 12, 'bold'), 
                            width=15, height=2, relief=tk.RAISED, bd=2, cursor='hand2')
        reset_btn.grid(row=3, column=3, rowspan=2, padx=20, pady=5, sticky='n')
        
        info_frame.columnconfigure(1, weight=1)
        
        # === PHẦN HIỂN THỊ ĐÃ QUÉT ===
        count_frame = tk.Frame(main_frame, bg=bg_color)
        count_frame.grid(row=1, column=0, sticky='ew', pady=(0, 5))
        
        # Hiển thị số dòng đã quét trong table
        tk.Label(count_frame, text="Đã quét:", bg=bg_color, fg=label_required_fg, font=('Arial', 11, 'bold')).grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.so_tua_da_quet_var = tk.StringVar(value="0")
        tk.Label(count_frame, textvariable=self.so_tua_da_quet_var, bg=bg_color, fg='#4CAF50', font=('Arial', 14, 'bold')).grid(row=0, column=1, padx=5, pady=5, sticky='w')
        
        # === PHẦN NHẬP ISBN (QUÉT MÃ VẠCH) - Đặt ở dưới cùng với grid ===
        scan_frame = tk.Frame(main_frame, bg=bg_color)
        scan_frame.grid(row=3, column=0, sticky='ew', pady=(10, 0))
        # scan_frame không expand, chỉ chiếm không gian cần thiết
        
        tk.Label(scan_frame, text="Quét/Nhập ISBN:", bg=bg_color, fg=label_fg, 
                font=('Arial', 13, 'bold')).grid(row=0, column=0, padx=5, pady=8, sticky='w')
        self.isbn_entry = tk.Entry(scan_frame, font=('Arial', 16, 'bold'), width=30, 
                                   bg='#FFFFFF', fg='#000000', relief=tk.SOLID, bd=3, insertbackground='#000000')
        # Tăng chiều cao bằng cách thêm padding
        self.isbn_entry.grid(row=0, column=1, padx=5, pady=8, sticky='ew', ipady=8)
        self.isbn_entry.bind('<Return>', self.on_isbn_entered)
        self.isbn_entry.focus()
        
        scan_frame.columnconfigure(1, weight=1)
        
        # === BẢNG DỮ LIỆU ===
        table_frame = tk.Frame(main_frame, bg=bg_color)
        # Sử dụng grid, row=2 giữa count_frame và scan_frame, với weight=1 để expand
        table_frame.grid(row=2, column=0, sticky='nsew', pady=(0, 0))
        # Đảm bảo row 2 (table_frame) có thể expand, nhưng row 3 (scan_frame) thì không
        main_frame.grid_rowconfigure(2, weight=1)  # Table frame có thể expand
        main_frame.grid_rowconfigure(3, weight=0)  # Scan frame không expand, chỉ chiếm không gian cần thiết
        
        # Tạo Treeview với scrollbar
        scrollbar_y = tk.Scrollbar(table_frame, orient=tk.VERTICAL, bg='#E0E0E0', troughcolor=bg_color)
        scrollbar_x = tk.Scrollbar(table_frame, orient=tk.HORIZONTAL, bg='#E0E0E0', troughcolor=bg_color)
        
        # Định nghĩa thứ tự cột cố định - Showroom có cột "Tình trạng" nhưng để trống
        columns = ('Số thứ tự', 'ISBN', 'Tựa', 'Tồn thực tế', 'Số thùng', 'Tồn tựa trong thùng', 'Tình trạng', 'Ghi chú', 'Xóa')
        # Thứ tự: 0=Số thứ tự, 1=ISBN, 2=Tựa, 3=Tồn thực tế, 4=Số thùng, 5=Tồn tựa trong thùng, 6=Tình trạng, 7=Ghi chú, 8=Xóa
        
        # Tạo style cho Treeview
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Treeview', background='#FFFFFF', foreground='#333333', 
                       fieldbackground='#FFFFFF', font=('Arial', 10), rowheight=25)
        style.configure('Treeview.Heading', background='#2196F3', foreground='white', 
                       font=('Arial', 10, 'bold'), relief=tk.FLAT)
        style.map('Treeview.Heading', background=[('active', '#1976D2')])
        
        # Giảm chiều cao mặc định để đảm bảo có không gian cho ô nhập ISBN
        self.tree = ttk.Treeview(table_frame, columns=columns, show='headings', 
                                 yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set,
                                 height=10, style='Treeview')
        
        # Scrollbar - bỏ logic highlight
        def yview_scroll(*args):
            self.tree.yview(*args)
        
        def xview_scroll(*args):
            self.tree.xview(*args)
        
        scrollbar_y.config(command=yview_scroll)
        scrollbar_x.config(command=xview_scroll)
        
        # Định nghĩa các cột
        self.tree.heading('Số thứ tự', text='STT')
        self.tree.heading('ISBN', text='ISBN')
        self.tree.heading('Tựa', text='Tựa')
        self.tree.heading('Tồn thực tế', text='Tồn thực tế')
        self.tree.heading('Số thùng', text='Số thùng')
        self.tree.heading('Tồn tựa trong thùng', text='Tồn tựa trong thùng')
        self.tree.heading('Tình trạng', text='Tình trạng')
        self.tree.heading('Ghi chú', text='Ghi chú')
        self.tree.heading('Xóa', text='Xóa')
        
        self.tree.column('Số thứ tự', width=80, anchor='center')
        self.tree.column('ISBN', width=150, anchor='w')
        self.tree.column('Tựa', width=300, anchor='w')
        self.tree.column('Tồn thực tế', width=120, anchor='center')
        self.tree.column('Số thùng', width=100, anchor='center')
        self.tree.column('Tồn tựa trong thùng', width=150, anchor='center')
        self.tree.column('Tình trạng', width=100, anchor='center')
        self.tree.column('Ghi chú', width=200, anchor='w')
        self.tree.column('Xóa', width=80, anchor='center')
        
        # Tag để highlight màu đỏ cho cả row khi có lỗi (không dùng nữa, chỉ highlight 2 cell)
        # self.tree.tag_configure('error', background='#FFEBEE', foreground='#C62828')
        
        self.tree.grid(row=0, column=0, sticky='nsew')
        scrollbar_y.grid(row=0, column=1, sticky='ns')
        scrollbar_x.grid(row=1, column=0, sticky='ew')
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # Bind click để edit trực tiếp tất cả các cột có thể edit
        self.tree.bind('<Button-1>', self.on_item_click)
        # Double click cũng mở edit (để dễ sử dụng hơn)
        self.tree.bind('<Double-1>', self.on_item_click)
        
        # Showroom: Bỏ logic highlight - không cần bind events
        
        # === TAB 2: TỔNG HỢP ===
        tab_tonghop = tk.Frame(self.notebook, bg=bg_color)
        self.notebook.add(tab_tonghop, text="Tổng hợp")
        
        # Frame chính cho tab Tổng hợp
        tonghop_main_frame = tk.Frame(tab_tonghop, bg=bg_color)
        tonghop_main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Tiêu đề
        title_frame = tk.Frame(tonghop_main_frame, bg=bg_color)
        title_frame.pack(fill=tk.X, pady=(0, 10))
        
        title_label = tk.Label(title_frame, text="TỔNG HỢP DỮ LIỆU NHẬP", 
                              bg=bg_color, fg='#000000', font=('Arial', 16, 'bold'))
        title_label.pack()
        
        # Frame chứa nút và ô tìm kiếm
        button_frame = tk.Frame(tonghop_main_frame, bg=bg_color)
        button_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Ô tìm kiếm ISBN
        search_frame = tk.Frame(button_frame, bg=bg_color)
        search_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        tk.Label(search_frame, text="Tìm kiếm ISBN:", bg=bg_color, fg=label_fg, 
                font=('Arial', 11, 'bold')).pack(side=tk.LEFT, padx=(0, 5))
        
        self.tong_hop_search_entry = tk.Entry(search_frame, font=('Arial', 12), width=30, 
                                              bg='#FFFFFF', fg='#000000', relief=tk.SOLID, bd=2)
        self.tong_hop_search_entry.pack(side=tk.LEFT, padx=5, ipady=5)
        self.tong_hop_search_entry.bind('<Return>', self.on_tong_hop_search)
        self.tong_hop_search_entry.bind('<KeyRelease>', self.on_tong_hop_search_keyrelease)
        
        search_btn = tk.Button(search_frame, text="Tìm", command=self.on_tong_hop_search,
                              bg='#2196F3', fg='white', font=('Arial', 10, 'bold'), 
                              width=8, relief=tk.RAISED, bd=2, cursor='hand2')
        search_btn.pack(side=tk.LEFT, padx=5)
        
        # Nút Tải file excel tổng hợp
        export_btn = tk.Button(button_frame, text="Tải file excel tổng hợp", 
                              command=self.export_tong_hop_excel,
                              bg='#4CAF50', fg='white', font=('Arial', 12, 'bold'), 
                              width=25, height=2, relief=tk.RAISED, bd=2, cursor='hand2')
        export_btn.pack(side=tk.RIGHT, padx=10)
        
        # Bảng tổng hợp
        tonghop_table_frame = tk.Frame(tonghop_main_frame, bg=bg_color)
        tonghop_table_frame.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar cho bảng tổng hợp
        tonghop_scrollbar_y = tk.Scrollbar(tonghop_table_frame, orient=tk.VERTICAL, bg='#E0E0E0', troughcolor=bg_color)
        tonghop_scrollbar_x = tk.Scrollbar(tonghop_table_frame, orient=tk.HORIZONTAL, bg='#E0E0E0', troughcolor=bg_color)
        
        # Cột cho bảng tổng hợp
        tonghop_columns = ('N/X', 'Số phiếu', 'Ngày', 'Vị trí mới', 'ISBN', 'Tựa', 'Tồn thực tế', 'Số thùng', 'Tình trạng', 'Ghi chú', 'Note thùng')
        
        # Tạo Treeview cho tổng hợp với tối ưu hiệu suất
        self.tong_hop_tree = ttk.Treeview(tonghop_table_frame, columns=tonghop_columns, show='headings', 
                                          yscrollcommand=tonghop_scrollbar_y.set, xscrollcommand=tonghop_scrollbar_x.set,
                                          height=20, style='Treeview')
        
        # Tối ưu hiệu suất: tắt một số tính năng không cần thiết
        # Không cần selectmode vì không có selection logic phức tạp
        
        # Cấu hình các cột
        column_widths = {'N/X': 250, 'Số phiếu': 120, 'Ngày': 100, 'Vị trí mới': 120, 'ISBN': 150, 
                        'Tựa': 250, 'Tồn thực tế': 100, 'Số thùng': 120, 'Tình trạng': 100, 'Ghi chú': 200, 'Note thùng': 200}
        
        for col in tonghop_columns:
            self.tong_hop_tree.heading(col, text=col)
            # Tối ưu: không stretch, không minwidth để tăng tốc độ render và scroll
            self.tong_hop_tree.column(col, width=column_widths.get(col, 100), anchor='w', 
                                     stretch=False, minwidth=0)
        
        # Tag để highlight dòng tìm thấy
        self.tong_hop_tree.tag_configure('search_highlight', background='#FFF9C4')  # Màu vàng nhạt
        
        # Tối ưu scrollbar để scroll mượt hơn - không có delay
        # Sử dụng trực tiếp method của treeview để giảm overhead
        tonghop_scrollbar_y.config(command=self.tong_hop_tree.yview)
        tonghop_scrollbar_x.config(command=self.tong_hop_tree.xview)
        
        self.tong_hop_tree.grid(row=0, column=0, sticky='nsew')
        tonghop_scrollbar_y.grid(row=0, column=1, sticky='ns')
        tonghop_scrollbar_x.grid(row=1, column=0, sticky='ew')
        
        tonghop_table_frame.grid_rowconfigure(0, weight=1)
        tonghop_table_frame.grid_columnconfigure(0, weight=1)
        
        # Bind events để cho phép chỉnh sửa và xóa dòng trong tab Tổng hợp
        self.tong_hop_tree.bind('<Double-1>', self.on_tong_hop_item_click)
        self.tong_hop_tree.bind('<Button-1>', self.on_tong_hop_item_click)
        self.tong_hop_tree.bind('<Delete>', self.on_tong_hop_delete)
        self.tong_hop_tree.bind('<Key-Delete>', self.on_tong_hop_delete)
    
    def get_all_box_numbers(self):
        """Lấy danh sách tất cả mã thùng từ dữ liệu đầu vào"""
        if self.df is None or self.df.empty:
            return set()
        
        # Tìm cột số thùng
        box_col = None
        for col in self.df.columns:
            col_lower = str(col).lower().strip()
            if ('số thùng' in col_lower or 'so thung' in col_lower or 
                col_lower == 'thùng' or col_lower == 'thung'):
                box_col = col
                break
        
        if box_col is None:
            return set()
        
        # Lấy tất cả giá trị số thùng, loại bỏ NaN và chuyển thành string
        box_numbers = self.df[box_col].dropna().astype(str).str.strip()
        # Loại bỏ các giá trị rỗng
        box_numbers = box_numbers[box_numbers != '']
        return set(box_numbers.unique())
    
    def validate_vi_tri_moi(self):
        """Kiểm tra mã thùng mới có trùng với dữ liệu đầu vào không"""
        vi_tri_moi = self.vi_tri_moi_var.get().strip()
        
        # Nếu rỗng thì không cần kiểm tra
        if not vi_tri_moi:
            return True
        
        # Lấy danh sách tất cả mã thùng từ dữ liệu đầu vào
        existing_box_numbers = self.get_all_box_numbers()
        
        # Kiểm tra xem mã thùng mới có trùng với mã thùng nào trong dữ liệu không
        if vi_tri_moi in existing_box_numbers:
            messagebox.showerror(
                "Lỗi", 
                f"Mã thùng mới '{vi_tri_moi}' đã tồn tại trong dữ liệu đầu vào!\n\n"
                f"Vui lòng nhập mã thùng khác với các mã thùng hiện có.\n\n"
                f"Các mã thùng hiện có: {', '.join(sorted(existing_box_numbers)[:10])}"
                + (f" và {len(existing_box_numbers) - 10} mã khác..." if len(existing_box_numbers) > 10 else "")
            )
            # Xóa giá trị và focus lại vào ô nhập
            self.vi_tri_moi_var.set('')
            self.vi_tri_moi_entry.focus()
            return False
        
        return True
    
    def load_box_data(self):
        """Load dữ liệu của thùng được nhập"""
        so_thung = self.so_thung_var.get().strip()
        if not so_thung:
            messagebox.showwarning("Cảnh báo", "Vui lòng nhập số thùng!")
            return
        
        # Kiểm tra nếu đang có dữ liệu đã quét
        if self.scanned_items and len(self.scanned_items) > 0:
            # Đếm số item có tồn thực tế đã nhập
            items_with_data = sum(1 for item in self.scanned_items.values() if item.get('ton_thuc_te', '').strip())
            
            if items_with_data > 0:
                # Hiển thị dialog hỏi người dùng
                result = messagebox.askyesnocancel(
                    "Cảnh báo", 
                    f"Bạn đang có {len(self.scanned_items)} item đã quét, trong đó {items_with_data} item đã nhập tồn thực tế.\n\n"
                    "Bạn có muốn lưu dữ liệu trước khi load thùng mới không?\n\n"
                    "• Có: Lưu dữ liệu hiện tại\n"
                    "• Không: Xóa dữ liệu và load thùng mới\n"
                    "• Hủy: Không làm gì"
                )
                
                if result is True:  # Người dùng chọn "Có" - Lưu
                    self.save_data()
                    # Sau khi lưu, tiếp tục load thùng mới
                elif result is False:  # Người dùng chọn "Không" - Reset
                    # Reset dữ liệu
                    self.scanned_items = {}
                    self.clear_table()
                    # Xóa phần đếm Số tựa - không còn sử dụng
                    # Tiếp tục load thùng mới
                else:  # Người dùng chọn "Hủy" - Không làm gì
                    return
        
        # Showroom: Chỉ lưu số thùng người dùng nhập, không cần load từ Excel
        self.current_box_number = so_thung
        
        # Showroom: Set current_box_data = self.df (toàn bộ dữ liệu Excel)
        if self.df is not None and not self.df.empty:
            self.current_box_data = self.df.copy()
            # Hiển thị số tựa tổng từ Excel
            # Xóa phần đếm Số tựa - không còn sử dụng
        else:
            self.current_box_data = None
            self.so_tua_var.set("0")
            messagebox.showwarning("Cảnh báo", "Chưa load dữ liệu Excel. Vui lòng đảm bảo file DuLieuDauVaoShowroom.xlsx có trong thư mục.")
            return
        
        # Cập nhật "Đã quét" sau khi load: số dòng trong Tổng hợp + số dòng trong Kiểm kê
        self.update_da_quet_counter()
        
        # Focus vào ô nhập ISBN để sẵn sàng quét
        if hasattr(self, 'isbn_entry'):
            self.isbn_entry.focus()
    
    def count_valid_scanned_isbns(self):
        """Đếm số ISBN hợp lệ (tồn tại trong Excel) đã được quét - không đếm ISBN không tồn tại"""
        if not self.scanned_items or self.df is None:
            return 0
        
        count = 0
        if 'isbn' not in self.df.columns:
            return 0
        
        for isbn, item_data in self.scanned_items.items():
            # Bỏ qua nếu là ISBN không hợp lệ (không tồn tại trong Excel)
            if item_data.get('is_invalid_isbn', False) or item_data.get('is_new_isbn_not_in_data', False):
                continue
            
            # Nếu ISBN không bị đánh dấu là invalid, coi như hợp lệ
            count += 1
        
        return count
    
    def count_scanned_titles_for_box(self, so_thung):
        """Đếm số tựa đã quét từ tab Tổng hợp cho một thùng cụ thể - chỉ đếm ISBN tồn tại"""
        if not so_thung or not self.tong_hop_data:
            return 0
        
        so_thung_clean = str(so_thung).strip()
        count = 0
        
        for data in self.tong_hop_data:
            # So sánh số thùng (có thể là 'Số thùng' hoặc 'Vị trí mới') - không phân biệt chữ hoa/thường
            so_thung_in_data = str(data.get('Số thùng', '')).strip()
            vi_tri_moi_in_data = str(data.get('Vị trí mới', '')).strip()
            
            # Kiểm tra nếu số thùng khớp (so sánh không phân biệt chữ hoa/thường)
            if (so_thung_in_data.lower() == so_thung_clean.lower() or 
                vi_tri_moi_in_data.lower() == so_thung_clean.lower()):
                
                # QUAN TRỌNG: Chỉ đếm các ISBN tồn tại (có _is_valid_isbn = True)
                # ISBN không tồn tại không được tính vào số tựa đã quét
                is_valid = data.get('_is_valid_isbn', True)  # Mặc định là True để tương thích với dữ liệu cũ
                
                # Nếu không có field _is_valid_isbn, coi như hợp lệ (tương thích ngược)
                if is_valid is not False:  # Chỉ đếm nếu is_valid = True hoặc không có field
                    count += 1
        
        return count
    
    def count_all_rows_for_box_in_tong_hop(self, so_thung):
        """Đếm TẤT CẢ số dòng đã lưu trong Tổng hợp cho một thùng cụ thể (bao gồm cả ISBN không tồn tại)"""
        if not so_thung or not self.tong_hop_data:
            return 0
        
        so_thung_clean = str(so_thung).strip()
        count = 0
        
        for data in self.tong_hop_data:
            # So sánh số thùng (có thể là 'Số thùng' hoặc 'Vị trí mới') - không phân biệt chữ hoa/thường
            so_thung_in_data = str(data.get('Số thùng', '')).strip()
            vi_tri_moi_in_data = str(data.get('Vị trí mới', '')).strip()
            
            # Kiểm tra nếu số thùng khớp (so sánh không phân biệt chữ hoa/thường)
            if (so_thung_in_data.lower() == so_thung_clean.lower() or 
                vi_tri_moi_in_data.lower() == so_thung_clean.lower()):
                # Đếm TẤT CẢ các dòng, không phân biệt ISBN hợp lệ hay không
                count += 1
        
        return count
    
    def update_da_quet_counter(self):
        """Cập nhật số "Đã quét": chỉ đếm số dòng hiện tại trong bảng Kiểm kê"""
        if not hasattr(self, 'so_tua_da_quet_var') or not self.so_tua_da_quet_var:
            return
        
        # Đếm số dòng hiện tại trong table Kiểm kê
        so_dong_trong_kiem_ke = len(self.tree.get_children()) if hasattr(self, 'tree') else 0
        
        # Chỉ hiển thị số dòng đang quét trong Kiểm kê
        self.so_tua_da_quet_var.set(str(so_dong_trong_kiem_ke))
    
    def on_so_thung_changed(self, event=None):
        """Xử lý khi Số thùng thay đổi - cập nhật lại "Đã quét" để đếm lại số dòng trong Tổng hợp cho số thùng mới"""
        # Chỉ cập nhật nếu UI đã được khởi tạo hoàn chỉnh
        if hasattr(self, 'so_tua_da_quet_var') and self.so_tua_da_quet_var:
            # Cập nhật current_box_number để đồng bộ với giá trị mới
            so_thung_moi = self.so_thung_var.get().strip() if hasattr(self, 'so_thung_var') and self.so_thung_var else ''
            if so_thung_moi:
                self.current_box_number = so_thung_moi
            # Cập nhật lại "Đã quét"
            self.update_da_quet_counter()
    
    def on_so_thung_focus_in(self, event=None):
        """Xử lý khi focus vào input số thùng - kiểm tra nếu đã có dữ liệu quét thì chặn sửa"""
        if self.scanned_items and len(self.scanned_items) > 0:
            # Đã có dữ liệu quét, không cho phép sửa
            # Đảm bảo input ở chế độ readonly với màu nền đẹp hơn (xanh nhạt, không phải đen)
            if hasattr(self, 'so_thung_entry'):
                self.so_thung_entry.config(state='readonly', bg='#E8F4F8', fg='#1565C0', relief=tk.SOLID, bd=1)
            # Hiển thị cảnh báo
            messagebox.showwarning(
                "Cảnh báo", 
                "Không thể sửa số thùng khi đã có dữ liệu đã quét!\n\n"
                "Vui lòng SAVE hoặc RESET trước khi nhập số thùng mới."
            )
            # Focus ra khỏi input số thùng, chuyển sang input ISBN
            if hasattr(self, 'isbn_entry'):
                self.isbn_entry.focus()
    
    def is_isbn_in_input_data(self, isbn):
        """Kiểm tra xem ISBN có tồn tại trong dữ liệu đầu vào (tong_hop_data) không - không kiểm tra số thùng"""
        if not isbn or not self.tong_hop_data:
            return False
        
        try:
            isbn_clean = str(isbn).strip()
            isbn_clean_digits = ''.join(filter(str.isdigit, isbn_clean))
            
            for data in self.tong_hop_data:
                isbn_in_data = str(data.get('ISBN', '')).strip()
                
                # So sánh nhanh: khớp chính xác trước
                if isbn_in_data == isbn_clean:
                    return True
                
                # So sánh với digits
                isbn_in_data_digits = ''.join(filter(str.isdigit, isbn_in_data))
                if isbn_in_data_digits and isbn_clean_digits and isbn_in_data_digits == isbn_clean_digits:
                    return True
                
                # So sánh với endswith/startswith (chậm hơn, chỉ khi cần)
                if (isbn_in_data.endswith(isbn_clean) or isbn_clean.endswith(isbn_in_data)):
                    return True
            
            return False
        except Exception as e:
            # Nếu có lỗi, trả về False để không block quét
            print(f"Lỗi khi kiểm tra ISBN trong dữ liệu đầu vào: {str(e)}")
            return False
    
    def is_isbn_already_scanned(self, isbn, so_thung):
        """Kiểm tra xem ISBN đã được quét và lưu trong tab Tổng hợp cho thùng này chưa - tối ưu"""
        if not isbn or not so_thung or not self.tong_hop_data:
            return False
        
        try:
            isbn_clean = str(isbn).strip()
            isbn_clean_digits = ''.join(filter(str.isdigit, isbn_clean))
            so_thung_clean = str(so_thung).strip()
            
            # Tối ưu: chỉ kiểm tra các item có số thùng khớp trước (không phân biệt chữ hoa/thường)
            so_thung_clean_lower = so_thung_clean.lower()
            for data in self.tong_hop_data:
                # Kiểm tra số thùng trước (nhanh hơn)
                so_thung_in_data = str(data.get('Số thùng', '')).strip()
                vi_tri_moi_in_data = str(data.get('Vị trí mới', '')).strip()
                
                # Nếu số thùng không khớp, bỏ qua ngay (so sánh không phân biệt chữ hoa/thường)
                if (so_thung_in_data.lower() != so_thung_clean_lower and 
                    vi_tri_moi_in_data.lower() != so_thung_clean_lower):
                    continue
                
                # Kiểm tra ISBN chỉ khi số thùng khớp
                isbn_in_data = str(data.get('ISBN', '')).strip()
                
                # So sánh nhanh: khớp chính xác trước
                if isbn_in_data == isbn_clean:
                    return True
                
                # So sánh với digits
                isbn_in_data_digits = ''.join(filter(str.isdigit, isbn_in_data))
                if isbn_in_data_digits and isbn_clean_digits and isbn_in_data_digits == isbn_clean_digits:
                    return True
                
                # So sánh với endswith/startswith (chậm hơn, chỉ khi cần)
                if (isbn_in_data.endswith(isbn_clean) or isbn_clean.endswith(isbn_in_data)):
                    return True
            
            return False
        except Exception as e:
            # Nếu có lỗi, trả về False để không block quét
            print(f"Lỗi khi kiểm tra ISBN đã quét: {str(e)}")
            return False
    
    def on_isbn_entered(self, event=None):
        """Xử lý khi nhập/quét ISBN - tối ưu để tránh freeze"""
        try:
            isbn = self.isbn_entry.get().strip()
            if not isbn:
                return
            
            # Showroom: Kiểm tra dữ liệu Excel đã load chưa
            if self.df is None or self.df.empty:
                # Sử dụng after để không block UI
                self.root.after(10, lambda: messagebox.showwarning("Cảnh báo", "Vui lòng đảm bảo file Excel đã được load!"))
                self.isbn_entry.delete(0, tk.END)
                return
            
            # Showroom: Kiểm tra số thùng đã nhập chưa
            so_thung_input = self.so_thung_var.get().strip() if hasattr(self, 'so_thung_var') and self.so_thung_var else ''
            if not so_thung_input:
                self.root.after(10, lambda: messagebox.showwarning("Cảnh báo", "Vui lòng nhập số thùng trước khi quét ISBN!"))
                self.isbn_entry.delete(0, tk.END)
                return
            
            # QUAN TRỌNG: Cập nhật current_box_number từ so_thung_var để đồng bộ
            # Đảm bảo khi quét ISBN, current_box_number luôn đúng với giá trị người dùng nhập
            self.current_box_number = so_thung_input
            
            # Lưu giá trị số thùng ban đầu khi quét ISBN đầu tiên (nếu chưa có dữ liệu)
            if not self.scanned_items:
                self.so_thung_original_value = so_thung_input
                # Disable input số thùng khi đã bắt đầu quét (không cho phép sửa)
                # Set màu nền đẹp hơn khi readonly (xanh nhạt, không phải đen)
                if hasattr(self, 'so_thung_entry'):
                    self.so_thung_entry.config(state='readonly', bg='#E8F4F8', fg='#1565C0', relief=tk.SOLID, bd=1)
            
            # Showroom: Tìm ISBN trong toàn bộ self.df (không cần tìm theo số thùng)
            # Tìm tựa trong toàn bộ dữ liệu Excel - tối ưu với vectorization
            if 'isbn' in self.df.columns:
                isbn_clean = str(isbn).strip()
                isbn_clean_digits = ''.join(filter(str.isdigit, isbn_clean))
                matched_row = None
                
                # Showroom: Tìm trong toàn bộ self.df (không cần tìm theo số thùng)
                # Tối ưu: sử dụng vectorization thay vì iterrows() (nhanh hơn nhiều)
                try:
                    # Chuyển đổi ISBN sang số để so sánh nhanh hơn
                    isbn_col = self.df['isbn'].astype(str).str.strip()
                    
                    # Tìm khớp chính xác trước (nhanh nhất)
                    exact_match = isbn_col == isbn_clean
                    if exact_match.any():
                        matched_row = self.df[exact_match].iloc[0]
                    else:
                        # Tìm khớp với digits
                        isbn_col_digits = isbn_col.str.replace(r'\D', '', regex=True)
                        digit_match = isbn_col_digits == isbn_clean_digits
                        if digit_match.any():
                            matched_row = self.df[digit_match].iloc[0]
                        else:
                            # Tìm khớp với endswith/startswith (chậm hơn nhưng cần thiết)
                            for idx, row in self.df.iterrows():
                                row_isbn = str(row.get('isbn', '')).strip()
                                row_isbn_clean = ''.join(filter(str.isdigit, row_isbn))
                                
                                if (row_isbn == isbn_clean or 
                                    row_isbn.endswith(isbn_clean) or 
                                    isbn_clean.endswith(row_isbn) or
                                    (row_isbn_clean and isbn_clean_digits and row_isbn_clean == isbn_clean_digits)):
                                    matched_row = row
                                    break
                except Exception as e:
                    # Fallback về cách cũ nếu vectorization lỗi
                    print(f"Lỗi khi tìm ISBN với vectorization: {str(e)}")
                    for idx, row in self.df.iterrows():
                        row_isbn = str(row.get('isbn', '')).strip()
                        row_isbn_clean = ''.join(filter(str.isdigit, row_isbn))
                        
                        if (row_isbn == isbn_clean or 
                            row_isbn.endswith(isbn_clean) or 
                            isbn_clean.endswith(row_isbn) or
                            (row_isbn_clean and isbn_clean_digits and row_isbn_clean == isbn_clean_digits)):
                            matched_row = row
                            break
                
                # Showroom: Nếu không tìm thấy ISBN, coi như ISBN không tồn tại
                # BỎ logic kiểm tra ISBN đã quét trong Tổng hợp - cho phép quét lại và cộng dồn
                is_invalid_isbn = False
                
                if matched_row is None:
                    is_invalid_isbn = True
                
                # Đảm bảo pandas đã được import
                if self.pd is None:
                    try:
                        import pandas as pd
                        self.pd = pd
                    except ImportError:
                        messagebox.showerror("Lỗi", "Không thể import pandas!")
                        return
                
                pd = self.pd  # Alias để dùng trong hàm này
                
                # Kiểm tra xem ISBN có tồn tại trong dữ liệu đầu vào không
                isbn_not_in_input_data = False
                if matched_row is None:
                    # Chỉ kiểm tra nếu ISBN không tồn tại trong thùng
                    isbn_not_in_input_data = not self.is_isbn_in_input_data(isbn_clean)
                
                # Kiểm tra xem ISBN đã được quét trước đó chưa (để cho phép tăng số lượng)
                # Nếu đã quét rồi và ISBN tồn tại trong thùng, luôn cho phép tăng số lượng, không cần kiểm tra is_over_limit
                is_already_scanned = isbn_clean in self.scanned_items
                
                # Kiểm tra nếu ISBN không hợp lệ (không thuộc thùng, không tồn tại trong dữ liệu, hoặc quét quá số tựa)
                # Trong trường hợp này, coi như tạo dòng trống mới, chỉ điền số thùng đang load vào cột "Số thùng"
                # KHÔNG thêm thông báo lỗi vào Ghi chú
                # LƯU Ý: 
                # - Nếu ISBN đã được quét trước đó (is_already_scanned) và tồn tại trong thùng (matched_row is not None),
                #   thì không coi là quá giới hạn vì đây là quét lại để tăng số lượng
                # - QUAN TRỌNG: Chỉ tìm ISBN trong thùng hiện tại, không lấy dữ liệu từ thùng khác
                # - Nếu ISBN không thuộc thùng hiện tại, coi như không tồn tại
                # - Chỉ đếm các ISBN hợp lệ (tồn tại trong thùng hiện tại) khi kiểm tra giới hạn
                should_treat_as_invalid = False
                if matched_row is None:
                    should_treat_as_invalid = True  # ISBN không tồn tại trong thùng hiện tại
                elif is_invalid_isbn:
                    should_treat_as_invalid = True  # ISBN không hợp lệ
                # Nếu ISBN tồn tại trong thùng hiện tại (matched_row is not None), luôn coi là hợp lệ
                # Chỉ đếm các ISBN hợp lệ (tồn tại trong thùng hiện tại)
                # ISBN không tồn tại sẽ không được đếm vào số tựa
                
                if should_treat_as_invalid:
                    # Để trống các cột: Tựa, Tồn thực tế, Tồn tựa trong thùng, Tình trạng, Ghi chú
                    # Tự động lấy số thùng từ input "Số thùng" ở trên
                    tua = ''
                    ton_trong_thung = 0
                    ton_thuc_te_value = ''  # Để trống thay vì mặc định là 1
                    tinh_trang_value = ''  # Bỏ hiển thị trạng thái
                    ghi_chu_value = ''  # Để trống, không thêm thông báo lỗi
                    # Tự động lấy số thùng từ input "Số thùng" ở trên
                    so_thung_hien_thi = str(self.current_box_number) if self.current_box_number else ''
                    is_existing_item = False
                else:
                    # Lấy thông tin từ matched_row (ISBN hợp lệ)
                    # Tìm cột 'tua' (có thể là 'tua' sau khi mapping hoặc tên gốc)
                    tua = ''
                    for col in matched_row.index:
                        col_lower = str(col).lower().strip()
                        if 'tựa' in col_lower or 'tua' in col_lower or 'tên' in col_lower or 'titles' in col_lower:
                            tua = str(matched_row[col]) if pd.notna(matched_row[col]) else ''
                            break
                    
                    # Tìm cột 'ton_tung_tua' (có thể là 'ton_tung_tua' sau khi mapping hoặc tên gốc)
                    ton_trong_thung = 0
                    for col in matched_row.index:
                        col_lower = str(col).lower().strip()
                        if ('tồn' in col_lower and 'tựa' in col_lower) or ('ton' in col_lower and 'tua' in col_lower) or 'qty tựa trong thùng' in col_lower or 'qty tua trong thung' in col_lower:
                            ton_trong_thung = matched_row[col] if pd.notna(matched_row[col]) else 0
                            try:
                                ton_trong_thung = int(float(ton_trong_thung))  # Chuyển thành số nguyên
                            except:
                                ton_trong_thung = 0
                            break
                    
                    so_thung = self.current_box_number
                    
                    # Kiểm tra xem đã quét chưa để tự động tăng số lượng
                    ton_thuc_te_value = '1'  # Mặc định là 1 khi quét lần đầu
                    ghi_chu_value = ''
                    tinh_trang_value = ''
                    is_existing_item = False
                
                # Xử lý item đã tồn tại: nếu ISBN đã được quét (cả ISBN hợp lệ và không tồn tại), cho phép tăng số lượng
                # (không cần kiểm tra is_over_limit vì đây là quét lại để tăng số lượng)
                old_stt = None  # Lưu STT cũ khi cộng dồn
                if is_already_scanned:
                    # ISBN đã tồn tại - tăng số lượng lên 1
                    is_existing_item = True
                    item_id_old = self.scanned_items[isbn_clean]['item_id']
                    
                    # Lấy giá trị "Tồn thực tế" từ cả scanned_items và tree để đảm bảo chính xác
                    old_ton_thuc_te_from_items = self.scanned_items[isbn_clean].get('ton_thuc_te', '')
                    old_values = list(self.tree.item(item_id_old, 'values'))
                    old_ton_thuc_te_from_tree = old_values[3] if len(old_values) > 3 else ''  # Tồn thực tế ở index 3
                    
                    # Giữ lại STT cũ để dùng lại khi thêm lại dòng
                    old_stt = old_values[0] if len(old_values) > 0 else None
                    
                    # Chuyển đổi sang string và strip để so sánh
                    # QUAN TRỌNG: Kiểm tra cả None và empty string, nhưng không bỏ qua giá trị '0'
                    old_ton_thuc_te_from_items_str = ''
                    if old_ton_thuc_te_from_items is not None:
                        old_ton_thuc_te_from_items_str = str(old_ton_thuc_te_from_items).strip()
                    
                    old_ton_thuc_te_from_tree_str = ''
                    if old_ton_thuc_te_from_tree is not None:
                        old_ton_thuc_te_from_tree_str = str(old_ton_thuc_te_from_tree).strip()
                    
                    # Ưu tiên lấy từ scanned_items (chính xác nhất), nếu không có hoặc rỗng thì lấy từ tree
                    # Kiểm tra bằng cách thử parse, không chỉ kiểm tra empty string
                    old_ton_thuc_te = ''
                    
                    # Thử parse từ scanned_items trước (ưu tiên)
                    if old_ton_thuc_te_from_items_str != '':
                        try:
                            test_num = float(old_ton_thuc_te_from_items_str)
                            old_ton_thuc_te = old_ton_thuc_te_from_items_str
                        except (ValueError, TypeError):
                            pass
                    
                    # Nếu không có từ scanned_items, thử từ tree
                    if not old_ton_thuc_te and old_ton_thuc_te_from_tree_str != '':
                        try:
                            test_num = float(old_ton_thuc_te_from_tree_str)
                            old_ton_thuc_te = old_ton_thuc_te_from_tree_str
                        except (ValueError, TypeError):
                            pass
                    
                    try:
                        # Tăng số lượng lên 1
                        if old_ton_thuc_te:
                            # Parse giá trị và tăng lên 1
                            old_ton_thuc_te_num = int(float(old_ton_thuc_te))
                            ton_thuc_te_value = str(old_ton_thuc_te_num + 1)
                        else:
                            # Nếu không có giá trị, mặc định là 1
                            ton_thuc_te_value = '1'
                    except (ValueError, TypeError) as e:
                        # Nếu không parse được, mặc định là 1
                        ton_thuc_te_value = '1'
                    
                    # Showroom: Giữ lại giá trị Ghi chú và Tình trạng cũ
                    tinh_trang_value = old_values[6] if len(old_values) > 6 else ''  # Tình trạng ở index 6
                    ghi_chu_value = old_values[7] if len(old_values) > 7 else ''  # Ghi chú ở index 7
                    
                    # Xóa highlight cũ trước khi xóa item để tránh lỗi khi click vào highlight
                    try:
                        self.remove_error_highlights(item_id_old)
                    except:
                        pass
                    
                    # Xóa item cũ
                    try:
                        self.tree.delete(item_id_old)
                    except:
                        pass  # Item có thể đã bị xóa rồi
                    del self.scanned_items[isbn_clean]
                    
                    # Nếu là ISBN không tồn tại, tự động cập nhật "Tồn tựa trong thùng" bằng với "Tồn thực tế" mới
                    if should_treat_as_invalid:
                        try:
                            ton_thuc_te_num = int(float(ton_thuc_te_value)) if ton_thuc_te_value else 0
                            if ton_thuc_te_num > 0:
                                ton_trong_thung = ton_thuc_te_num
                        except (ValueError, TypeError):
                            pass
                
                # Thêm vào bảng với đầy đủ thông tin
                # Chuyển ton_trong_thung thành số nguyên để hiển thị
                # Nếu ISBN không hợp lệ, để trống thay vì hiển thị 0 (trừ khi đã được cộng dồn)
                if should_treat_as_invalid:
                    # Nếu đã được cộng dồn (is_existing_item = True), hiển thị giá trị đã cập nhật
                    if is_existing_item and ton_trong_thung > 0:
                        ton_trong_thung_display = ton_trong_thung
                    else:
                        ton_trong_thung_display = ''
                else:
                    ton_trong_thung_display = int(ton_trong_thung) if ton_trong_thung else 0
                
                # Chỉ xử lý vi_tri_moi cho ISBN hợp lệ
                if not should_treat_as_invalid:
                    # Kiểm tra nếu có "Thùng / vị trí mới" thì dùng giá trị đó, không thì dùng số thùng hiện tại
                    vi_tri_moi = self.vi_tri_moi_var.get().strip()
                    
                    # Validation: Mã thùng mới phải khác với tất cả mã thùng trong dữ liệu đầu vào
                    if vi_tri_moi:
                        # Kiểm tra lại một lần nữa khi quét ISBN (để đảm bảo)
                        existing_box_numbers = self.get_all_box_numbers()
                        if vi_tri_moi in existing_box_numbers:
                            # Fix closure issue: capture giá trị vào biến local
                            vi_tri_msg = vi_tri_moi
                            existing_list = ', '.join(sorted(existing_box_numbers)[:10])
                            existing_count = len(existing_box_numbers)
                            existing_suffix = f" và {existing_count - 10} mã khác..." if existing_count > 10 else ""
                            self.root.after(10, lambda v=vi_tri_msg, e=existing_list, s=existing_suffix: messagebox.showerror(
                                "Lỗi", 
                                f"Mã thùng mới '{v}' đã tồn tại trong dữ liệu đầu vào!\n\n"
                                f"Vui lòng nhập mã thùng khác với các mã thùng hiện có.\n\n"
                                f"Các mã thùng hiện có: {e}{s}"
                            ))
                            self.isbn_entry.delete(0, tk.END)
                            return
                        
                        so_thung_hien_thi = vi_tri_moi
                    else:
                        so_thung_hien_thi = so_thung
                # Nếu ISBN không hợp lệ, so_thung_hien_thi đã được set thành '' ở trên
                
                # Showroom: Đảm bảo thứ tự đúng với columns: Số thứ tự, ISBN, Tựa, Tồn thực tế, Số thùng, Tồn tựa trong thùng, Ghi chú, Xóa
                # Tính số thứ tự: nếu là cộng dồn (đã có STT cũ), giữ nguyên STT cũ, nếu không thì tính mới
                if old_stt is not None:
                    try:
                        so_thu_tu = int(old_stt)  # Giữ nguyên STT cũ khi cộng dồn (chuyển sang int để so sánh)
                    except (ValueError, TypeError):
                        so_thu_tu = len(self.tree.get_children()) + 1
                else:
                    so_thu_tu = len(self.tree.get_children()) + 1  # Tính STT mới cho item mới
                
                # Đảm bảo ton_thuc_te_value là string
                # Nếu ISBN không hợp lệ, giữ nguyên giá trị rỗng
                # Nếu ISBN hợp lệ, mặc định là '1' nếu rỗng
                if should_treat_as_invalid:
                    # Giữ nguyên giá trị rỗng cho ISBN không hợp lệ
                    ton_thuc_te_value = str(ton_thuc_te_value).strip() if ton_thuc_te_value is not None else ''
                else:
                    # Đối với ISBN hợp lệ, mặc định là '1' nếu rỗng
                    if not ton_thuc_te_value or str(ton_thuc_te_value).strip() == '':
                        ton_thuc_te_value = '1'
                    else:
                        ton_thuc_te_value = str(ton_thuc_te_value).strip()
                
                # Đảm bảo ISBN được format đúng và hiển thị
                isbn_display = str(isbn_clean).strip() if isbn_clean else ''
                if not isbn_display:
                    isbn_display = str(isbn).strip() if isbn else ''
                
                # Showroom: Lấy số thùng từ input "Số thùng"
                so_thung_hien_thi = str(self.so_thung_var.get().strip()) if self.so_thung_var.get().strip() else ''
                
                item_id = self.tree.insert('', tk.END, values=(
                    str(so_thu_tu),            # 0: Số thứ tự
                    isbn_display,               # 1: ISBN - đảm bảo hiển thị đúng
                    str(tua) if tua else '',   # 2: Tựa
                    ton_thuc_te_value,         # 3: Tồn thực tế - tự động điền 1 hoặc tăng lên (hoặc rỗng cho ISBN không hợp lệ)
                    so_thung_hien_thi,         # 4: Số thùng (lấy từ input "Số thùng")
                    str(ton_trong_thung_display) if ton_trong_thung_display != '' else '',  # 5: Tồn tựa trong thùng (rỗng cho ISBN không hợp lệ)
                    '',                         # 6: Tình trạng - để trống
                    ghi_chu_value if ghi_chu_value else '',              # 7: Ghi chú - để trống, người dùng tự nhập
                    'Xóa'                       # 8: Xóa - nút xóa dòng
                ), tags=('',))
                
                # Đảm bảo giá trị ISBN được hiển thị đúng ngay sau khi insert
                current_values_check = list(self.tree.item(item_id, 'values'))
                if len(current_values_check) >= 2 and current_values_check[1] != isbn_display:
                    current_values_check[1] = isbn_display
                    self.tree.item(item_id, values=current_values_check)
                
                # Đảm bảo giá trị được hiển thị đúng ngay sau khi insert
                # Cập nhật lại để đảm bảo đồng bộ
                current_values_check = list(self.tree.item(item_id, 'values'))
                if len(current_values_check) >= 4 and current_values_check[3] != ton_thuc_te_value:
                    current_values_check[3] = ton_thuc_te_value
                    self.tree.item(item_id, values=current_values_check)
                
                # Lưu thông tin
                # Lưu cả số thùng gốc và số thùng hiển thị (vị trí mới nếu có)
                # Lưu cả vi_tri_moi để có thể sử dụng khi lưu (nếu người dùng chỉnh sửa trực tiếp)
                vi_tri_moi_value = self.vi_tri_moi_var.get().strip()
                # Đảm bảo so_thung được định nghĩa cho cả trường hợp ISBN không hợp lệ
                # Khi ISBN không hợp lệ, lấy số thùng từ input "Số thùng" ở trên
                if should_treat_as_invalid:
                    so_thung_goc = str(self.current_box_number) if self.current_box_number else ''  # Lấy từ input "Số thùng"
                else:
                    so_thung_goc = so_thung if 'so_thung' in locals() else self.current_box_number
                
                # Nếu là ISBN không tồn tại và đã được cộng dồn, cập nhật ton_trong_thung bằng với ton_thuc_te
                if should_treat_as_invalid and is_existing_item:
                    try:
                        ton_thuc_te_num = int(float(ton_thuc_te_value)) if ton_thuc_te_value else 0
                        if ton_thuc_te_num > 0:
                            ton_trong_thung = ton_thuc_te_num
                    except (ValueError, TypeError):
                        pass
                
                self.scanned_items[isbn_clean] = {
                    'item_id': item_id,
                    'tua': tua,
                    'ton_thuc_te': ton_thuc_te_value,  # Lưu giá trị đã tự động điền hoặc đã tăng (có thể rỗng cho ISBN không hợp lệ)
                    'so_thung': so_thung_hien_thi,  # Lưu số thùng hiển thị (có thể là vị trí mới, hoặc rỗng cho ISBN không hợp lệ)
                    'so_thung_goc': so_thung_goc,  # Lưu số thùng gốc từ dữ liệu (hoặc rỗng cho ISBN không hợp lệ)
                    'vi_tri_moi': vi_tri_moi_value,  # Lưu giá trị từ ô "Thùng / vị trí mới" khi quét
                    'ton_trong_thung': ton_trong_thung,  # Có thể là 0 hoặc giá trị thực tế (hoặc bằng ton_thuc_te cho ISBN không tồn tại đã cộng dồn)
                    'tinh_trang': tinh_trang_value,  # Giữ lại tình trạng cũ nếu có (hoặc rỗng cho ISBN không hợp lệ)
                    'ghi_chu': ghi_chu_value,  # Giữ lại ghi chú cũ nếu có (hoặc rỗng cho ISBN không hợp lệ)
                    'is_invalid_isbn': should_treat_as_invalid,  # Đánh dấu ISBN không hợp lệ để cho phép edit cột Tựa
                    'is_new_isbn_not_in_data': isbn_not_in_input_data  # Đánh dấu ISBN không tồn tại trong cả thùng và dữ liệu đầu vào
                }
                
                # Tính lại STT cho tất cả các dòng để đảm bảo chính xác (đặc biệt khi cộng dồn)
                # Sắp xếp lại các dòng theo STT hiện tại, sau đó đánh số lại từ 1
                children = list(self.tree.get_children())
                
                # Lấy tất cả các dòng với STT của chúng
                items_with_stt = []
                for child in children:
                    child_values = list(self.tree.item(child, 'values'))
                    if len(child_values) >= 1:
                        try:
                            stt = int(child_values[0]) if child_values[0] else 9999
                        except (ValueError, TypeError):
                            stt = 9999
                        items_with_stt.append((stt, child, child_values))
                
                # Sắp xếp theo STT
                items_with_stt.sort(key=lambda x: x[0])
                
                # Đánh số lại STT từ 1 và cập nhật lại tree
                # Xóa tất cả và insert lại theo thứ tự đúng
                all_items_data = []
                for _, child, child_values in items_with_stt:
                    # Lưu lại tất cả thông tin cần thiết
                    isbn_val = child_values[1] if len(child_values) > 1 else ''
                    all_items_data.append((child_values, isbn_val, child))
                
                # Xóa tất cả items cũ
                for child in children:
                    self.tree.delete(child)
                
                # Insert lại theo thứ tự đúng với STT mới
                for idx, (child_values, isbn_val, old_child_id) in enumerate(all_items_data, start=1):
                    # Cập nhật STT mới
                    if len(child_values) >= 9:
                        child_values[0] = str(idx)
                    # Insert lại vào tree
                    new_item_id = self.tree.insert('', tk.END, values=child_values, tags=('',))
                    # Cập nhật item_id trong scanned_items nếu có
                    if isbn_val and isbn_val in self.scanned_items:
                        self.scanned_items[isbn_val]['item_id'] = new_item_id
                        # Nếu là dòng vừa cộng dồn, cập nhật item_id mới
                        if isbn_clean == isbn_val:
                            item_id = new_item_id
                
                # Cập nhật số dòng đã quét: số dòng trong Tổng hợp + số dòng hiện tại trong Kiểm kê
                self.update_da_quet_counter()
                
                # Scroll đến item để người dùng thấy được item vừa thêm/cập nhật
                self.tree.selection_set(item_id)
                self.tree.focus(item_id)
                self.tree.see(item_id)
                
                # Nếu là item mới (quét lần đầu), tự động focus vào ô "Tồn thực tế" để người dùng nhập và Enter
                # Áp dụng cho cả ISBN hợp lệ và không hợp lệ - người dùng cần điền tồn thực tế
                # Nếu là item đã tồn tại (quét lại), chỉ cộng dồn số lượng, không cần focus
                if not is_existing_item:
                    # Tự động focus vào ô "Tồn thực tế" để người dùng có thể chỉnh sửa và Enter
                    # Áp dụng cho cả ISBN hợp lệ và không hợp lệ
                    # Fix closure issue: capture item_id vào biến local
                    item_id_to_edit = item_id
                    self.root.after(100, lambda i=item_id_to_edit: self.auto_edit_ton_thuc_te(i))
                else:
                    # Nếu là item đã tồn tại (đã cộng dồn), đảm bảo giá trị được hiển thị đúng
                    # Cập nhật lại giá trị trong tree ngay lập tức để đảm bảo hiển thị đúng
                    current_values = list(self.tree.item(item_id, 'values'))
                    if len(current_values) >= 4:
                        # Đảm bảo giá trị Tồn thực tế đúng
                        if current_values[3] != ton_thuc_te_value:
                            current_values[3] = ton_thuc_te_value
                            self.tree.item(item_id, values=current_values)
                            # Force update để hiển thị ngay
                            self.root.update_idletasks()
                    
                    # Showroom: Bỏ logic kiểm tra chênh lệch và highlight
                
            else:
                self.root.after(10, lambda: messagebox.showerror("Lỗi", "Không tìm thấy cột 'ISBN' trong dữ liệu!"))
            
        except Exception as e:
            # Xử lý lỗi để tránh crash
            error_msg = str(e)
            print(f"Lỗi khi quét ISBN: {error_msg}")
            import traceback
            traceback.print_exc()
            # Fix closure issue: capture error_msg vào biến local
            self.root.after(10, lambda e=error_msg: messagebox.showerror("Lỗi", f"Lỗi khi quét ISBN: {e}"))
        
        # Clear ô nhập ISBN để sẵn sàng quét tiếp
        self.isbn_entry.delete(0, tk.END)
        self.isbn_entry.focus()
    
    def ensure_values_format(self, values):
        """Đảm bảo values có đủ 9 cột và cột cuối cùng là 'Xóa' - Showroom có cột Tình trạng để trống"""
        # Chuyển sang list nếu chưa phải
        values = list(values) if values else []
        # Đảm bảo có đủ 9 cột (có cột Tình trạng ở index 6)
        while len(values) < 9:
            values.append('')
        # Đảm bảo cột cuối cùng (index 8) là 'Xóa'
        values[8] = 'Xóa'
        return values
    
    def delete_row(self, item_id):
        """Xóa dòng khỏi bảng và scanned_items"""
        try:
            # Lấy giá trị từ item
            values = list(self.tree.item(item_id, 'values'))
            
            # Lấy ISBN từ cột index 1
            if len(values) < 2:
                return
            
            isbn = str(values[1]).strip()
            
            # Xóa highlight nếu có
            if item_id in self.error_highlights:
                try:
                    for widget in self.error_highlights[item_id]:
                        try:
                            widget.destroy()
                        except:
                            pass
                    del self.error_highlights[item_id]
                except:
                    pass
            
            # Xóa khỏi scanned_items
            if isbn in self.scanned_items:
                del self.scanned_items[isbn]
            
            # Xóa khỏi tree
            self.tree.delete(item_id)
            
            # Cập nhật lại số thứ tự (STT) của các dòng còn lại
            children = self.tree.get_children()
            for idx, child in enumerate(children, start=1):
                child_values = list(self.tree.item(child, 'values'))
                if len(child_values) >= 9:  # Đảm bảo có đủ cột bao gồm cột Xóa
                    child_values[0] = str(idx)  # Cập nhật STT
                    self.tree.item(child, values=child_values)
            
            # Cập nhật số dòng đã quét: số dòng trong Tổng hợp + số dòng hiện tại trong Kiểm kê
            self.update_da_quet_counter()
            
        except Exception as e:
            print(f"Lỗi khi xóa dòng: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def on_item_click(self, event):
        """Xử lý click để edit trực tiếp các cột có thể chỉnh sửa"""
        # Hủy edit cũ nếu có
        if self.edit_entry:
            self.finish_edit()
        
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        
        item = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        column_index = int(column.replace('#', '')) - 1
        
        if not item:
            return
        
        # Showroom: Nếu click vào cột "Xóa" (index 8), xóa dòng đó
        if column_index == 8:
            self.delete_row(item)
            return
        
        # Lấy giá trị hiện tại từ tree - đảm bảo lấy giá trị mới nhất
        values = list(self.tree.item(item, 'values'))
        # Showroom: Đảm bảo có đủ 9 cột (bao gồm cột Tình trạng ở index 6 và cột Xóa ở index 8)
        while len(values) < 9:
            values.append('')
        
        # Lấy giá trị từ đúng cột được click
        # QUAN TRỌNG: Với cột "Ghi chú" (index 7), đảm bảo chỉ lấy từ values[7], KHÔNG lấy từ note_thung_var
        if column_index == 7:  # Cột "Ghi chú"
            # Chỉ lấy giá trị từ values[7], không lấy từ bất kỳ nguồn nào khác
            # Đảm bảo không bị ảnh hưởng bởi note_thung_var
            current_value = ''
            if len(values) > 7:
                raw_value = values[7]
                if raw_value is not None:
                    current_value = str(raw_value).strip()
            
            # Nếu ISBN có trong scanned_items, lấy giá trị từ đó (nếu có), nhưng ưu tiên giá trị từ tree
            isbn_value = values[1] if len(values) > 1 else ''
            if isbn_value and isbn_value in self.scanned_items:
                saved_ghi_chu = self.scanned_items[isbn_value].get('ghi_chu', '')
                # Chỉ dùng giá trị từ scanned_items nếu giá trị trong tree rỗng
                if not current_value and saved_ghi_chu:
                    current_value = str(saved_ghi_chu).strip()
            
            # QUAN TRỌNG: Đảm bảo KHÔNG lấy từ note_thung_var - chỉ dùng giá trị đã lấy ở trên
            # Nếu current_value vẫn rỗng, để rỗng (không lấy từ note_thung_var)
        else:
            # Các cột khác: lấy bình thường từ values
            current_value = str(values[column_index]) if column_index < len(values) and values[column_index] is not None else ''
        
        # Kiểm tra xem ISBN có hợp lệ không bằng cách kiểm tra flag trong scanned_items
        # hoặc kiểm tra các giá trị trong row nếu không có trong scanned_items
        isbn_value = values[1] if len(values) > 1 else ''  # ISBN ở index 1
        is_invalid_isbn_row = False
        
        # Kiểm tra flag trong scanned_items trước (chính xác nhất)
        if isbn_value and isbn_value in self.scanned_items:
            is_invalid_isbn_row = self.scanned_items[isbn_value].get('is_invalid_isbn', False)
        else:
            # Nếu không có trong scanned_items, kiểm tra bằng cách xem các giá trị trong row
            # Nếu Tựa (2), Tồn thực tế (3), Số thùng (4), Tồn tựa trong thùng (5) đều trống/rỗng
            # thì có thể là ISBN không hợp lệ (không thuộc thùng, không tồn tại trong data, hoặc quét quá số tựa)
            if len(values) >= 6:
                tua = str(values[2]).strip() if len(values) > 2 else ''
                ton_thuc_te = str(values[3]).strip() if len(values) > 3 else ''
                so_thung = str(values[4]).strip() if len(values) > 4 else ''
                ton_trong_thung = str(values[5]).strip() if len(values) > 5 else ''
                # Nếu tất cả đều trống, có thể là ISBN không hợp lệ
                if not tua and not ton_thuc_te and not so_thung and (not ton_trong_thung or ton_trong_thung == '0' or ton_trong_thung == ''):
                    is_invalid_isbn_row = True
        
        # Cho phép edit: Tồn thực tế (3), Số thùng (4), Tồn tựa trong thùng (5), Ghi chú (7)
        # Cho phép edit Tựa (2) nếu ISBN không hợp lệ
        # Không cho edit: Số thứ tự (0), ISBN (1), Tình trạng (6) - chỉ đọc
        editable_columns = [3, 4, 5, 7]  # Tồn thực tế, Số thùng, Tồn tựa trong thùng, Ghi chú
        if is_invalid_isbn_row:
            editable_columns.append(2)  # Thêm cột Tựa nếu ISBN không hợp lệ
        
        if column_index not in editable_columns:
            return
        
        # Lấy vị trí của cell
        bbox = self.tree.bbox(item, column)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        # Tạo Entry widget để edit trực tiếp
        # QUAN TRỌNG: Với cột "Ghi chú", đảm bảo không bị ảnh hưởng bởi clipboard từ "Note thùng"
        if column_index == 7:
            # Xóa clipboard để tránh paste nhầm từ "Note thùng" khi nhấn Ctrl+V hoặc khi focus vào
            try:
                self.root.clipboard_clear()
            except:
                pass
        
        self.edit_entry = tk.Entry(self.tree, font=('Arial', 10), 
                                   relief=tk.FLAT, bd=0, bg='#FFFFFF', fg='#000000')
        # Insert giá trị đã được xử lý (chỉ từ values[7], không từ note_thung_var)
        self.edit_entry.insert(0, str(current_value))
        self.edit_entry.select_range(0, tk.END)
        self.edit_entry.place(x=x, y=y, width=width, height=height)
        self.edit_entry.focus()
        self.editing_item = item
        self.editing_column = column_index
        
        def finish_on_enter(event):
            self.finish_edit()
        
        def finish_on_focus_out(event):
            # Delay một chút để tránh conflict với click events
            self.root.after(100, self.finish_edit)
        
        self.edit_entry.bind('<Return>', finish_on_enter)
        self.edit_entry.bind('<FocusOut>', finish_on_focus_out)
        self.edit_entry.bind('<Escape>', lambda e: self.cancel_edit())
    
    def finish_edit(self):
        """Hoàn tất việc chỉnh sửa"""
        # Tránh xử lý 2 lần nếu đang trong quá trình xử lý
        if self.is_processing_edit:
            return
        
        if not self.edit_entry or not self.editing_item:
            return
        
        # Đặt flag để tránh xử lý lại
        self.is_processing_edit = True
        
        try:
            new_value = self.edit_entry.get().strip()
            item = self.editing_item
            column_index = self.editing_column
            
            # Kiểm tra xem item có còn tồn tại trong tree không (có thể đã bị xóa)
            try:
                # Lấy giá trị hiện tại và đảm bảo có đủ 8 cột
                values = list(self.tree.item(item, 'values'))
            except tk.TclError:
                # Item đã bị xóa, hủy edit
                self.cancel_edit()
                return
            
            values = self.ensure_values_format(values)
            
            isbn = values[1] if len(values) > 1 else ''  # ISBN ở index 1 (sau Số thứ tự)
            
            # KHÔNG tự động điền giá trị cho cột "Ghi chú" - để người dùng tự nhập
            # Nếu đang edit cột "Ghi chú", chỉ lấy giá trị từ tree, không lấy từ scanned_items
            
            # Xử lý theo từng cột
            if column_index == 2:  # Tựa (chỉ cho phép edit khi ISBN không hợp lệ)
                # Đảm bảo chỉ update cột "Tựa" (index 2), không động đến các cột khác
                # Lấy lại values từ tree để đảm bảo có giá trị mới nhất
                values = list(self.tree.item(item, 'values'))
                values = self.ensure_values_format(values)
                # Chỉ update cột "Tựa"
                values[2] = new_value
                if isbn in self.scanned_items:
                    self.scanned_items[isbn]['tua'] = new_value
                    # Lưu backup khi có thay đổi
                    self.save_backup_on_change()
                # Cập nhật tree với giá trị mới - chỉ cột "Tựa" được thay đổi
                self.tree.item(item, values=values)
                # Reset flag và cleanup
                self.is_processing_edit = False
                if self.edit_entry:
                    self.edit_entry.destroy()
                    self.edit_entry = None
                    self.editing_item = None
                return
            
            elif column_index == 3:  # Tồn thực tế - Showroom: Bỏ logic check chênh lệch
                values[3] = new_value  # Đảm bảo đúng index
                
                # Showroom: Lưu giá trị và tự động điền "Tồn tựa trong thùng" cho ISBN không tồn tại
                if isbn in self.scanned_items:
                    self.scanned_items[isbn]['ton_thuc_te'] = new_value
                    
                    # Kiểm tra xem ISBN có tồn tại trong dữ liệu không
                    is_new_isbn_not_in_data = self.scanned_items[isbn].get('is_new_isbn_not_in_data', False)
                    
                    # Nếu ISBN không tồn tại trong dữ liệu và có giá trị Tồn thực tế, tự động điền Tồn tựa trong thùng
                    if is_new_isbn_not_in_data and new_value:
                        try:
                            ton_thuc_te_num = float(new_value) if new_value else 0
                            if ton_thuc_te_num > 0:
                                # Tự động đặt ton_trong_thung = ton_thuc_te
                                ton_trong_thung_value = str(int(ton_thuc_te_num))
                                # Đảm bảo có đủ 9 cột
                                values = self.ensure_values_format(values)
                                values[5] = ton_trong_thung_value  # Cột 5: Tồn tựa trong thùng
                                
                                # Cập nhật trong scanned_items
                                self.scanned_items[isbn]['ton_trong_thung'] = ton_thuc_te_num
                        except (ValueError, TypeError):
                            # Nếu không parse được số, không làm gì
                            pass
                    
                    # Lưu backup khi có thay đổi
                    self.save_backup_on_change()
                    
                    # Đảm bảo có đủ 9 cột
                    values = self.ensure_values_format(values)
                    
                    # Đảm bảo Tình trạng và Ghi chú giữ nguyên (không tự động điền)
                    # Tình trạng để trống
                    if len(values) > 6:
                        # Giữ nguyên giá trị Tình trạng hiện tại (có thể rỗng)
                        pass
                    
                    # Cập nhật tree
                    self.tree.item(item, values=values)
            
            elif column_index == 4:  # Số thùng
                # Validation: Mã thùng mới phải khác với tất cả mã thùng trong dữ liệu đầu vào
                if new_value.strip():
                    existing_box_numbers = self.get_all_box_numbers()
                    if new_value.strip() in existing_box_numbers:
                        messagebox.showerror(
                            "Lỗi", 
                            f"Mã thùng '{new_value.strip()}' đã tồn tại trong dữ liệu đầu vào!\n\n"
                            f"Vui lòng nhập mã thùng khác với các mã thùng hiện có.\n\n"
                            f"Các mã thùng hiện có: {', '.join(sorted(existing_box_numbers)[:10])}"
                            + (f" và {len(existing_box_numbers) - 10} mã khác..." if len(existing_box_numbers) > 10 else "")
                        )
                        # Khôi phục giá trị cũ
                        values[4] = self.scanned_items[isbn]['so_thung'] if isbn in self.scanned_items else ''  # Số thùng ở index 4
                        self.tree.item(item, values=values)
                        # Reset flag và cleanup trước khi return
                        self.is_processing_edit = False
                        if self.edit_entry:
                            self.edit_entry.destroy()
                            self.edit_entry = None
                            self.editing_item = None
                        return
                
                values[4] = new_value  # Đảm bảo đúng index (Số thùng ở index 4)
                if isbn in self.scanned_items:
                    # Khi chỉnh sửa trực tiếp, cập nhật số thùng hiển thị
                    # Nhưng giữ nguyên số thùng gốc (từ dữ liệu đầu vào) - KHÔNG BAO GIỜ thay đổi
                    new_value_clean = new_value.strip()
                    self.scanned_items[isbn]['so_thung'] = new_value_clean
                else:
                    # Nếu ISBN không có trong scanned_items, không làm gì
                    pass
                
                # Đảm bảo so_thung_goc luôn được giữ nguyên (không thay đổi khi chỉnh sửa)
                # Chỉ set so_thung_goc nếu chưa có (lần đầu quét)
                if 'so_thung_goc' not in self.scanned_items[isbn] or not self.scanned_items[isbn].get('so_thung_goc'):
                    # Nếu chưa có, lấy từ current_box_number (số thùng đã load)
                    if self.current_box_number:
                        self.scanned_items[isbn]['so_thung_goc'] = self.current_box_number
                    else:
                        # Nếu không có current_box_number, dùng giá trị hiện tại (fallback)
                        self.scanned_items[isbn]['so_thung_goc'] = new_value_clean
                # Nếu đã có so_thung_goc, KHÔNG BAO GIỜ thay đổi nó
                
                # QUAN TRỌNG: Giữ nguyên vi_tri_moi khi chỉnh sửa trực tiếp
                # Nếu chưa có vi_tri_moi, lấy từ ô input hiện tại
                if 'vi_tri_moi' not in self.scanned_items[isbn] or not self.scanned_items[isbn].get('vi_tri_moi'):
                    vi_tri_moi_current = self.vi_tri_moi_var.get().strip()
                    if vi_tri_moi_current:
                        self.scanned_items[isbn]['vi_tri_moi'] = vi_tri_moi_current
        
            elif column_index == 5:  # Tồn tựa trong thùng - Showroom: Bỏ logic check chênh lệch
                # Chuyển thành số nguyên
                try:
                    new_value_int = int(float(new_value)) if new_value else 0
                    values[5] = str(new_value_int)  # Đảm bảo đúng index và chuyển thành string (Tồn tựa trong thùng ở index 5)
                    if isbn in self.scanned_items:
                        self.scanned_items[isbn]['ton_trong_thung'] = new_value_int
                        # Lưu backup khi có thay đổi
                        self.save_backup_on_change()
                except:
                    values[5] = new_value  # Đảm bảo đúng index
                    if isbn in self.scanned_items:
                        self.scanned_items[isbn]['ton_trong_thung'] = new_value
                        # Lưu backup khi có thay đổi
                        self.save_backup_on_change()
                
                # Đảm bảo có đủ 9 cột
                values = self.ensure_values_format(values)
                # Cập nhật tree
                self.tree.item(item, values=values)
            
            elif column_index == 7:  # Ghi chú - Showroom: Bỏ logic check chênh lệch
                # Đảm bảo có đủ 9 cột
                values = self.ensure_values_format(values)
                # Chỉ lưu giá trị mới vào values và scanned_items - không có logic phức tạp
                values[7] = new_value
                if isbn in self.scanned_items:
                    self.scanned_items[isbn]['ghi_chu'] = new_value
                    # Lưu backup khi có thay đổi
                    self.save_backup_on_change()
                # Cập nhật tree với giá trị mới
                self.tree.item(item, values=values)
        finally:
            # Reset flag và cleanup
            self.is_processing_edit = False
            self._finish_edit_scheduled = False  # Reset flag finish_edit
            # Xóa Entry widget
            if self.edit_entry:
                try:
                    self.edit_entry.destroy()
                except:
                    pass  # Widget có thể đã bị destroy
                self.edit_entry = None
                self.editing_item = None
            # Focus lại vào ISBN entry
            try:
                self.isbn_entry.focus()
            except:
                pass  # Widget có thể không tồn tại
    
    def cancel_edit(self):
        """Hủy việc chỉnh sửa"""
        if self.edit_entry:
            try:
                self.edit_entry.destroy()
            except:
                pass  # Widget có thể đã bị destroy
            self.edit_entry = None
            self.editing_item = None
        self.is_processing_edit = False  # Reset flag khi hủy
        self._finish_edit_scheduled = False  # Reset flag finish_edit
        try:
            self.isbn_entry.focus()
        except:
            pass  # Widget có thể không tồn tại
    
    def _check_and_update_status_after_increment(self, item_id, isbn):
        """Showroom: Bỏ logic check chênh lệch - hàm này không làm gì cả"""
        pass  # Không làm gì cả
    
    def highlight_error_cells(self, item_id):
        """Tô đỏ 2 ô: Tồn thực tế và Tình trạng"""
        # Xóa highlight cũ nếu có
        self.remove_error_highlights(item_id)
        
        # Lấy giá trị từ tree
        values = list(self.tree.item(item_id, 'values'))
        
        # Tô đỏ ô "Tồn thực tế" (cột 4, index 3)
        bbox_ton = self.tree.bbox(item_id, '#4')
        if bbox_ton:
            x, y, width, height = bbox_ton
            ton_value = values[3] if len(values) > 3 else ''
            highlight1 = tk.Label(self.tree, bg='#FFCDD2', fg='#C62828', 
                                  text=str(ton_value), font=('Arial', 10, 'bold'), 
                                  relief=tk.FLAT, anchor='center')
            highlight1.place(x=x, y=y, width=width, height=height)
            # Cho phép click qua để edit
            # Fix closure issue: capture item_id và column vào biến local
            item_id_click = item_id
            highlight1.bind('<Button-1>', lambda e, i=item_id_click, c='#4': self.on_highlight_click(e, i, c))
        
        # Tô đỏ ô "Tình trạng" (cột 7, index 6)
        bbox_tinh_trang = self.tree.bbox(item_id, '#7')
        if bbox_tinh_trang:
            x, y, width, height = bbox_tinh_trang
            tinh_trang_value = values[6] if len(values) > 6 else ''
            highlight2 = tk.Label(self.tree, bg='#FFCDD2', fg='#C62828', 
                                 text=str(tinh_trang_value), font=('Arial', 10, 'bold'), 
                                 relief=tk.FLAT, anchor='center')
            highlight2.place(x=x, y=y, width=width, height=height)
            # Không cho phép edit cột Tình trạng (chỉ đọc)
        
        # Lưu các highlight widgets
        if item_id not in self.error_highlights:
            self.error_highlights[item_id] = []
        if bbox_ton:
            self.error_highlights[item_id].append(highlight1)
        if bbox_tinh_trang:
            self.error_highlights[item_id].append(highlight2)
        
        # Cập nhật lại highlight khi scroll hoặc resize
        # Fix closure issue: capture item_id vào biến local
        item_id_to_update = item_id
        self.root.after(100, lambda i=item_id_to_update: self.update_error_highlights(i))
    
    def on_highlight_click(self, event, item_id, column):
        """Xử lý click vào highlight để edit cell"""
        # Hủy edit cũ nếu có
        if self.edit_entry:
            self.finish_edit()
        
        column_index = int(column.replace('#', '')) - 1
        
        # Không cho phép edit cột Tình trạng (6) - chỉ đọc
        if column_index == 6:
            return
        
        # Kiểm tra xem item có tồn tại không
        try:
            # Kiểm tra item có tồn tại bằng cách lấy children
            all_items = self.tree.get_children()
            if item_id not in all_items:
                return  # Item không tồn tại, không làm gì cả
        except:
            return  # Lỗi khi kiểm tra, không làm gì cả
        
        # Lấy giá trị hiện tại
        try:
            values = list(self.tree.item(item_id, 'values'))
            current_value = values[column_index] if column_index < len(values) else ''
        except Exception as e:
            # Item có thể đã bị xóa, không làm gì cả
            return
        
        # Lấy vị trí của cell
        bbox = self.tree.bbox(item_id, column)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        # Tạo Entry widget để edit trực tiếp
        self.edit_entry = tk.Entry(self.tree, font=('Arial', 10), 
                                   relief=tk.FLAT, bd=0, bg='#FFFFFF', fg='#000000')
        self.edit_entry.insert(0, str(current_value))
        self.edit_entry.select_range(0, tk.END)
        self.edit_entry.place(x=x, y=y, width=width, height=height)
        self.edit_entry.focus()
        self.editing_item = item_id
        self.editing_column = column_index
        
        def finish_on_enter(event):
            self.finish_edit()
        
        def finish_on_focus_out(event):
            self.root.after(100, self.finish_edit)
        
        self.edit_entry.bind('<Return>', finish_on_enter)
        self.edit_entry.bind('<FocusOut>', finish_on_focus_out)
        self.edit_entry.bind('<Escape>', lambda e: self.cancel_edit())
    
    def update_error_highlights(self, item_id):
        """Cập nhật lại vị trí highlight khi scroll"""
        if item_id not in self.error_highlights:
            return
        
        # Kiểm tra xem item còn tồn tại không
        try:
            if item_id not in self.tree.get_children(''):
                return
        except:
            return
        
        # Lấy giá trị từ tree
        values = list(self.tree.item(item_id, 'values'))
        
        widgets = self.error_highlights[item_id]
        if len(widgets) >= 2:
            # Cập nhật ô Tồn thực tế
            bbox_ton = self.tree.bbox(item_id, '#4')
            if bbox_ton and widgets[0].winfo_exists():
                x, y, width, height = bbox_ton
                ton_value = values[3] if len(values) > 3 else ''
                widgets[0].config(text=str(ton_value))
                widgets[0].place(x=x, y=y, width=width, height=height)
            
            # Cập nhật ô Tình trạng
            bbox_tinh_trang = self.tree.bbox(item_id, '#7')
            if bbox_tinh_trang and len(widgets) > 1 and widgets[1].winfo_exists():
                x, y, width, height = bbox_tinh_trang
                tinh_trang_value = values[6] if len(values) > 6 else ''
                widgets[1].config(text=str(tinh_trang_value))
                widgets[1].place(x=x, y=y, width=width, height=height)
    
    def update_all_highlights(self):
        """Cập nhật tất cả các highlight khi scroll hoặc resize"""
        # Delay một chút để đảm bảo tree đã cập nhật
        self.root.after(10, self._do_update_all_highlights)
    
    def _do_update_all_highlights(self):
        """Thực hiện cập nhật tất cả highlights"""
        for item_id in list(self.error_highlights.keys()):
            try:
                # Kiểm tra item còn tồn tại không
                if item_id in self.tree.get_children(''):
                    self.update_error_highlights(item_id)
                else:
                    # Xóa highlight nếu item không còn tồn tại
                    self.remove_error_highlights(item_id)
            except:
                pass
    
    def remove_error_highlights(self, item_id):
        """Xóa highlight của các ô lỗi"""
        if item_id in self.error_highlights:
            for widget in self.error_highlights[item_id]:
                try:
                    widget.destroy()
                except:
                    pass
            del self.error_highlights[item_id]
    
    def auto_edit_ton_thuc_te(self, item_id):
        """Tự động mở edit cho cột 'Tồn thực tế' sau khi thêm item mới"""
        if not item_id:
            return
        
        # Đảm bảo không có edit đang diễn ra
        if self.edit_entry:
            self.finish_edit()
        
        # Lấy vị trí của cell "Tồn thực tế" (column index 3)
        column = '#4'  # Column index 3 (0-indexed) + 1
        bbox = self.tree.bbox(item_id, column)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        # Lấy giá trị hiện tại từ tree - đảm bảo lấy đúng từ cột "Tồn thực tế" (index 3)
        values = list(self.tree.item(item_id, 'values'))
        values = self.ensure_values_format(values)
        # Lấy giá trị từ cột "Tồn thực tế" (index 3), không phải từ cột khác
        current_value = str(values[3]) if len(values) > 3 and values[3] is not None else ''
        
        # Tạo Entry widget để edit trực tiếp
        self.edit_entry = tk.Entry(self.tree, font=('Arial', 10), 
                                   relief=tk.FLAT, bd=0, bg='#FFFFFF', fg='#000000')
        self.edit_entry.insert(0, str(current_value))
        self.edit_entry.select_range(0, tk.END)
        self.edit_entry.place(x=x, y=y, width=width, height=height)
        self.edit_entry.focus()
        self.editing_item = item_id
        self.editing_column = 3  # Tồn thực tế ở index 3
        
        # Reset flag finish_edit khi bắt đầu edit mới
        self._finish_edit_scheduled = False
        
        def finish_on_enter(event):
            if not self._finish_edit_scheduled:
                self._finish_edit_scheduled = True
                self.finish_edit()
        
        def finish_on_focus_out(event):
            if not self._finish_edit_scheduled:
                self._finish_edit_scheduled = True
                self.root.after(100, self.finish_edit)
        
        self.edit_entry.bind('<Return>', finish_on_enter)
        self.edit_entry.bind('<FocusOut>', finish_on_focus_out)
        self.edit_entry.bind('<Escape>', lambda e: self.cancel_edit())
    
    
    def clear_table(self):
        """Xóa tất cả items trong bảng"""
        # Xóa tất cả highlights trước
        for item_id in list(self.error_highlights.keys()):
            self.remove_error_highlights(item_id)
        
        # Xóa tất cả items
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Cập nhật "Đã quét": số dòng trong Tổng hợp + số dòng trong Kiểm kê (0 vì đã clear)
        self.update_da_quet_counter()
    
    def reset_scanned_data(self):
        """Reset lại tất cả dữ liệu đã quét và nhập để làm lại"""
        # Xác nhận với người dùng
        result = messagebox.askyesno(
            "Xác nhận Reset", 
            "Bạn có chắc chắn muốn reset lại tất cả dữ liệu đã quét?\n\n"
            "Tất cả các ISBN đã quét và dữ liệu nhập sẽ bị xóa.\n"
            "Các thông tin như Số thùng, Ngày, Tổ sẽ được giữ lại."
        )
        
        if not result:
            return
        
        # Xóa tất cả items trong bảng
        self.clear_table()
        
        # Xóa tất cả scanned items
        self.scanned_items = {}
        
        # Enable lại input số thùng sau khi RESET - restore màu nền ban đầu
        if hasattr(self, 'so_thung_entry'):
            self.so_thung_entry.config(state='normal', bg='#FFF9C4', fg='#000000')
        if hasattr(self, 'so_thung_original_value'):
            self.so_thung_original_value = ''
        
        # Xóa input ISBN
        if hasattr(self, 'isbn_entry'):
            self.isbn_entry.delete(0, tk.END)
            self.isbn_entry.focus()
        
        # Hủy edit nếu đang edit
        if self.edit_entry:
            self.cancel_edit()
        
        # Cập nhật "Đã quét"
        self.update_da_quet_counter()
        
        # Thông báo thành công
        messagebox.showinfo("Thành công", "Đã reset lại tất cả dữ liệu đã quét.\nBạn có thể bắt đầu quét lại từ đầu.")
    
    def on_enter_pressed(self, event):
        """Xử lý phím Enter"""
        widget = self.root.focus_get()
        if widget == self.isbn_entry:
            self.on_isbn_entered()
    
    def save_data(self):
        """Lưu dữ liệu đã kiểm tra vào tab Tổng hợp"""
        if not self.scanned_items:
            messagebox.showwarning("Cảnh báo", "Chưa có dữ liệu để lưu!")
            return
        
        # Kiểm tra ràng buộc: Tổ là bắt buộc
        to_value = self.to_var.get().strip() if hasattr(self, 'to_var') and self.to_var.get() else ''
        if not to_value:
            messagebox.showerror("Lỗi", "Vui lòng nhập 'Tổ' trước khi lưu!")
            # Focus vào ô input Tổ
            if hasattr(self, 'to_entry'):
                self.to_entry.focus()
                self.to_entry.select_range(0, tk.END)
            return
        
        # Kiểm tra ràng buộc: Số tựa đã quét phải bằng số tựa trong thùng
        if self.current_box_data is not None and not self.current_box_data.empty and self.current_box_number:
            so_tua_trong_thung = len(self.current_box_data)
            # QUAN TRỌNG: Chỉ đếm các ISBN hợp lệ (tồn tại trong thùng), không đếm ISBN không tồn tại
            so_tua_da_quet_lan_nay = self.count_valid_scanned_isbns()
            
            # QUAN TRỌNG: Đếm cả các tựa đã lưu trong tab Tổng hợp từ các lần save trước
            # (ISBN không tồn tại không được lưu vào tổng hợp nên không được đếm)
            so_tua_da_luu_truoc = self.count_scanned_titles_for_box(self.current_box_number) if self.current_box_number else 0
            
            # Tổng số tựa đã quét = số tựa hợp lệ quét lần này + số tựa đã lưu trước đó
            so_tua_da_quet_tong = so_tua_da_quet_lan_nay + so_tua_da_luu_truoc
            
            if so_tua_da_quet_tong < so_tua_trong_thung:
                # Tạo dialog tùy chỉnh với 2 nút
                dialog = tk.Toplevel(self.root)
                dialog.title("Cảnh báo")
                dialog.geometry("600x400")
                dialog.resizable(False, False)
                dialog.transient(self.root)
                dialog.grab_set()
                dialog.configure(bg='#f5f5f5')
                
                # Đặt dialog ở giữa màn hình
                dialog.update_idletasks()
                x = (dialog.winfo_screenwidth() // 2) - (600 // 2)
                y = (dialog.winfo_screenheight() // 2) - (400 // 2)
                dialog.geometry(f"600x400+{x}+{y}")
                
                # Frame chính với padding đủ để các nút không bị che
                main_frame = tk.Frame(dialog, padx=30, pady=30, bg='#f5f5f5')
                main_frame.pack(fill=tk.BOTH, expand=True)
                
                # Icon cảnh báo
                icon_frame = tk.Frame(main_frame, bg='#f5f5f5')
                icon_frame.pack(pady=(0, 15))
                icon_label = tk.Label(icon_frame, text="⚠️", font=('Arial', 32), bg='#f5f5f5')
                icon_label.pack()
                
                # Thông điệp - hiển thị cả số tựa đã lưu trước đó
                if so_tua_da_luu_truoc > 0:
                    message_text = (
                        f"Chưa quét đủ số tựa trong thùng!\n\n"
                        f"Thùng {self.current_box_number} có {so_tua_trong_thung} tựa.\n"
                        f"Đã quét lần này: {so_tua_da_quet_lan_nay} tựa.\n"
                        f"Đã lưu trước đó: {so_tua_da_luu_truoc} tựa.\n"
                        f"Tổng đã quét: {so_tua_da_quet_tong} tựa.\n"
                        f"Còn thiếu: {so_tua_trong_thung - so_tua_da_quet_tong} tựa.\n\n"
                        f"Bạn muốn tiếp tục lưu hay hủy để quét tiếp?"
                    )
                else:
                    message_text = (
                        f"Chưa quét đủ số tựa trong thùng!\n\n"
                        f"Thùng {self.current_box_number} có {so_tua_trong_thung} tựa.\n"
                        f"Đã quét: {so_tua_da_quet_lan_nay} tựa.\n"
                        f"Còn thiếu: {so_tua_trong_thung - so_tua_da_quet_lan_nay} tựa.\n\n"
                        f"Bạn muốn tiếp tục lưu hay hủy để quét tiếp?"
                    )
                message_frame = tk.Frame(main_frame, bg='#f5f5f5')
                message_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
                message_label = tk.Label(
                    message_frame, 
                    text=message_text,
                    font=('Arial', 12),
                    justify=tk.CENTER,
                    bg='#f5f5f5',
                    fg='#333333',
                    wraplength=540
                )
                message_label.pack()
                
                # Frame chứa các nút với padding đủ để không bị che
                button_frame = tk.Frame(main_frame, bg='#f5f5f5')
                button_frame.pack(pady=(15, 25))
                
                # Biến để lưu kết quả
                result = {'value': None}
                
                def on_continue_save():
                    result['value'] = 'continue'
                    dialog.destroy()
                
                def on_cancel():
                    result['value'] = 'cancel'
                    dialog.destroy()
                
                # Nút "Tiếp tục và lưu"
                continue_btn = tk.Button(
                    button_frame,
                    text="Tiếp tục và lưu",
                    command=on_continue_save,
                    font=('Arial', 12, 'bold'),
                    bg='#4CAF50',
                    fg='white',
                    padx=25,
                    pady=12,
                    relief=tk.FLAT,
                    cursor='hand2',
                    activebackground='#45a049',
                    activeforeground='white',
                    bd=0,
                    highlightthickness=0
                )
                continue_btn.pack(side=tk.LEFT, padx=15)
                
                # Nút "Cancel"
                cancel_btn = tk.Button(
                    button_frame,
                    text="Cancel",
                    command=on_cancel,
                    font=('Arial', 12, 'bold'),
                    bg='#f44336',
                    fg='white',
                    padx=25,
                    pady=12,
                    relief=tk.FLAT,
                    cursor='hand2',
                    activebackground='#da190b',
                    activeforeground='white',
                    bd=0,
                    highlightthickness=0
                )
                cancel_btn.pack(side=tk.LEFT, padx=15)
                
                # Đợi dialog đóng
                dialog.wait_window()
                
                # Xử lý kết quả
                if result['value'] == 'cancel':
                    # Focus vào ô nhập ISBN để tiếp tục quét
                    if hasattr(self, 'isbn_entry'):
                        self.isbn_entry.focus()
                    return
                # Nếu chọn "Tiếp tục và lưu", tiếp tục với logic lưu bên dưới
        
        # Lấy giá trị từ các input
        vi_tri_moi_global = self.vi_tri_moi_var.get().strip() if hasattr(self, 'vi_tri_moi_var') else ''
        ngay_value = self.ngay_var.get().strip() if hasattr(self, 'ngay_var') else ''
        nhap_xuat_value = self.nhap_xuat_var.get().strip() if hasattr(self, 'nhap_xuat_var') else ''
        note_thung_value = self.note_thung_var.get().strip() if hasattr(self, 'note_thung_var') else ''
        
        # Tạo số phiếu theo format P-DD/MM/YYYY
        from datetime import datetime
        try:
            if ngay_value:
                # Parse ngày từ format DD/MM/YY hoặc DD/MM/YYYY
                if len(ngay_value.split('/')) == 3:
                    parts = ngay_value.split('/')
                    if len(parts[2]) == 2:
                        # Format DD/MM/YY -> DD/MM/YYYY
                        parts[2] = '20' + parts[2]
                    ngay_parsed = datetime.strptime('/'.join(parts), "%d/%m/%Y")
                else:
                    ngay_parsed = datetime.now()
            else:
                ngay_parsed = datetime.now()
        except:
            ngay_parsed = datetime.now()
        
        so_phieu = f"P-{ngay_parsed.strftime('%d/%m/%Y')}"
        
        # Thêm dữ liệu vào tổng hợp - tối ưu cho dữ liệu lớn
        items_to_add = []
        for isbn, info in self.scanned_items.items():
            try:
                # Kiểm tra xem item có tồn thực tế chưa
                if not info.get('ton_thuc_te', '').strip():
                    continue  # Bỏ qua item chưa nhập tồn thực tế
                
                # Lưu cả ISBN tồn tại và không tồn tại vào tổng hợp
                # (ISBN không tồn tại vẫn được lưu, nhưng không được tính vào số tựa đã quét)
                
                # Lấy số thùng gốc
                so_thung_goc = info.get('so_thung_goc', '')
                if not so_thung_goc:
                    so_thung_goc = self.current_box_number if self.current_box_number else info.get('so_thung', '')
                so_thung_goc_clean = str(so_thung_goc).strip()
                
                # Xác định số thùng mới (vị trí mới)
                # QUAN TRỌNG: "Vị trí mới" chỉ lấy từ vi_tri_moi, KHÔNG fallback về so_thung_goc
                so_thung_moi = ''
                
                # Ưu tiên 1: vi_tri_moi_global (từ ô input "Thùng / vị trí mới")
                if vi_tri_moi_global and vi_tri_moi_global.strip():
                    so_thung_moi = vi_tri_moi_global.strip()
                else:
                    # Ưu tiên 2: vi_tri_moi từ scanned_items (đã lưu khi quét)
                    vi_tri_moi_saved = info.get('vi_tri_moi', '').strip()
                    if vi_tri_moi_saved:
                        so_thung_moi = vi_tri_moi_saved
                    # Nếu không có vi_tri_moi, để trống (không lấy từ so_thung gốc)
                
                # Lấy giá trị từ input "Nhập/Xuất" ở tab Kiểm kê và hiển thị trực tiếp vào cột N/X
                nx_value = nhap_xuat_value.strip() if nhap_xuat_value else ""
                
                # Lấy flag đánh dấu ISBN có tồn tại không (để đếm số tựa đã quét)
                is_invalid_isbn = info.get('is_invalid_isbn', False)
                is_valid_isbn = not is_invalid_isbn  # ISBN hợp lệ (tồn tại) = không phải invalid
                
                # Thêm vào danh sách để append sau (hiệu quả hơn)
                items_to_add.append({
                    'N/X': nx_value,
                    'Số phiếu': so_phieu,
                    'Ngày': ngay_value,
                    'Vị trí mới': so_thung_moi,  # Chỉ lấy từ vi_tri_moi, không fallback về so_thung_goc
                    'ISBN': isbn,
                    'Tựa': info.get('tua', ''),
                    'Tồn thực tế': info.get('ton_thuc_te', ''),
                    'Số thùng': so_thung_goc_clean,
                    'Tình trạng': info.get('tinh_trang', ''),  # Lấy từ scanned_items
                    'Ghi chú': info.get('ghi_chu', ''),  # Ghi chú do người dùng tự nhập
                    'Note thùng': note_thung_value,  # Note thùng từ input
                    '_is_valid_isbn': is_valid_isbn  # Đánh dấu ISBN có tồn tại (để đếm số tựa đã quét)
                })
            except Exception as e:
                # Bỏ qua item lỗi và tiếp tục
                print(f"Lỗi khi xử lý item {isbn}: {str(e)}")
                continue
        
        # Lưu số thùng hiện tại trước khi reset để cập nhật số tựa đã quét
        saved_box_number = self.current_box_number
        items_count = len(items_to_add)
        
        # Xử lý với dữ liệu lớn (10000+ dòng)
        try:
            # Hiển thị thông báo cho dữ liệu lớn
            if items_count > 5000:
                # Disable các nút để tránh click nhiều lần
                if hasattr(self, 'save_button'):
                    self.save_button.config(state='disabled')
                self.root.update()
                
                # Hiển thị message trong status (nếu có)
                status_msg = f"Đang lưu {items_count:,} dòng... Vui lòng đợi..."
                print(status_msg)
            
            # Thêm tất cả items vào tổng hợp cùng lúc (hiệu quả hơn append từng cái)
            # Với dữ liệu cực lớn, extend vẫn nhanh hơn append từng cái
            self.tong_hop_data.extend(items_to_add)
            
            # Lưu backup ngay sau khi extend để tránh mất dữ liệu nếu crash
            try:
                if items_count > 1000:
                    # Lưu backup ngay lập tức cho dữ liệu lớn
                    self.save_backup()
            except Exception as backup_err:
                # Log nhưng không chặn quá trình
                print(f"Lỗi khi lưu backup: {str(backup_err)}")
            
            # Cập nhật bảng tổng hợp (có xử lý lỗi và batch processing bên trong)
            self.update_tong_hop_table()
            
        except MemoryError:
            # Xử lý lỗi memory
            messagebox.showerror("Lỗi", 
                f"Không đủ bộ nhớ để lưu {items_count:,} dòng!\n\n"
                f"Vui lòng thử lưu với số lượng ít hơn hoặc đóng các ứng dụng khác.")
            return
        except Exception as e:
            # Xử lý lỗi chung
            error_msg = f"Lỗi khi lưu dữ liệu: {str(e)}\n\nSố dòng đang lưu: {items_count:,}"
            messagebox.showerror("Lỗi", error_msg)
            print(f"Chi tiết lỗi: {traceback.format_exc()}")
            return
        finally:
            # Enable lại nút save
            if hasattr(self, 'save_button'):
                self.save_button.config(state='normal')
        
        # Xóa dữ liệu đã quét
        self.scanned_items.clear()
        self.clear_table()
        
        # Enable lại input số thùng và reset về rỗng sau khi SAVE để cho phép nhập số thùng mới
        # Restore màu nền ban đầu
        if hasattr(self, 'so_thung_entry'):
            self.so_thung_entry.config(state='normal', bg='#FFF9C4', fg='#000000')
        if hasattr(self, 'so_thung_var'):
            self.so_thung_var.set("")
        if hasattr(self, 'so_thung_original_value'):
            self.so_thung_original_value = ''
        
        # Reset các input: Thùng / vị trí mới, và Note thùng
        # QUAN TRỌNG: KHÔNG reset so_thung_var ở đây để giữ lại số thùng hiện tại
        # Nếu reset so_thung_var, sẽ làm mất số thùng và không thể đếm được số dòng trong Tổng hợp
        # Người dùng có thể tự xóa và nhập số thùng mới khi cần
        if hasattr(self, 'vi_tri_moi_var'):
            self.vi_tri_moi_var.set("")
        if hasattr(self, 'note_thung_var'):
            self.note_thung_var.set("")
        
        # Reset current_box_number về None, nhưng giữ lại số thùng trong so_thung_var
        # Để khi người dùng nhập số thùng mới, có thể đếm được số dòng trong Tổng hợp
        self.current_box_number = None
        self.current_box_data = None
        
        # Cập nhật "Đã quét" sau khi save: số dòng trong Tổng hợp + số dòng trong Kiểm kê (0 vì đã clear)
        # QUAN TRỌNG: Cập nhật SAU KHI giữ lại so_thung_var để đếm đúng
        # Nếu so_thung_var vẫn còn giá trị, sẽ đếm được số dòng trong Tổng hợp cho số thùng đó
        self.update_da_quet_counter()
        
        # Chuyển sang tab Tổng hợp
        self.notebook.select(1)
        
        # Thông báo thành công với format số cho dữ liệu lớn
        total_count = len(self.tong_hop_data)
        messagebox.showinfo("Thành công", 
            f"Đã lưu {items_count:,} dòng mới vào Tổng hợp!\nTổng cộng: {total_count:,} dòng")
    
    def _aggregate_tong_hop_data(self):
        """Cộng dồn các dòng có cùng ISBN và cùng Số thùng trong tong_hop_data"""
        if not self.tong_hop_data:
            return []
        
        # Dictionary để nhóm các dòng theo (ISBN, Số thùng)
        aggregated = {}
        
        for data in self.tong_hop_data:
            isbn = str(data.get('ISBN', '')).strip()
            so_thung = str(data.get('Số thùng', '')).strip()
            
            # Tạo key từ ISBN và Số thùng
            key = (isbn, so_thung)
            
            if key not in aggregated:
                # Dòng đầu tiên - giữ nguyên tất cả thông tin
                aggregated[key] = data.copy()
                # Đảm bảo "Tồn thực tế" là số
                try:
                    aggregated[key]['Tồn thực tế'] = float(data.get('Tồn thực tế', 0) or 0)
                except (ValueError, TypeError):
                    aggregated[key]['Tồn thực tế'] = 0.0
            else:
                # Dòng trùng - cộng dồn "Tồn thực tế"
                try:
                    ton_thuc_te_new = float(data.get('Tồn thực tế', 0) or 0)
                except (ValueError, TypeError):
                    ton_thuc_te_new = 0.0
                
                aggregated[key]['Tồn thực tế'] = aggregated[key].get('Tồn thực tế', 0) + ton_thuc_te_new
        
        # Chuyển dictionary thành list và chuyển "Tồn thực tế" về string để hiển thị
        result = []
        for data in aggregated.values():
            # Chuyển "Tồn thực tế" về string (loại bỏ .0 nếu là số nguyên)
            ton_thuc_te = data.get('Tồn thực tế', 0)
            if isinstance(ton_thuc_te, float):
                if ton_thuc_te == int(ton_thuc_te):
                    data['Tồn thực tế'] = str(int(ton_thuc_te))
                else:
                    data['Tồn thực tế'] = str(ton_thuc_te)
            else:
                data['Tồn thực tế'] = str(ton_thuc_te)
            result.append(data)
        
        return result
    
    def update_tong_hop_table(self):
        """Cập nhật bảng tổng hợp - tối ưu cho dữ liệu lớn"""
        if not self.tong_hop_tree:
            return
        
        try:
            # Tắt cập nhật UI để tăng tốc độ
            self.tong_hop_tree.config(cursor='wait')
            self.root.config(cursor='wait')
            
            # Xóa tất cả items cũ - sử dụng cách nhanh hơn
            children = self.tong_hop_tree.get_children()
            if children:
                self.tong_hop_tree.delete(*children)
            
            # Tắt update trong quá trình insert để tăng tốc độ
            self.root.update_idletasks()
            
            # Cộng dồn dữ liệu: các dòng có cùng ISBN và cùng Số thùng
            aggregated_data = self._aggregate_tong_hop_data()
            
            # Kiểm tra số lượng dữ liệu (sau khi cộng dồn)
            total_items = len(aggregated_data)
            
            # Tối ưu batch size dựa trên số lượng dữ liệu
            # Với dữ liệu lớn (>1000 items), sử dụng batch insert để tránh freeze UI
            if total_items > 1000:
                # Điều chỉnh batch size dựa trên số lượng dữ liệu
                if total_items > 10000:
                    batch_size = 500  # Batch lớn hơn cho dữ liệu cực lớn
                elif total_items > 5000:
                    batch_size = 300
                elif total_items > 2000:
                    batch_size = 200
                else:
                    batch_size = 100
                
                # Hiển thị thông báo cho dữ liệu lớn
                if total_items > 5000:
                    self.root.update()
                    status_msg = f"Đang cập nhật {total_items:,} dòng... Vui lòng đợi..."
                    # Có thể thêm progress bar nếu cần
                
                processed = 0
                for i in range(0, total_items, batch_size):
                    batch = aggregated_data[i:i + batch_size]
                    
                    # Tạo values list cho batch trước để tối ưu
                    batch_values = []
                    for data in batch:
                        try:
                            batch_values.append((
                                data.get('N/X', ''),
                                data.get('Số phiếu', ''),
                                data.get('Ngày', ''),
                                data.get('Vị trí mới', ''),
                                data.get('ISBN', ''),
                                data.get('Tựa', ''),
                                data.get('Tồn thực tế', ''),
                                data.get('Số thùng', ''),
                                data.get('Tình trạng', ''),
                                data.get('Ghi chú', ''),
                                data.get('Note thùng', '')
                            ))
                        except Exception as e:
                            # Bỏ qua item lỗi và tiếp tục
                            print(f"Lỗi khi chuẩn bị item: {str(e)}")
                            continue
                    
                    # Insert batch vào tree
                    for values in batch_values:
                        try:
                            self.tong_hop_tree.insert('', 'end', values=values)
                            processed += 1
                        except Exception as e:
                            # Bỏ qua item lỗi và tiếp tục
                            print(f"Lỗi khi insert item: {str(e)}")
                            continue
                    
                    # Update UI mỗi batch để không freeze và hiển thị tiến trình
                    if i + batch_size < total_items:
                        self.root.update_idletasks()
                        # Cho phép xử lý events để UI không bị đóng băng
                        if processed % (batch_size * 2) == 0:  # Update mỗi 2 batch
                            self.root.update()
            else:
                # Với dữ liệu nhỏ, insert trực tiếp
                for data in aggregated_data:
                    try:
                        self.tong_hop_tree.insert('', 'end', values=(
                            data.get('N/X', ''),
                            data.get('Số phiếu', ''),
                            data.get('Ngày', ''),
                            data.get('Vị trí mới', ''),
                            data.get('ISBN', ''),
                            data.get('Tựa', ''),
                            data.get('Tồn thực tế', ''),
                            data.get('Số thùng', ''),
                            data.get('Tình trạng', ''),
                            data.get('Ghi chú', ''),
                            data.get('Note thùng', '')
                        ))
                    except Exception as e:
                        # Bỏ qua item lỗi và tiếp tục
                        print(f"Lỗi khi insert item: {str(e)}")
                        continue
            
            # Bật lại cursor bình thường
            self.tong_hop_tree.config(cursor='')
            self.root.config(cursor='')
            self.root.update_idletasks()
            
        except Exception as e:
            # Xử lý lỗi để tránh crash
            self.tong_hop_tree.config(cursor='')
            self.root.config(cursor='')
            messagebox.showerror("Lỗi", f"Không thể cập nhật bảng tổng hợp: {str(e)}\n\nSố lượng dữ liệu: {len(self.tong_hop_data)}")
            print(f"Lỗi khi update_tong_hop_table: {str(e)}")
    
    def on_tong_hop_item_click(self, event):
        """Xử lý click để edit trực tiếp các cột có thể chỉnh sửa trong tab Tổng hợp"""
        # Hủy edit cũ nếu có
        if self.tong_hop_edit_entry:
            self.finish_tong_hop_edit()
        
        region = self.tong_hop_tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        
        item = self.tong_hop_tree.identify_row(event.y)
        column = self.tong_hop_tree.identify_column(event.x)
        column_index = int(column.replace('#', '')) - 1
        
        # Cho phép edit: Vị trí mới (3), Tồn thực tế (6), Tình trạng (8), Ghi chú (9), Note thùng (10)
        # Không cho edit: N/X (0), Số phiếu (1), Ngày (2), ISBN (4), Tựa (5), Số thùng (7) - chỉ đọc
        editable_columns = [3, 6, 8, 9, 10]  # Vị trí mới, Tồn thực tế, Tình trạng, Ghi chú, Note thùng
        if column_index not in editable_columns:
            return
        
        if not item:
            return
        
        # Lấy giá trị hiện tại
        values = list(self.tong_hop_tree.item(item, 'values'))
        current_value = values[column_index] if column_index < len(values) else ''
        
        # Lấy vị trí của cell
        bbox = self.tong_hop_tree.bbox(item, column)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        # Tạo Entry widget để edit trực tiếp
        self.tong_hop_edit_entry = tk.Entry(self.tong_hop_tree, font=('Arial', 10), 
                                           relief=tk.FLAT, bd=0, bg='#FFFFFF', fg='#000000')
        self.tong_hop_edit_entry.insert(0, str(current_value))
        self.tong_hop_edit_entry.select_range(0, tk.END)
        self.tong_hop_edit_entry.place(x=x, y=y, width=width, height=height)
        self.tong_hop_edit_entry.focus()
        self.tong_hop_editing_item = item
        self.tong_hop_editing_column = column_index
        
        def finish_on_enter(event):
            # Hủy scheduled call nếu có
            if self._tong_hop_finish_scheduled is not None:
                try:
                    # Chỉ cancel nếu ID hợp lệ
                    if isinstance(self._tong_hop_finish_scheduled, str) and self._tong_hop_finish_scheduled:
                        self.root.after_cancel(self._tong_hop_finish_scheduled)
                except (ValueError, TypeError):
                    # Bỏ qua lỗi nếu ID không hợp lệ
                    pass
                self._tong_hop_finish_scheduled = None
            self.finish_tong_hop_edit()
        
        def finish_on_focus_out(event):
            # Kiểm tra entry còn tồn tại và chưa đang xử lý
            if not self.tong_hop_edit_entry or self.is_processing_tong_hop_edit:
                return
            
            # Kiểm tra widget còn tồn tại
            try:
                if not self.tong_hop_edit_entry.winfo_exists():
                    return
            except:
                return
            
            # Hủy scheduled call cũ nếu có (nếu có)
            if self._tong_hop_finish_scheduled is not None:
                try:
                    if isinstance(self._tong_hop_finish_scheduled, str) and self._tong_hop_finish_scheduled:
                        self.root.after_cancel(self._tong_hop_finish_scheduled)
                except (ValueError, TypeError, tk.TclError):
                    pass
                self._tong_hop_finish_scheduled = None
            
            # Gọi trực tiếp - KHÔNG dùng after() hoặc after_idle() để tránh lỗi
            # Sử dụng try-except để bắt mọi lỗi có thể xảy ra
            try:
                self.finish_tong_hop_edit()
            except Exception as e:
                # Nếu có lỗi, chỉ log và cleanup
                print(f"Error in finish_on_focus_out: {e}")
                try:
                    if self.tong_hop_edit_entry:
                        self.tong_hop_edit_entry.destroy()
                except:
                    pass
                self.tong_hop_edit_entry = None
                self.tong_hop_editing_item = None
                self.tong_hop_editing_column = None
                self.is_processing_tong_hop_edit = False
        
        self.tong_hop_edit_entry.bind('<Return>', finish_on_enter)
        self.tong_hop_edit_entry.bind('<FocusOut>', finish_on_focus_out)
        self.tong_hop_edit_entry.bind('<Escape>', lambda e: self.cancel_tong_hop_edit())
    
    def _do_finish_tong_hop_edit(self):
        """Wrapper để reset scheduled flag trước khi gọi finish_tong_hop_edit"""
        self._tong_hop_finish_scheduled = None
        # Gọi finish trực tiếp
        self.finish_tong_hop_edit()
    
    def finish_tong_hop_edit(self):
        """Hoàn tất việc chỉnh sửa trong tab Tổng hợp"""
        # Tránh xử lý 2 lần nếu đang trong quá trình xử lý
        if self.is_processing_tong_hop_edit:
            return
        
        # Hủy scheduled call nếu có
        if self._tong_hop_finish_scheduled is not None:
            try:
                if isinstance(self._tong_hop_finish_scheduled, str) and self._tong_hop_finish_scheduled:
                    self.root.after_cancel(self._tong_hop_finish_scheduled)
            except (ValueError, TypeError, tk.TclError):
                # Bỏ qua lỗi nếu ID không hợp lệ hoặc đã bị cancel
                pass
            self._tong_hop_finish_scheduled = None
        
        # Kiểm tra entry và item còn tồn tại
        if not self.tong_hop_edit_entry or not self.tong_hop_editing_item:
            return
        
        # Kiểm tra entry widget còn tồn tại trong window
        try:
            if not self.tong_hop_edit_entry.winfo_exists():
                # Entry đã bị destroy, cleanup và return
                self.tong_hop_edit_entry = None
                self.tong_hop_editing_item = None
                self.tong_hop_editing_column = None
                return
        except:
            # Nếu không thể kiểm tra, giả sử đã bị destroy
            self.tong_hop_edit_entry = None
            self.tong_hop_editing_item = None
            self.tong_hop_editing_column = None
            return
        
        # Đặt flag để tránh xử lý lại
        self.is_processing_tong_hop_edit = True
        
        try:
            # Lấy giá trị mới
            new_value = self.tong_hop_edit_entry.get()
            
            # Lấy item ID và column index
            item_id = self.tong_hop_editing_item
            column_index = self.tong_hop_editing_column
            
            # Lấy giá trị hiện tại của dòng
            values = list(self.tong_hop_tree.item(item_id, 'values'))
            
            # Cập nhật giá trị trong tree
            values[column_index] = new_value
            self.tong_hop_tree.item(item_id, values=values)
            
            # Tìm index của item trong tree để map với tong_hop_data
            all_items = list(self.tong_hop_tree.get_children())
            if item_id in all_items:
                data_index = all_items.index(item_id)
                
                # Cập nhật trong tong_hop_data theo index
                if 0 <= data_index < len(self.tong_hop_data):
                    # Map column index sang tên cột trong data
                    column_mapping = {
                        3: 'Vị trí mới',
                        6: 'Tồn thực tế',
                        8: 'Tình trạng',
                        9: 'Ghi chú',
                        10: 'Note thùng'
                    }
                    
                    column_name = column_mapping.get(column_index)
                    if column_name:
                        self.tong_hop_data[data_index][column_name] = new_value
                        
                        # Showroom: Bỏ logic so sánh và tự động điền Tình trạng/Ghi chú khi sửa Tồn thực tế
                        # Nếu là cột "Tồn thực tế" (column_index == 6), chỉ lưu giá trị, không check chênh lệch
                        
                        # Lưu backup khi chỉnh sửa (gọi trực tiếp, không dùng after)
                        try:
                            self.save_backup()
                        except Exception as backup_error:
                            # Không hiển thị lỗi cho người dùng, chỉ log
                            print(f"Error saving backup: {backup_error}")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể cập nhật dữ liệu: {str(e)}")
        finally:
            # Reset flag
            self.is_processing_tong_hop_edit = False
            
            # Xóa entry widget
            if self.tong_hop_edit_entry:
                try:
                    self.tong_hop_edit_entry.destroy()
                except:
                    pass
                self.tong_hop_edit_entry = None
            self.tong_hop_editing_item = None
            self.tong_hop_editing_column = None
    
    def _check_and_update_tinh_trang_tong_hop(self, data_index, ton_thuc_te_new, values, item_id):
        """Showroom: Bỏ logic so sánh - hàm này không làm gì cả"""
        pass  # Không làm gì cả
    
    def cancel_tong_hop_edit(self):
        """Hủy việc chỉnh sửa trong tab Tổng hợp"""
        # Hủy scheduled call nếu có
        if self._tong_hop_finish_scheduled is not None:
            try:
                # Chỉ cancel nếu ID hợp lệ
                if isinstance(self._tong_hop_finish_scheduled, str) and self._tong_hop_finish_scheduled:
                    self.root.after_cancel(self._tong_hop_finish_scheduled)
            except (ValueError, TypeError):
                # Bỏ qua lỗi nếu ID không hợp lệ
                pass
            self._tong_hop_finish_scheduled = None
        
        if self.tong_hop_edit_entry:
            try:
                self.tong_hop_edit_entry.destroy()
            except:
                pass
            self.tong_hop_edit_entry = None
        self.tong_hop_editing_item = None
        self.tong_hop_editing_column = None
    
    def on_tong_hop_delete(self, event):
        """Xóa dòng được chọn trong tab Tổng hợp"""
        selected_items = self.tong_hop_tree.selection()
        if not selected_items:
            return
        
        # Xác nhận xóa
        result = messagebox.askyesno("Xác nhận", f"Bạn có chắc chắn muốn xóa {len(selected_items)} dòng đã chọn?")
        if not result:
            return
        
        try:
            # Lấy tất cả items trong tree theo thứ tự
            all_items = list(self.tong_hop_tree.get_children())
            
            # Tìm index của các items được chọn
            selected_indices = []
            for item_id in selected_items:
                if item_id in all_items:
                    selected_indices.append(all_items.index(item_id))
            
            # Sắp xếp theo thứ tự ngược lại để xóa từ cuối lên (tránh lỗi index)
            selected_indices.sort(reverse=True)
            
            # Xóa từ tong_hop_data trước (theo index)
            for idx in selected_indices:
                if 0 <= idx < len(self.tong_hop_data):
                    del self.tong_hop_data[idx]
            
            # Xóa khỏi tree
            for item_id in selected_items:
                self.tong_hop_tree.delete(item_id)
            
            # Lưu backup sau khi xóa
            self.save_backup()
            
            messagebox.showinfo("Thành công", f"Đã xóa {len(selected_items)} dòng!")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xóa dòng: {str(e)}")
            traceback.print_exc()
    
    def on_tong_hop_search(self, event=None):
        """Tìm kiếm ISBN trong tab Tổng hợp"""
        if not hasattr(self, 'tong_hop_search_entry') or not self.tong_hop_search_entry:
            return
        
        search_isbn = self.tong_hop_search_entry.get().strip()
        if not search_isbn:
            return
        
        # Lấy tất cả items trong tree
        all_items = list(self.tong_hop_tree.get_children())
        
        if not all_items:
            messagebox.showinfo("Thông báo", "Không có dữ liệu trong bảng tổng hợp!")
            return
        
        # Lấy item hiện tại được chọn (nếu có)
        current_selection = self.tong_hop_tree.selection()
        start_index = 0
        
        # Nếu đang có item được chọn, tìm từ item tiếp theo
        if current_selection:
            try:
                current_item = current_selection[0]
                if current_item in all_items:
                    current_index = all_items.index(current_item)
                    start_index = current_index + 1
            except:
                pass
        
        # Chuẩn hóa ISBN để tìm kiếm
        search_isbn_clean = str(search_isbn).strip()
        search_isbn_digits = ''.join(filter(str.isdigit, search_isbn_clean))
        
        # Tìm từ vị trí start_index
        found_item = None
        for i in range(start_index, len(all_items)):
            item_id = all_items[i]
            values = list(self.tong_hop_tree.item(item_id, 'values'))
            
            # ISBN ở cột index 4
            if len(values) > 4:
                isbn_value = str(values[4]).strip()
                isbn_value_digits = ''.join(filter(str.isdigit, isbn_value))
                
                # So sánh ISBN
                if (isbn_value == search_isbn_clean or 
                    isbn_value.endswith(search_isbn_clean) or 
                    search_isbn_clean.endswith(isbn_value) or
                    (isbn_value_digits and search_isbn_digits and isbn_value_digits == search_isbn_digits)):
                    found_item = item_id
                    break
        
        # Nếu không tìm thấy từ start_index, tìm từ đầu
        if not found_item:
            for i in range(0, min(start_index, len(all_items))):
                item_id = all_items[i]
                values = list(self.tong_hop_tree.item(item_id, 'values'))
                
                if len(values) > 4:
                    isbn_value = str(values[4]).strip()
                    isbn_value_digits = ''.join(filter(str.isdigit, isbn_value))
                    
                    if (isbn_value == search_isbn_clean or 
                        isbn_value.endswith(search_isbn_clean) or 
                        search_isbn_clean.endswith(isbn_value) or
                        (isbn_value_digits and search_isbn_digits and isbn_value_digits == search_isbn_digits)):
                        found_item = item_id
                        break
        
        if found_item:
            # Xóa highlight cũ trước
            for item in self.tong_hop_tree.get_children():
                tags = list(self.tong_hop_tree.item(item, 'tags'))
                if 'search_highlight' in tags:
                    tags.remove('search_highlight')
                    self.tong_hop_tree.item(item, tags=tags)
            
            # Highlight dòng tìm thấy
            current_tags = list(self.tong_hop_tree.item(found_item, 'tags'))
            if 'search_highlight' not in current_tags:
                current_tags.append('search_highlight')
            self.tong_hop_tree.item(found_item, tags=current_tags)
            
            # Scroll đến item và chọn nó
            self.tong_hop_tree.selection_set(found_item)
            self.tong_hop_tree.focus(found_item)
            self.tong_hop_tree.see(found_item)
            
            # Focus vào tree để có thể chỉnh sửa ngay (double-click vào cột để edit)
            self.tong_hop_tree.focus()
        else:
            # Xóa highlight cũ nếu không tìm thấy
            for item in self.tong_hop_tree.get_children():
                tags = list(self.tong_hop_tree.item(item, 'tags'))
                if 'search_highlight' in tags:
                    tags.remove('search_highlight')
                    self.tong_hop_tree.item(item, tags=tags)
            
            # Tìm lại từ đầu nếu không tìm thấy
            messagebox.showinfo("Thông báo", f"Không tìm thấy ISBN: {search_isbn}")
            # Xóa selection để có thể tìm lại từ đầu lần sau
            self.tong_hop_tree.selection_set([])
    
    def on_tong_hop_search_keyrelease(self, event=None):
        """Tự động tìm kiếm khi gõ trong ô tìm kiếm (tùy chọn - có thể bỏ qua)"""
        # Có thể thêm logic tự động tìm kiếm khi gõ nếu muốn
        pass
    
    def update_excel_file_metadata(self, file_path):
        """
        Cập nhật metadata của file Excel (timestamps và properties) thành ngày hiện tại
        Để khi người dùng xem Properties của file sẽ thấy ngày tải về thay vì ngày từ file gốc
        """
        try:
            file_path_obj = Path(file_path)
            if not file_path_obj.exists():
                return
            
            # 1. Cập nhật file system timestamps (created, modified) thành thời điểm hiện tại
            current_time = time.time()
            os.utime(str(file_path_obj), (current_time, current_time))
            
            # 2. Cập nhật Excel file properties (metadata bên trong file)
            try:
                from openpyxl import load_workbook
                from datetime import datetime
                
                wb = load_workbook(str(file_path_obj))
                
                # Cập nhật các properties thành ngày hiện tại
                now = datetime.now()
                wb.properties.created = now
                wb.properties.modified = now
                wb.properties.lastModifiedBy = os.getenv('USERNAME', os.getenv('USER', 'User'))
                
                # Lưu lại file với properties mới
                wb.save(str(file_path_obj))
                wb.close()
            except Exception as excel_prop_error:
                # Nếu không thể cập nhật Excel properties, chỉ log và tiếp tục
                # File system timestamps đã được cập nhật ở trên
                print(f"Không thể cập nhật Excel properties (không ảnh hưởng): {str(excel_prop_error)}")
                
        except Exception as e:
            # Log lỗi nhưng không chặn quá trình save
            print(f"Lỗi khi cập nhật metadata file: {str(e)}")
            traceback.print_exc()
    
    def export_tong_hop_excel(self):
        """Xuất file Excel tổng hợp (logic giống save_data cũ) - sử dụng dữ liệu đã cộng dồn"""
        # Sử dụng dữ liệu đã cộng dồn (các dòng có cùng ISBN và cùng Số thùng)
        aggregated_data = self._aggregate_tong_hop_data()
        if not aggregated_data:
            messagebox.showwarning("Cảnh báo", "Chưa có dữ liệu tổng hợp để xuất!")
            return
        
        # Sử dụng pandas đã được import trong __init__
        if self.pd is None:
            try:
                import pandas as pd
                self.pd = pd
            except ImportError:
                messagebox.showerror("Lỗi", "Không thể import pandas! Vui lòng cài đặt: pip install pandas")
                return
        
        pd = self.pd  # Alias để dùng trong hàm này
        df_save = pd.DataFrame(aggregated_data)
        
        # Tạo tên file theo format
        from datetime import datetime
        ngay_hien_tai = datetime.now().strftime("%d/%m/%Y")
        to_value = self.to_var.get().strip() if hasattr(self, 'to_var') and self.to_var.get() else ""
        
        # Thay thế ký tự không hợp lệ trong tên file
        to_safe = to_value.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
        ngay_safe = ngay_hien_tai.replace('/', '_')
        
        # File 1: Kiemke_dd/mm/yyyy_Tổ.xlsx (file chính)
        ten_file_1 = f"Kiemke_{ngay_safe}_{to_safe}.xlsx" if to_safe else f"Kiemke_{ngay_safe}.xlsx"
        
        # File 2: Kiemkecuoinam_dd/mm/yyyy_Tổ.xlsx (file ngầm)
        ten_file_2 = f"Kiemkecuoinam_{ngay_safe}_{to_safe}.xlsx" if to_safe else f"Kiemkecuoinam_{ngay_safe}.xlsx"
        
        # Kiểm tra có cấu hình đầy đủ không
        if not self.template_file_path or not os.path.exists(self.template_file_path):
            messagebox.showerror("Lỗi", "Chưa cấu hình CHỌN ĐƯỜNG DẪN FILE TỔNG HỢP MẶC ĐỊNH! Vui lòng cấu hình lại.")
            return
        
        # Cho phép người dùng chọn đường dẫn để lưu file
        # KHÔNG set initialdir để không hiển thị đường dẫn đã chọn trước đó
        # Mở dialog ở thư mục mặc định của hệ thống (Documents hoặc Desktop)
        user_home = Path.home()
        if sys.platform == 'win32':
            # Windows: mở ở Documents (không phải đường dẫn đã chọn trước đó)
            default_dir = str(user_home / "Documents")
        else:
            # macOS/Linux: mở ở Home
            default_dir = str(user_home)
        
        filename = filedialog.asksaveasfilename(
            title="Chọn đường dẫn để lưu file Excel tổng hợp",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=ten_file_1,
            initialdir=default_dir  # Luôn mở ở thư mục mặc định, không phải đường dẫn cũ
        )
        
        if not filename:
            # Người dùng đã hủy
            return
        
        # Lưu file vào đường dẫn đã chọn
        try:
            # QUAN TRỌNG: Lưu backup trước khi save để tránh mất dữ liệu nếu có lỗi
            try:
                self.save_backup()
            except Exception as backup_err:
                # Log lỗi nhưng không chặn quá trình save
                print(f"Lỗi khi lưu backup trước khi save: {str(backup_err)}")
            
            # Normalize path cho Windows
            filename = str(Path(filename).resolve())
            template_path_normalized = str(Path(self.template_file_path).resolve())
            
            # File 1: Chỉ copy file template từ đường dẫn đã cấu hình và đổi tên (KHÔNG ghi đè data)
            # Windows-specific: Retry nếu file bị lock
            max_retries = 5
            retry_count = 0
            copy_success = False
            
            while retry_count < max_retries and not copy_success:
                try:
                    shutil.copy2(template_path_normalized, filename)
                    # Cập nhật metadata file thành ngày hiện tại (để hiển thị ngày tải về)
                    self.update_excel_file_metadata(filename)
                    copy_success = True
                except PermissionError as pe:
                    retry_count += 1
                    if retry_count < max_retries:
                        time.sleep(0.5)  # Đợi 0.5 giây trước khi thử lại
                    else:
                        messagebox.showerror("Lỗi", 
                            f"Không thể copy file template!\n\n"
                            f"File có thể đang được mở trong Excel hoặc chương trình khác.\n"
                            f"Vui lòng đóng file và thử lại.\n\n"
                            f"Lỗi: {str(pe)}")
                        return
                except Exception as e:
                    messagebox.showerror("Lỗi", f"Không thể copy file template: {str(e)}")
                    traceback.print_exc()
                    return
            
            # KHÔNG ghi đè dữ liệu vào file 1 - giữ nguyên data từ template
            
            # File 2: Tự động lưu vào thư mục đã cấu hình (nếu có)
            if self.auto_save_folder:
                try:
                    # Tạo đường dẫn file 2 với tên file đúng format - Windows-safe
                    file2_path = str(Path(self.auto_save_folder) / ten_file_2)
                    
                    # Windows-specific: Retry nếu file bị lock
                    max_retries = 5
                    retry_count = 0
                    save_success = False
                    
                    # Với dữ liệu lớn (>5000 dòng), hiển thị progress bar
                    total_rows = len(df_save)
                    show_progress = total_rows > 5000
                    progress_window = None
                    progress_bar = None
                    
                    if show_progress:
                        # Tạo progress window
                        progress_window = tk.Toplevel(self.root)
                        progress_window.title("Đang lưu file...")
                        progress_window.geometry("400x100")
                        progress_window.resizable(False, False)
                        progress_window.transient(self.root)
                        progress_window.grab_set()
                        progress_window.configure(bg='#f5f5f5')
                        
                        # Đặt ở giữa màn hình
                        progress_window.update_idletasks()
                        x = (progress_window.winfo_screenwidth() // 2) - (400 // 2)
                        y = (progress_window.winfo_screenheight() // 2) - (100 // 2)
                        progress_window.geometry(f"400x100+{x}+{y}")
                        
                        # Label
                        label = tk.Label(progress_window, text=f"Đang lưu {total_rows:,} dòng dữ liệu...", 
                                       font=('Arial', 11), bg='#f5f5f5')
                        label.pack(pady=10)
                        
                        # Progress bar
                        progress_bar = ttk.Progressbar(progress_window, mode='indeterminate', length=350)
                        progress_bar.pack(pady=5)
                        progress_bar.start()
                        
                        progress_window.update()
                    
                    while retry_count < max_retries and not save_success:
                        temp_file = None
                        try:
                            # Lưu vào file tạm trước (atomic operation để tránh mất dữ liệu nếu crash)
                            # Tạo temp file với extension .xlsx để engine openpyxl có thể xử lý
                            file2_path_obj = Path(file2_path)
                            temp_file = file2_path_obj.parent / (file2_path_obj.stem + '_temp.xlsx')
                            
                            # Xóa temp file cũ nếu tồn tại
                            if temp_file.exists():
                                try:
                                    temp_file.unlink()
                                except:
                                    pass
                            
                            # Lưu file tạm với data tổng hợp
                            df_save.to_excel(str(temp_file), index=False, engine='openpyxl')
                            
                            # Nếu file cũ tồn tại, xóa nó trước
                            if file2_path_obj.exists():
                                try:
                                    file2_path_obj.unlink()
                                except:
                                    pass
                            
                            # Rename file tạm thành file chính (atomic operation)
                            temp_file.rename(file2_path)
                            
                            # Cập nhật metadata file thành ngày hiện tại (để hiển thị ngày tải về)
                            self.update_excel_file_metadata(file2_path)
                            
                            save_success = True
                        except PermissionError as pe:
                            retry_count += 1
                            if retry_count < max_retries:
                                time.sleep(0.5)  # Đợi 0.5 giây trước khi thử lại
                            else:
                                # Log lỗi nhưng không hiển thị cho người dùng
                                print(f"Lỗi khi lưu file tự động (file có thể đang mở): {str(pe)}")
                                traceback.print_exc()
                                break
                        except Exception as e2:
                            # Log lỗi nhưng không hiển thị cho người dùng
                            print(f"Lỗi khi lưu file tự động: {str(e2)}")
                            traceback.print_exc()
                            break
                        finally:
                            # Xóa file tạm nếu còn tồn tại (nếu có lỗi hoặc chưa rename thành công)
                            if temp_file is not None and temp_file.exists():
                                try:
                                    temp_file.unlink()
                                except:
                                    pass
                    
                    # Đóng progress window nếu có
                    if progress_window:
                        try:
                            progress_bar.stop()
                            progress_window.destroy()
                        except:
                            pass
                    
                    # Hiển thị thông báo thành công
                    messagebox.showinfo("Thành công", f"Đã lưu file tổng hợp thành công!")
                except Exception as e2:
                    # Nếu lỗi khi lưu file 2, chỉ log lỗi nhưng không hiển thị cho người dùng
                    print(f"Lỗi khi lưu file tự động: {str(e2)}")
                    traceback.print_exc()
                    # Vẫn hiển thị thành công cho file 1
                    messagebox.showinfo("Thành công", f"Đã lưu file tổng hợp thành công!")
            else:
                # Không có thư mục tự động, chỉ lưu file 1
                messagebox.showinfo("Thành công", f"Đã lưu file tổng hợp thành công!")
                
        except Exception as e:
            error_msg = str(e)
            traceback.print_exc()
            messagebox.showerror("Lỗi", f"Không thể lưu file: {error_msg}")
    
    def _get_backup_file_path_init(self):
        """Lấy đường dẫn file backup khi khởi tạo - chỉ gọi một lần"""
        if getattr(sys, 'frozen', False):
            # Chạy từ executable - lưu cùng thư mục với file .exe
            return Path(sys.executable).parent / "kiem_kho_showroom_backup.json"
        else:
            # Chạy từ source code - lưu cùng thư mục với file .py
            try:
                return Path(__file__).parent / "kiem_kho_showroom_backup.json"
            except NameError:
                # Nếu __file__ không có (ví dụ khi chạy từ interactive shell)
                # Sử dụng thư mục hiện tại hoặc thư mục làm việc
                return Path.cwd() / "kiem_kho_showroom_backup.json"
    
    def get_backup_file_path(self):
        """Lấy đường dẫn file backup - sử dụng đường dẫn đã lưu khi khởi tạo"""
        # Sử dụng đường dẫn đã lưu khi khởi tạo để tránh lỗi __file__ không được định nghĩa
        # (ví dụ khi chạy từ atexit handler)
        if hasattr(self, '_backup_file_path'):
            return self._backup_file_path
        else:
            # Fallback nếu chưa có (không nên xảy ra)
            return Path.cwd() / "kiem_kho_showroom_backup.json"
    
    def save_backup(self):
        """Tự động lưu backup dữ liệu (scanned_items và tong_hop_data) - tối ưu cho dữ liệu lớn"""
        try:
            backup_file = self.get_backup_file_path()
            backup_data = {
                'scanned_items': self.scanned_items,
                'tong_hop_data': self.tong_hop_data,
                'current_box_number': self.current_box_number,
                'timestamp': time.time()
            }
            
            # Kiểm tra kích thước dữ liệu
            data_size = len(self.tong_hop_data)
            
            # Với dữ liệu lớn (>10000 dòng), tối ưu cách lưu
            if data_size > 10000:
                # Giảm indent để file nhỏ hơn và nhanh hơn
                indent_value = None  # Không indent cho dữ liệu lớn để tăng tốc
            else:
                indent_value = 2  # Indent bình thường cho dữ liệu nhỏ
            
            # Lưu vào file tạm trước, sau đó rename để tránh mất dữ liệu khi crash
            temp_file = backup_file.with_suffix('.tmp')
            
            # Với dữ liệu cực lớn, sử dụng buffering lớn hơn
            if data_size > 50000:
                # Sử dụng buffer lớn hơn cho dữ liệu cực lớn
                buffer_size = 65536  # 64KB buffer
            else:
                buffer_size = 8192   # 8KB buffer mặc định
            
            try:
                with open(temp_file, 'w', encoding='utf-8', buffering=buffer_size) as f:
                    json.dump(backup_data, f, ensure_ascii=False, indent=indent_value)
            except MemoryError:
                # Nếu không đủ memory, thử lưu không indent
                print(f"Cảnh báo: Không đủ bộ nhớ, đang lưu backup không indent...")
                with open(temp_file, 'w', encoding='utf-8', buffering=buffer_size) as f:
                    json.dump(backup_data, f, ensure_ascii=False, indent=None)
            except Exception as e:
                # Thử lại với cách đơn giản hơn
                print(f"Lỗi khi lưu backup, thử lại không indent: {str(e)}")
                with open(temp_file, 'w', encoding='utf-8') as f:
                    json.dump(backup_data, f, ensure_ascii=False, indent=None)
            
            # Rename file tạm thành file chính (atomic operation)
            if backup_file.exists():
                try:
                    backup_file.unlink()
                except Exception as e:
                    # Nếu không xóa được file cũ, thử đổi tên file cũ
                    print(f"Không thể xóa file backup cũ: {str(e)}")
                    old_backup = backup_file.with_suffix('.old')
                    try:
                        if old_backup.exists():
                            old_backup.unlink()
                        backup_file.rename(old_backup)
                    except:
                        pass  # Bỏ qua nếu vẫn lỗi
            
            temp_file.rename(backup_file)
            
        except MemoryError:
            # Xử lý lỗi memory riêng
            print(f"Lỗi: Không đủ bộ nhớ để lưu backup ({len(self.tong_hop_data)} dòng)")
        except Exception as e:
            # Không hiển thị lỗi cho người dùng vì đây là auto-save
            print(f"Lỗi khi lưu backup: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def check_and_restore_backup(self):
        """Kiểm tra và khôi phục dữ liệu backup nếu có"""
        try:
            backup_file = self.get_backup_file_path()
            if not backup_file.exists():
                return  # Không có backup
            
            # Đọc backup - tối ưu cho dữ liệu lớn
            # Kiểm tra kích thước file trước
            file_size = backup_file.stat().st_size
            
            # Với file lớn (>10MB), sử dụng buffer lớn hơn
            if file_size > 10 * 1024 * 1024:  # >10MB
                buffer_size = 65536  # 64KB buffer
            else:
                buffer_size = 8192   # 8KB buffer mặc định
            
            try:
                with open(backup_file, 'r', encoding='utf-8', buffering=buffer_size) as f:
                    backup_data = json.load(f)
            except MemoryError:
                messagebox.showerror("Lỗi", 
                    f"Không đủ bộ nhớ để khôi phục backup!\n\n"
                    f"Kích thước file: {file_size / (1024*1024):.2f} MB\n\n"
                    f"Vui lòng đóng các ứng dụng khác và thử lại.")
                return
            except json.JSONDecodeError as e:
                messagebox.showerror("Lỗi", 
                    f"File backup bị lỗi hoặc không hợp lệ!\n\n"
                    f"Lỗi: {str(e)}\n\n"
                    f"Vui lòng kiểm tra lại file backup.")
                return
            
            scanned_items_backup = backup_data.get('scanned_items', {})
            tong_hop_data_backup = backup_data.get('tong_hop_data', [])
            current_box_number_backup = backup_data.get('current_box_number')
            timestamp = backup_data.get('timestamp', 0)
            
            # Kiểm tra xem có dữ liệu để khôi phục không
            has_scanned_items = scanned_items_backup and len(scanned_items_backup) > 0
            has_tong_hop_data = tong_hop_data_backup and len(tong_hop_data_backup) > 0
            
            if not has_scanned_items and not has_tong_hop_data:
                return  # Không có dữ liệu để khôi phục
            
            # Hiển thị dialog cho phép người dùng chọn khôi phục
            dialog = tk.Toplevel(self.root)
            dialog.title("Khôi phục dữ liệu")
            dialog.geometry("600x400")
            dialog.resizable(False, False)
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.configure(bg='#f5f5f5')
            
            # Đặt dialog ở phía trên cửa sổ chính để không che các nút SAVE/RESET
            dialog.update_idletasks()
            # Lấy vị trí của cửa sổ chính
            root_x = self.root.winfo_x()
            root_y = self.root.winfo_y()
            root_width = self.root.winfo_width()
            # Đặt dialog ở giữa theo chiều ngang của cửa sổ chính, nhưng ở phía trên
            x = root_x + (root_width // 2) - (600 // 2)
            y = root_y + 100  # Đặt cách đỉnh cửa sổ chính 100px để không che các nút
            dialog.geometry(f"600x400+{x}+{y}")
            
            # Frame chính với padding đủ để các nút không bị che
            main_frame = tk.Frame(dialog, padx=30, pady=30, bg='#f5f5f5')
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Icon thông tin
            icon_frame = tk.Frame(main_frame, bg='#f5f5f5')
            icon_frame.pack(pady=(0, 15))
            icon_label = tk.Label(icon_frame, text="💾", font=('Arial', 32), bg='#f5f5f5')
            icon_label.pack()
            
            # Thông điệp
            from datetime import datetime
            backup_time = datetime.fromtimestamp(timestamp).strftime("%d/%m/%Y %H:%M:%S") if timestamp else "Không xác định"
            
            message_text = (
                f"Phát hiện dữ liệu backup từ lần chạy trước!\n\n"
                f"Thời gian backup: {backup_time}\n\n"
            )
            
            if has_scanned_items:
                message_text += f"• Dữ liệu đang quét: {len(scanned_items_backup)} tựa\n"
            if has_tong_hop_data:
                message_text += f"• Dữ liệu tổng hợp: {len(tong_hop_data_backup)} dòng\n"
            if current_box_number_backup:
                message_text += f"• Thùng đang kiểm kê: {current_box_number_backup}\n"
            
            message_text += "\nBạn có muốn khôi phục dữ liệu này không?"
            
            message_frame = tk.Frame(main_frame, bg='#f5f5f5')
            message_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
            message_label = tk.Label(
                message_frame, 
                text=message_text,
                font=('Arial', 12),
                justify=tk.CENTER,
                bg='#f5f5f5',
                fg='#333333',
                wraplength=540
            )
            message_label.pack()
            
            # Frame chứa các nút với padding đủ để không bị che
            button_frame = tk.Frame(main_frame, bg='#f5f5f5')
            button_frame.pack(pady=(15, 25))
            
            result = {'value': None}
            
            def on_restore():
                result['value'] = 'restore'
                dialog.destroy()
            
            def on_discard():
                result['value'] = 'discard'
                dialog.destroy()
            
            # Nút "Khôi phục"
            restore_btn = tk.Button(
                button_frame,
                text="Khôi phục",
                command=on_restore,
                font=('Arial', 12, 'bold'),
                bg='#4CAF50',
                fg='white',
                padx=25,
                pady=12,
                relief=tk.FLAT,
                cursor='hand2',
                activebackground='#45a049',
                activeforeground='white',
                bd=0,
                highlightthickness=0
            )
            restore_btn.pack(side=tk.LEFT, padx=15)
            
            # Nút "Bỏ qua"
            discard_btn = tk.Button(
                button_frame,
                text="Bỏ qua",
                command=on_discard,
                font=('Arial', 12, 'bold'),
                bg='#757575',
                fg='white',
                padx=25,
                pady=12,
                relief=tk.FLAT,
                cursor='hand2',
                activebackground='#616161',
                activeforeground='white',
                bd=0,
                highlightthickness=0
            )
            discard_btn.pack(side=tk.LEFT, padx=15)
            
            # Đợi dialog đóng
            dialog.wait_window()
            
            # Xử lý kết quả
            if result['value'] == 'restore':
                # Khôi phục dữ liệu
                self.scanned_items = scanned_items_backup
                self.tong_hop_data = tong_hop_data_backup
                self.current_box_number = current_box_number_backup
                
                # QUAN TRỌNG: Cập nhật số thùng vào input field TRƯỚC các thao tác khác
                # Để đảm bảo update_da_quet_counter() có thể đọc đúng số thùng
                if hasattr(self, 'so_thung_var') and self.current_box_number:
                    self.so_thung_var.set(str(self.current_box_number).strip())
                else:
                    # Nếu không có current_box_number nhưng có tong_hop_data, thử lấy số thùng từ dữ liệu đầu tiên
                    if self.tong_hop_data and len(self.tong_hop_data) > 0:
                        first_data = self.tong_hop_data[0]
                        so_thung_from_data = str(first_data.get('Số thùng', '')).strip()
                        if so_thung_from_data and hasattr(self, 'so_thung_var'):
                            self.so_thung_var.set(so_thung_from_data)
                            self.current_box_number = so_thung_from_data
                
                # Đảm bảo Tkinter variable được sync
                self.root.update_idletasks()
                
                # Cập nhật UI sau khi khôi phục
                if hasattr(self, 'tong_hop_tree') and self.tong_hop_tree:
                    self.update_tong_hop_table()
                
                # Xóa tất cả items hiện tại trong tree (không gọi clear_table để tránh gọi update_da_quet_counter nhiều lần)
                for item in self.tree.get_children():
                    self.tree.delete(item)
                
                # Nếu có dữ liệu đang quét, hiển thị lại trong bảng
                if self.scanned_items:
                    for isbn, info in self.scanned_items.items():
                        # Tạo lại item trong tree
                        item_id = self.tree.insert('', tk.END, values=(
                            len(self.tree.get_children()) + 1,  # STT
                            isbn,
                            info.get('tua', ''),
                            info.get('ton_thuc_te', ''),
                            info.get('so_thung', ''),
                            info.get('ton_trong_thung', ''),
                            info.get('tinh_trang', ''),
                            info.get('ghi_chu', ''),
                            'Xóa'  # Nút xóa
                        ))
                        # Cập nhật item_id trong scanned_items
                        info['item_id'] = item_id
                    
                    # Re-index STT sau khi insert tất cả items
                    children = list(self.tree.get_children())
                    for idx, child in enumerate(children, start=1):
                        child_values = list(self.tree.item(child, 'values'))
                        if len(child_values) >= 1:
                            child_values[0] = str(idx)
                            self.tree.item(child, values=child_values)
                
                # Đảm bảo UI được cập nhật trước khi đếm
                self.root.update_idletasks()
                
                # QUAN TRỌNG: Luôn cập nhật "Đã quét" sau khi khôi phục (cả khi có hoặc không có scanned_items)
                # Để hiển thị số dòng trong Tổng hợp cho cùng Số thùng + số dòng trong Kiểm kê
                # Đảm bảo cập nhật sau khi:
                # 1. tong_hop_data đã được khôi phục
                # 2. so_thung_var đã được set
                # 3. Tất cả items đã được insert vào tree
                self.update_da_quet_counter()
                
                messagebox.showinfo("Thành công", 
                    f"Đã khôi phục dữ liệu!\n\n"
                    f"Dữ liệu đang quét: {len(self.scanned_items)} tựa\n"
                    f"Dữ liệu tổng hợp: {len(self.tong_hop_data)} dòng")
            else:
                # Người dùng không muốn khôi phục - giữ nguyên file backup (không xóa)
                pass
        
        except Exception as e:
            # Không hiển thị lỗi cho người dùng, chỉ log
            print(f"Lỗi khi khôi phục backup: {str(e)}")
    
    def start_auto_save(self):
        """Bắt đầu auto-save định kỳ (mỗi 30 giây)"""
        def auto_save_periodic():
            # Chỉ lưu nếu có dữ liệu
            if self.scanned_items or self.tong_hop_data:
                self.save_backup()
            # Lên lịch lại sau 30 giây
            self.root.after(30000, auto_save_periodic)
        
        # Bắt đầu auto-save ngay lập tức, sau đó mỗi 30 giây
        auto_save_periodic()  # Chạy ngay lần đầu
    
    def save_backup_on_change(self):
        """Lưu backup ngay lập tức khi có thay đổi dữ liệu"""
        # Delay một chút để tránh lưu quá nhiều lần
        if hasattr(self, '_backup_scheduled') and self._backup_scheduled is not None:
            try:
                # Chỉ cancel nếu ID hợp lệ
                if isinstance(self._backup_scheduled, str) and self._backup_scheduled:
                    self.root.after_cancel(self._backup_scheduled)
            except (ValueError, TypeError, tk.TclError):
                # Bỏ qua lỗi nếu ID không hợp lệ
                pass
            self._backup_scheduled = None
        
        def do_save():
            try:
                self.save_backup()
            except Exception as e:
                # Không hiển thị lỗi cho người dùng
                print(f"Error saving backup: {e}")
            finally:
                self._backup_scheduled = None
        
        try:
            after_id = self.root.after(2000, do_save)  # Lưu sau 2 giây
            # Chỉ lưu nếu after() trả về giá trị hợp lệ
            if after_id is not None:
                self._backup_scheduled = after_id
            else:
                # Nếu after() trả về None, lưu ngay
                self.save_backup()
        except Exception as e:
            # Nếu có lỗi với after(), lưu ngay
            print(f"Error scheduling backup: {e}")
            try:
                self.save_backup()
            except:
                pass
    
    def setup_signal_handlers(self):
        """Đăng ký xử lý signal để lưu backup khi shutdown (cúp điện, tắt máy)"""
        def signal_handler(signum, frame):
            """Xử lý signal shutdown - lưu backup ngay lập tức"""
            try:
                # Hủy scheduled backup nếu có để tránh conflict
                if hasattr(self, '_backup_scheduled') and self._backup_scheduled:
                    try:
                        self.root.after_cancel(self._backup_scheduled)
                    except:
                        pass
                
                # Lưu backup ngay lập tức
                if self.scanned_items or self.tong_hop_data:
                    self.save_backup()
                    print(f"Đã lưu backup khi nhận signal {signum}")
            except Exception as e:
                print(f"Lỗi khi lưu backup trong signal handler: {str(e)}")
        
        def atexit_handler():
            """Xử lý khi exit - lưu backup"""
            try:
                if self.scanned_items or self.tong_hop_data:
                    self.save_backup()
                    print("Đã lưu backup khi exit")
            except Exception as e:
                print(f"Lỗi khi lưu backup trong atexit handler: {str(e)}")
        
        # Đăng ký signal handlers (chỉ trên Unix/Linux/macOS, Windows không hỗ trợ tốt)
        if hasattr(signal, 'SIGTERM'):
            try:
                signal.signal(signal.SIGTERM, signal_handler)
            except:
                pass
        
        if hasattr(signal, 'SIGINT'):
            try:
                signal.signal(signal.SIGINT, signal_handler)
            except:
                pass
        
        # Đăng ký atexit handler (hoạt động trên cả Windows và Unix)
        atexit.register(atexit_handler)
    
    def on_closing(self):
        """Xử lý sự kiện đóng cửa sổ - kiểm tra dữ liệu chưa lưu"""
        # Kiểm tra xem có dữ liệu chưa lưu không (cả scanned_items và tong_hop_data)
        has_scanned_items = self.scanned_items and len(self.scanned_items) > 0
        has_tong_hop_data = self.tong_hop_data and len(self.tong_hop_data) > 0
        
        if has_scanned_items or has_tong_hop_data:
            # Có dữ liệu chưa lưu, hiển thị dialog cảnh báo
            dialog = tk.Toplevel(self.root)
            dialog.title("Cảnh báo")
            dialog.geometry("600x320")
            dialog.resizable(False, False)
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.configure(bg='#f5f5f5')
            
            # Đặt dialog ở giữa màn hình
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (600 // 2)
            y = (dialog.winfo_screenheight() // 2) - (320 // 2)
            dialog.geometry(f"600x320+{x}+{y}")
            
            # Frame chính với padding lớn hơn
            main_frame = tk.Frame(dialog, padx=30, pady=25, bg='#f5f5f5')
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Icon cảnh báo
            icon_frame = tk.Frame(main_frame, bg='#f5f5f5')
            icon_frame.pack(pady=(0, 15))
            icon_label = tk.Label(icon_frame, text="⚠️", font=('Arial', 32), bg='#f5f5f5')
            icon_label.pack()
            
            # Thông điệp
            message_parts = []
            if has_scanned_items:
                message_parts.append(f"• {len(self.scanned_items)} tựa đang quét chưa được lưu vào Tổng hợp")
            if has_tong_hop_data:
                message_parts.append(f"• {len(self.tong_hop_data)} dòng dữ liệu trong tab Tổng hợp chưa được lưu backup")
            
            message_text = (
                f"Bạn có muốn lưu backup dữ liệu trước khi đóng phần mềm không?\n\n"
                + "\n".join(message_parts) + "\n\n"
                f"Bạn muốn làm gì?"
            )
            
            message_frame = tk.Frame(main_frame, bg='#f5f5f5')
            message_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 25))
            message_label = tk.Label(
                message_frame,
                text=message_text,
                font=('Arial', 12),
                justify=tk.CENTER,
                bg='#f5f5f5',
                fg='#333333',
                wraplength=540
            )
            message_label.pack()
            
            # Frame chứa các nút
            button_frame = tk.Frame(main_frame, bg='#f5f5f5')
            button_frame.pack(pady=(10, 0))
            
            result = {'value': None}
            
            def on_save():
                result['value'] = 'save'
                dialog.destroy()
            
            def on_close():
                result['value'] = 'close'
                dialog.destroy()
            
            # Nút "Lưu"
            save_btn = tk.Button(
                button_frame,
                text="Lưu",
                command=on_save,
                font=('Arial', 12, 'bold'),
                bg='#4CAF50',
                fg='white',
                padx=25,
                pady=12,
                relief=tk.FLAT,
                cursor='hand2',
                activebackground='#45a049',
                activeforeground='white',
                bd=0,
                highlightthickness=0
            )
            save_btn.pack(side=tk.LEFT, padx=15)
            
            # Nút "Đóng không lưu"
            close_btn = tk.Button(
                button_frame,
                text="Đóng không lưu",
                command=on_close,
                font=('Arial', 12, 'bold'),
                bg='#757575',
                fg='white',
                padx=25,
                pady=12,
                relief=tk.FLAT,
                cursor='hand2',
                activebackground='#616161',
                activeforeground='white',
                bd=0,
                highlightthickness=0
            )
            close_btn.pack(side=tk.LEFT, padx=15)
            
            # Đợi dialog đóng
            dialog.wait_window()
            
            # Xử lý kết quả
            if result['value'] == 'save':
                # Lưu vào backup file trước khi đóng
                try:
                    self.save_backup()
                    # Sau khi lưu backup xong, đóng phần mềm
                    self.root.quit()
                    self.root.destroy()
                except Exception as e:
                    # Nếu lưu backup lỗi, hỏi lại có muốn đóng không
                    error_result = messagebox.askyesno(
                        "Lỗi",
                        f"Không thể lưu backup: {str(e)}\n\n"
                        f"Bạn có muốn đóng phần mềm mà không lưu không?"
                    )
                    if error_result:
                        self.root.quit()
                        self.root.destroy()
            elif result['value'] == 'close':
                # Đóng phần mềm luôn, không lưu gì cả
                self.root.quit()
                self.root.destroy()
            # Nếu result['value'] là None (người dùng đóng dialog bằng X), không làm gì cả
        else:
            # Không có dữ liệu chưa lưu, đóng phần mềm bình thường
            self.root.quit()
            self.root.destroy()

def main():
    try:
        root = tk.Tk()
        # Hiển thị window ngay lập tức để tăng tốc độ khởi động (perceived speed)
        root.deiconify()
        root.update_idletasks()  # Cập nhật UI ngay lập tức
        
        # Tạo app (sẽ load dữ liệu sau khi UI hiển thị)
        app = KiemKhoApp(root)
        
        # Kiểm tra xem app đã được khởi tạo thành công chưa
        if not hasattr(app, 'template_file_path') or not app.template_file_path:
            print("App không được khởi tạo đúng cách")
            root.quit()
            return
        
        root.mainloop()
    except Exception as e:
        # Hiển thị lỗi nếu có
        import traceback
        error_msg = f"Lỗi khi khởi động ứng dụng:\n{str(e)}\n\n{traceback.format_exc()}"
        print(error_msg)
        try:
            # Tạo một root window mới để hiển thị lỗi
            error_root = tk.Tk()
            error_root.withdraw()  # Ẩn window chính
            messagebox.showerror("Lỗi", f"Lỗi khi khởi động ứng dụng:\n{str(e)}")
            error_root.destroy()
        except:
            # Nếu không thể hiển thị messagebox, in ra console
            print("Không thể hiển thị dialog lỗi")
            print(error_msg)

if __name__ == "__main__":
    main()

