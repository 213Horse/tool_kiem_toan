#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script kiểm tra và sửa file Excel
"""

import pandas as pd
from pathlib import Path
import sys

def check_excel_file(filepath):
    """Kiểm tra file Excel"""
    print(f"\n=== Kiểm tra file: {filepath} ===")
    
    if not Path(filepath).exists():
        print(f"File không tồn tại!")
        return False
    
    try:
        if filepath.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True)
            print(f"Số worksheet: {len(wb.sheetnames)}")
            if len(wb.sheetnames) == 0:
                print("⚠ File không có worksheet nào!")
                wb.close()
                return False
            print(f"Worksheet names: {wb.sheetnames}")
            ws = wb[wb.sheetnames[0]]
            print(f"Số dòng: {ws.max_row}, Số cột: {ws.max_column}")
            wb.close()
            
            # Thử đọc với pandas
            df = pd.read_excel(filepath, engine='openpyxl', header=0, nrows=10)
            print(f"\nĐọc được {len(df)} dòng với header=0")
            print(f"Các cột: {list(df.columns)[:10]}")
            
        elif filepath.endswith('.xls'):
            df = pd.read_excel(filepath, engine='xlrd', header=None, nrows=20)
            print(f"Đọc được {len(df)} dòng (không có header)")
            print("Tìm header row...")
            
            # Tìm header
            for i in range(min(20, len(df))):
                row_values = [str(c).lower() if pd.notna(c) else '' for c in df.iloc[i].values]
                row_str = ' '.join(row_values)
                if 'isbn' in row_str:
                    keywords = sum([
                        'isbn' in row_str,
                        'số thùng' in row_str or 'so thung' in row_str or 'thùng' in row_str,
                        'tựa' in row_str or 'tua' in row_str or 'titles' in row_str,
                        'tồn' in row_str or 'ton' in row_str or 'qty' in row_str
                    ])
                    if keywords >= 2:
                        print(f"✓ Tìm thấy header ở dòng {i+1}")
                        # Đọc với header này
                        df_header = pd.read_excel(filepath, engine='xlrd', header=i, nrows=50)
                        print(f"\nĐọc được {len(df_header)} dòng với header={i}")
                        print(f"Các cột: {list(df_header.columns)[:10]}")
                        
                        # Kiểm tra dữ liệu
                        if 'ISBN' in df_header.columns:
                            valid_isbn = df_header[df_header['ISBN'].notna()]
                            print(f"Số dòng có ISBN: {len(valid_isbn)}")
                            if len(valid_isbn) > 0:
                                print(f"\nMẫu dữ liệu:")
                                print(valid_isbn[['ISBN', 'Titles', 'Thùng']].head(5) if all(c in df_header.columns for c in ['ISBN', 'Titles', 'Thùng']) else valid_isbn.head(5))
                        break
            else:
                print("⚠ Không tìm thấy header hợp lệ")
                return False
        
        return True
        
    except Exception as e:
        print(f"✗ Lỗi: {e}")
        return False

if __name__ == "__main__":
    files = [
        'DuLieuDauVao.xlsx',
        'KIEM KE Năm -2025 - BP ONLINE.xls',
        'KIEM KE Năm -2025 - BP ONLINE copy.xls'
    ]
    
    print("=" * 60)
    print("KIỂM TRA FILE EXCEL")
    print("=" * 60)
    
    for filename in files:
        check_excel_file(filename)
    
    print("\n" + "=" * 60)
    print("KẾT LUẬN:")
    print("=" * 60)
    print("Nếu file DuLieuDauVao.xlsx bị hỏng, bạn có thể:")
    print("1. Mở file .xls trong Excel và lưu lại thành .xlsx")
    print("2. Hoặc copy dữ liệu từ file .xls sang file .xlsx mới")
    print("3. Đảm bảo file có các cột: Số thùng, ISBN, Tựa, Tồn từng tựa")


